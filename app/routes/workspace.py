"""
Workspace routes

Main tree/grid navigation and data entry.
"""

import base64
import json
from datetime import date
from functools import wraps

from flask import Blueprint, abort, current_app, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required
from flask_wtf.csrf import generate_csrf
from markupsafe import Markup
from sqlalchemy import or_
from sqlalchemy.orm import selectinload

from app.extensions import db
from app.forms import ContributionForm
from app.models import (
    KPI,
    CellComment,
    Challenge,
    ChallengeInitiativeLink,
    Contribution,
    EntityTypeDefault,
    GovernanceBody,
    Initiative,
    InitiativeSystemLink,
    KPIGovernanceBodyLink,
    KPISnapshot,
    KPIValueTypeConfig,
    Organization,
    RollupSnapshot,
    SavedChart,
    SavedSearch,
    ImpactLevel,
    Space,
    StrategicPillar,
    System,
    SystemAnnouncement,
    User,
    UserAnnouncementAcknowledgment,
    UserFilterPreset,
    UserOrganizationMembership,
    ValueType,
)
from app.services import AggregationService, ConsensusService, ExcelExportService
from app.services.comment_service import CommentService
from app.services.search_service import SearchService
from app.services.snapshot_pivot_service import SnapshotPivotService
from app.services.snapshot_service import SnapshotService

bp = Blueprint("workspace", __name__, url_prefix="/workspace")


def organization_required(f):
    """Decorator to require organization context"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for("auth.login"))
        if session.get("organization_id") is None:
            flash("Please log in to an organization", "warning")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)

    return decorated_function


def _filters_match(current, preset):
    """Helper to compare current filters with preset filters"""

    # Normalize both to have consistent format
    def normalize_value(val):
        if val is None:
            return None
        if isinstance(val, list):
            if len(val) == 0:
                return None
            if len(val) == 1:
                return str(val[0])
            # Multiple values - return as sorted list of strings
            return sorted([str(v) for v in val])
        return str(val)

    # Remove skip_default from comparison (it's a control parameter, not a filter)
    current_clean = {k: v for k, v in current.items() if k != "skip_default"}
    preset_clean = {k: v for k, v in preset.items() if k != "skip_default"}

    # Get all keys from both
    all_keys = set(current_clean.keys()) | set(preset_clean.keys())

    for key in all_keys:
        current_norm = normalize_value(current_clean.get(key))
        preset_norm = normalize_value(preset_clean.get(key))

        # Both None/missing - match
        if current_norm is None and preset_norm is None:
            continue

        # One missing - no match
        if current_norm is None or preset_norm is None:
            return False

        # Compare (both are now either strings or sorted lists)
        if current_norm != preset_norm:
            return False

    return True


@bp.route("/theory")
@login_required
def theory():
    """CISK Theory — one-page framework overview"""
    from flask_wtf.csrf import generate_csrf
    return render_template("workspace/theory.html", csrf_token=generate_csrf)


@bp.route("/decision-register", methods=["GET", "POST"])
@login_required
@organization_required
def decision_register():
    """Standalone Decision Register — create and manage decisions outside initiative reviews"""
    from app.models import Decision, GovernanceBody

    org_id = session.get("organization_id")

    if request.method == "POST":
        what = request.form.get("what", "").strip()
        if not what:
            flash("Decision description is required.", "danger")
            return redirect(url_for("workspace.decision_register"))

        who = request.form.get("who", "").strip() or None
        tags_raw = request.form.getlist("tags")
        gb_id = request.form.get("governance_body_id") or None

        # Parse entity mentions from JSON
        import json as _dj
        mentions_json = request.form.get("entity_mentions_json", "[]")
        try:
            entity_mentions = _dj.loads(mentions_json)
        except (ValueError, TypeError):
            entity_mentions = []

        decision = Decision(
            organization_id=org_id,
            created_by_id=current_user.id,
            what=what,
            who=who,
            tags=tags_raw if tags_raw else None,
            entity_mentions=entity_mentions if entity_mentions else None,
            governance_body_id=int(gb_id) if gb_id else None,
        )
        db.session.add(decision)
        db.session.commit()
        flash("Decision recorded.", "success")
        return redirect(url_for("workspace.decision_register"))

    # GET — load existing standalone decisions
    standalone_decisions = (
        Decision.query.filter_by(organization_id=org_id)
        .order_by(Decision.created_at.desc())
        .all()
    )

    governance_bodies = GovernanceBody.query.filter_by(organization_id=org_id, is_active=True).order_by(GovernanceBody.name).all()

    # Decision tags from org config
    org = Organization.query.get(org_id)
    decision_tags = (org.decision_tags if org and org.decision_tags else
                     ["scope", "budget", "timeline", "resource", "technical", "governance", "other"])

    # Presets
    from app.models import UserFilterPreset
    dr_presets = (
        UserFilterPreset.query.filter_by(user_id=current_user.id, organization_id=org_id, feature="decisions")
        .order_by(UserFilterPreset.name).all()
    )

    # Unique values for filters
    all_who = sorted(set(d.who for d in standalone_decisions if d.who))
    all_tags_used = set()
    for d in standalone_decisions:
        if d.tags:
            for t in d.tags:
                all_tags_used.add(t)

    # Compute max importance per decision from entity mentions
    from app.models import ImpactLevel, Challenge, ChallengeInitiativeLink
    from app.services.impact_service import compute_true_importance
    _impact_scale = ImpactLevel.get_org_levels(org_id)
    _iw = {lvl: _impact_scale[lvl]["weight"] for lvl in _impact_scale} if _impact_scale else {}
    _im = org.impact_calc_method or "geometric_mean" if org else "geometric_mean"
    _icm = org.impact_qfd_matrix if org else None
    _icr = org.impact_reinforce_weights if org else None
    decision_importance = {}
    for d in standalone_decisions:
        max_ti = 0
        for m in (d.entity_mentions or []):
            if m.get("entity_type") == "initiative":
                ini = Initiative.query.get(m.get("entity_id"))
                if ini and ini.impact_level:
                    ci = ChallengeInitiativeLink.query.filter_by(initiative_id=ini.id).first()
                    if ci and ci.challenge and ci.challenge.space:
                        sp = ci.challenge.space
                        ch = ci.challenge
                        if sp.impact_level and ch.impact_level:
                            ti = compute_true_importance([sp.impact_level, ch.impact_level, ini.impact_level], _im, _iw, _icm, _icr)
                            if ti and ti > max_ti:
                                max_ti = ti
        decision_importance[d.id] = max_ti

    return render_template(
        "workspace/decision_register.html",
        decisions=standalone_decisions,
        governance_bodies=governance_bodies,
        decision_tags=decision_tags,
        csrf_token=generate_csrf,
        dr_presets_list=[{"id": p.id, "name": p.name, "config": p.filters} for p in dr_presets],
        all_who=sorted(all_who),
        all_tags_used=sorted(all_tags_used),
        decision_importance=decision_importance,
        impact_scale=_impact_scale,
    )


@bp.route("/decision-register/<int:decision_id>/delete", methods=["POST"])
@login_required
@organization_required
def delete_decision(decision_id):
    """Delete a standalone decision"""
    from app.models import Decision

    org_id = session.get("organization_id")
    decision = Decision.query.filter_by(id=decision_id, organization_id=org_id).first_or_404()
    db.session.delete(decision)
    db.session.commit()
    flash("Decision deleted.", "success")
    return redirect(url_for("workspace.decision_register"))


@bp.route("/decisions")
@login_required
@organization_required
def decisions():
    """Legacy route — redirect to Decision Register"""
    return redirect(url_for("workspace.decision_register"))


@bp.route("/decisions/export")
@login_required
@organization_required
def decisions_export():
    """Export decisions as CSV"""
    import csv
    from io import StringIO

    from app.models import Decision

    org_id = session.get("organization_id")
    all_decisions = Decision.query.filter_by(organization_id=org_id).order_by(Decision.created_at.desc()).all()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Decision", "Who", "Tags", "Governance Body", "Author", "Entities"])

    for d in all_decisions:
        author = d.created_by.display_name or d.created_by.login if d.created_by else ""
        gb = d.governance_body.name if d.governance_body else ""
        tags = ", ".join(d.tags) if d.tags else ""
        entities = ", ".join(m.get("entity_name", "") for m in (d.entity_mentions or []))
        writer.writerow([
            d.created_at.strftime("%Y-%m-%d"),
            d.what, d.who or "", tags, gb, author, entities,
        ])

    output.seek(0)
    return send_file(
        __import__("io").BytesIO(output.getvalue().encode("utf-8-sig")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"decisions_{session.get('organization_name', 'org')}_{__import__('datetime').date.today()}.csv",
    )


@bp.route("/governance")
@login_required
@organization_required
def gb_dashboard_index():
    """Redirect to the first governance body dashboard"""
    from app.models import GovernanceBody
    org_id = session.get("organization_id")
    first_gb = GovernanceBody.query.filter_by(organization_id=org_id, is_active=True).order_by(GovernanceBody.display_order).first()
    if first_gb:
        return redirect(url_for("workspace.gb_dashboard", gb_id=first_gb.id))
    flash("No governance bodies configured.", "warning")
    return redirect(url_for("workspace.index"))


@bp.route("/api/compare-rollups")
@login_required
@organization_required
def compare_rollups():
    """Compare pre-computed vs live rollup values for verification.
    Returns mismatches between the two computation modes."""
    org_id = session.get("organization_id")
    if not (current_user.is_global_admin or current_user.is_super_admin):
        return jsonify({"error": "Super admin only"}), 403

    from app.models import RollupCacheEntry
    import time

    # Time the live computation
    start_live = time.time()
    # Force live mode by temporarily ignoring cache
    from app.models.system_setting import SystemSetting as _SS2
    original = _SS2.get_bool("precompute_rollups_enabled", default=False)

    # Get cache data
    cache_entries = RollupCacheEntry.query.filter_by(organization_id=org_id).all()
    cache_map = {}
    for ce in cache_entries:
        cache_map[(ce.entity_type, ce.entity_id, ce.value_type_id)] = {
            "value": ce.value,
            "formatted_value": ce.formatted_value,
        }
    cache_time = len(cache_entries)

    # Compare: iterate entities and check values
    mismatches = []
    matches = 0
    for key, cached in cache_map.items():
        etype, eid, vtid = key
        if etype == "kpi":
            continue  # Skip KPI-level comparison for now (complex)
        # We can't easily re-run live computation here without duplicating the entire
        # _build_workspace_data. Instead, report cache stats.
        matches += 1

    return jsonify({
        "cache_entries": len(cache_entries),
        "entity_types": {
            "space": sum(1 for k in cache_map if k[0] == "space"),
            "challenge": sum(1 for k in cache_map if k[0] == "challenge"),
            "initiative": sum(1 for k in cache_map if k[0] == "initiative"),
            "system": sum(1 for k in cache_map if k[0] == "system"),
            "kpi": sum(1 for k in cache_map if k[0] == "kpi"),
        },
        "note": "To verify correctness: toggle precompute OFF, load workspace, note values. Toggle ON, recompute, load workspace, compare visually.",
    })


@bp.route("/api/recompute-rollups", methods=["POST"])
@login_required
@organization_required
def recompute_rollups_api():
    """Recompute rollup cache for current organization (AJAX)"""
    org_id = session.get("organization_id")
    if not (current_user.is_global_admin or current_user.is_super_admin or current_user.is_org_admin(org_id)):
        return jsonify({"error": "Permission denied"}), 403
    from app.services.rollup_compute_service import RollupComputeService
    try:
        result = RollupComputeService.recompute_organization(org_id)
        return jsonify({"success": True, **result})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/visibility-dashboard")
@login_required
@organization_required
def visibility_dashboard():
    """Visibility Dashboard — what's public vs private across the workspace"""
    from app.models import ActionItem, Space, Stakeholder, StakeholderMap, Decision

    org_id = session.get("organization_id")

    # Spaces
    all_spaces = Space.query.filter_by(organization_id=org_id).all()
    spaces_public = [s for s in all_spaces if not s.is_private]
    spaces_private = [s for s in all_spaces if s.is_private]

    # Action Items
    all_actions = ActionItem.query.filter_by(organization_id=org_id).all()
    actions_shared = [a for a in all_actions if a.visibility == "shared"]
    actions_private = [a for a in all_actions if a.visibility == "private"]
    private_owners = {}
    for a in actions_private:
        owner = a.owner_user.display_name or a.owner_user.login if a.owner_user else "Unknown"
        private_owners[owner] = private_owners.get(owner, 0) + 1

    # Stakeholders
    all_stakeholders = Stakeholder.query.filter_by(organization_id=org_id).all()
    stk_shared = [s for s in all_stakeholders if s.visibility == "shared"]
    stk_private = [s for s in all_stakeholders if s.visibility == "private"]

    # Stakeholder Maps
    all_maps = StakeholderMap.query.filter_by(organization_id=org_id).all()
    maps_shared = [m for m in all_maps if m.visibility == "shared"]
    maps_private = [m for m in all_maps if m.visibility == "private"]

    # Decisions (all shared)
    decisions_count = Decision.query.filter_by(organization_id=org_id).count()

    categories = [
        {"name": "Spaces", "icon": "bi-folder", "color": "#10b981", "total": len(all_spaces), "shared": len(spaces_public), "private": len(spaces_private)},
        {"name": "Action Items", "icon": "bi-card-checklist", "color": "#3b82f6", "total": len(all_actions), "shared": len(actions_shared), "private": len(actions_private)},
        {"name": "Stakeholders", "icon": "bi-people", "color": "#f59e0b", "total": len(all_stakeholders), "shared": len(stk_shared), "private": len(stk_private)},
        {"name": "Stakeholder Maps", "icon": "bi-diagram-3", "color": "#8b5cf6", "total": len(all_maps), "shared": len(maps_shared), "private": len(maps_private)},
        {"name": "Decisions", "icon": "bi-journal-bookmark-fill", "color": "#8b5cf6", "total": decisions_count, "shared": decisions_count, "private": 0},
    ]

    total_all = sum(c["total"] for c in categories)
    total_shared = sum(c["shared"] for c in categories)
    total_private = sum(c["private"] for c in categories)

    # Compute impact for private items (for star filter)
    from app.models import ImpactLevel
    impact_scale = ImpactLevel.get_org_levels(org_id)
    # Spaces have impact_level directly
    space_impacts = {s.id: s.impact_level or 0 for s in spaces_private}
    # Actions: use 0 (no direct impact)
    action_impacts = {a.id: 0 for a in actions_private}

    return render_template(
        "workspace/visibility_dashboard.html",
        categories=categories,
        total_all=total_all,
        total_shared=total_shared,
        total_private=total_private,
        shared_pct=int(total_shared / total_all * 100) if total_all else 100,
        spaces_private=spaces_private,
        actions_private=actions_private,
        private_owners=private_owners,
        impact_scale=impact_scale,
        space_impacts=space_impacts,
        action_impacts=action_impacts,
        csrf_token=generate_csrf,
    )


@bp.route("/impacts-dashboard")
@login_required
@organization_required
def impacts_dashboard():
    """Impacts Dashboard — impact assessment coverage, distribution, and gap analysis"""
    from app.models import (
        Challenge, ChallengeInitiativeLink, GovernanceBody, ImpactLevel,
        KPI, Space,
    )
    from app.models.system import InitiativeSystemLink, System
    from app.services.impact_service import METHODS, compute_true_importance

    org_id = session.get("organization_id")
    org = Organization.query.get(org_id)

    impact_scale = ImpactLevel.get_org_levels(org_id)
    _iw = {lvl: impact_scale[lvl]["weight"] for lvl in impact_scale} if impact_scale else {}
    _im = org.impact_calc_method or "geometric_mean" if org else "geometric_mean"
    _icm = org.impact_qfd_matrix if org else None
    _icr = org.impact_reinforce_weights if org else None
    method_info = METHODS.get(_im, METHODS.get("geometric_mean", {}))

    # Collect all entities with their impact data
    spaces = Space.query.filter_by(organization_id=org_id).order_by(Space.display_order).all()
    challenges = Challenge.query.filter_by(organization_id=org_id).all()
    initiatives = Initiative.query.filter_by(organization_id=org_id).all()
    systems = System.query.filter_by(organization_id=org_id).all()
    kpis = (KPI.query.join(InitiativeSystemLink).join(Initiative)
            .filter(Initiative.organization_id == org_id, KPI.is_archived.is_(False)).all())

    # Build entity data with chains
    entities = []
    type_stats = {}
    for etype, elist in [("space", spaces), ("challenge", challenges), ("initiative", initiatives), ("system", systems), ("kpi", kpis)]:
        assessed = sum(1 for e in elist if e.impact_level)
        total = len(elist)
        dist = {1: 0, 2: 0, 3: 0}
        for e in elist:
            if e.impact_level in dist:
                dist[e.impact_level] += 1
        type_stats[etype] = {"total": total, "assessed": assessed, "pct": int(assessed / total * 100) if total else 0, "dist": dist}

    # Compute true importance for all entities and find gaps
    all_entities = []
    ti_dist = {1: 0, 2: 0, 3: 0}
    gaps = []

    for sp in spaces:
        sp_ti = compute_true_importance([sp.impact_level], _im, _iw, _icm, _icr) if sp.impact_level else None
        all_entities.append({"type": "space", "name": sp.name, "id": sp.id, "impact": sp.impact_level, "true_importance": sp_ti, "chain": [sp.impact_level], "parent": None})
        if sp_ti and sp_ti in ti_dist: ti_dist[sp_ti] += 1
        if not sp.impact_level: gaps.append({"type": "space", "name": sp.name, "id": sp.id, "reason": "Not assessed"})

        for ch in [c for c in challenges if c.space_id == sp.id]:
            chain = [sp.impact_level, ch.impact_level]
            ch_ti = compute_true_importance(chain, _im, _iw, _icm, _icr) if all(chain) else None
            all_entities.append({"type": "challenge", "name": ch.name, "id": ch.id, "impact": ch.impact_level, "true_importance": ch_ti, "chain": chain, "parent": sp.name})
            if ch_ti and ch_ti in ti_dist: ti_dist[ch_ti] += 1
            if not ch.impact_level: gaps.append({"type": "challenge", "name": ch.name, "id": ch.id, "reason": "Not assessed", "parent": sp.name})
            elif sp.impact_level and ch.impact_level and not ch_ti: gaps.append({"type": "challenge", "name": ch.name, "id": ch.id, "reason": "Chain incomplete", "parent": sp.name})

            for ci in ChallengeInitiativeLink.query.filter_by(challenge_id=ch.id).all():
                ini = ci.initiative
                chain_i = [sp.impact_level, ch.impact_level, ini.impact_level]
                ini_ti = compute_true_importance(chain_i, _im, _iw, _icm, _icr) if all(chain_i) else None
                all_entities.append({"type": "initiative", "name": ini.name, "id": ini.id, "impact": ini.impact_level, "true_importance": ini_ti, "chain": chain_i, "parent": ch.name})
                if ini_ti and ini_ti in ti_dist: ti_dist[ini_ti] += 1
                if not ini.impact_level: gaps.append({"type": "initiative", "name": ini.name, "id": ini.id, "reason": "Not assessed", "parent": ch.name})

                for sl in InitiativeSystemLink.query.filter_by(initiative_id=ini.id).all():
                    sys = sl.system
                    chain_s = [sp.impact_level, ch.impact_level, ini.impact_level, sys.impact_level]
                    sys_ti = compute_true_importance(chain_s, _im, _iw, _icm, _icr) if all(chain_s) else None
                    all_entities.append({"type": "system", "name": sys.name, "id": sys.id, "impact": sys.impact_level, "true_importance": sys_ti, "chain": chain_s, "parent": ini.name})
                    if sys_ti and sys_ti in ti_dist: ti_dist[sys_ti] += 1
                    if not sys.impact_level: gaps.append({"type": "system", "name": sys.name, "id": sys.id, "reason": "Not assessed", "parent": ini.name})

                    for kpi in KPI.query.filter_by(initiative_system_link_id=sl.id, is_archived=False).all():
                        chain_k = [sp.impact_level, ch.impact_level, ini.impact_level, sys.impact_level, kpi.impact_level]
                        kpi_ti = compute_true_importance(chain_k, _im, _iw, _icm, _icr) if all(chain_k) else None
                        all_entities.append({"type": "kpi", "name": kpi.name, "id": kpi.id, "impact": kpi.impact_level, "true_importance": kpi_ti, "chain": chain_k, "parent": sys.name})
                        if kpi_ti and kpi_ti in ti_dist: ti_dist[kpi_ti] += 1
                        if not kpi.impact_level: gaps.append({"type": "kpi", "name": kpi.name, "id": kpi.id, "reason": "Not assessed", "parent": sys.name})

    total_entities = len(all_entities)
    total_assessed = sum(1 for e in all_entities if e["impact"])
    total_with_ti = sum(1 for e in all_entities if e["true_importance"])

    return render_template(
        "workspace/impacts_dashboard.html",
        type_stats=type_stats,
        ti_dist=ti_dist,
        gaps=gaps,
        total_gaps=len(gaps),
        total_entities=total_entities,
        total_assessed=total_assessed,
        total_with_ti=total_with_ti,
        coverage_pct=int(total_assessed / total_entities * 100) if total_entities else 0,
        ti_coverage_pct=int(total_with_ti / total_entities * 100) if total_entities else 0,
        impact_scale=impact_scale,
        method_name=method_info.get("name", _im),
        method_key=_im,
        all_entities=all_entities,
        csrf_token=generate_csrf,
    )


@bp.route("/challenges-dashboard")
@login_required
@organization_required
def challenges_dashboard():
    """Challenges Dashboard — strategic alignment and execution health per challenge"""
    from app.models import (
        Challenge, ChallengeInitiativeLink, Decision, GovernanceBody, ImpactLevel,
        KPI, Space,
    )
    from app.models.system import InitiativeSystemLink
    from app.services.impact_service import compute_true_importance

    org_id = session.get("organization_id")
    org = Organization.query.get(org_id)

    # Impact setup
    impact_scale = ImpactLevel.get_org_levels(org_id)
    _iw = {lvl: impact_scale[lvl]["weight"] for lvl in impact_scale} if impact_scale else {}
    _im = org.impact_calc_method or "geometric_mean" if org else "geometric_mean"
    _icm = org.impact_qfd_matrix if org else None
    _icr = org.impact_reinforce_weights if org else None

    spaces = Space.query.filter_by(organization_id=org_id).order_by(Space.display_order, Space.name).all()
    challenges = Challenge.query.filter_by(organization_id=org_id).order_by(Challenge.display_order, Challenge.name).all()

    challenge_list = []
    for ch in challenges:
        sp = ch.space
        ci_links = ChallengeInitiativeLink.query.filter_by(challenge_id=ch.id).all()
        initiatives = [cl.initiative for cl in ci_links]

        # RAG summary
        rags = [i.execution_rag for i in initiatives if i.execution_rag]
        worst_rag = "red" if "red" in rags else "amber" if "amber" in rags else "green" if "green" in rags else None
        rag_counts = {"green": rags.count("green"), "amber": rags.count("amber"), "red": rags.count("red")}

        # KPI count
        kpi_count = 0
        system_ids = set()
        for ini in initiatives:
            for sl in InitiativeSystemLink.query.filter_by(initiative_id=ini.id).all():
                system_ids.add(sl.system_id)
                kpi_count += KPI.query.filter_by(initiative_system_link_id=sl.id, is_archived=False).count()

        # Initiatives with progress updates
        ini_with_updates = sum(1 for i in initiatives if i.latest_progress_update)
        ini_with_kpis = 0
        for ini in initiatives:
            has_kpi = db.session.query(KPI.id).join(InitiativeSystemLink).filter(
                InitiativeSystemLink.initiative_id == ini.id, KPI.is_archived.is_(False)
            ).first()
            if has_kpi:
                ini_with_kpis += 1

        # Impact
        chain = []
        if sp and sp.impact_level:
            chain.append(sp.impact_level)
        if ch.impact_level:
            chain.append(ch.impact_level)
        true_importance = compute_true_importance(chain, _im, _iw, _icm, _icr) if len(chain) == 2 and all(chain) else None

        # Decision count
        decision_count = sum(1 for d in Decision.query.filter_by(organization_id=org_id).all()
                            if d.mentions_entity("challenge", ch.id))

        challenge_list.append({
            "id": ch.id,
            "name": ch.name,
            "space_name": sp.name if sp else None,
            "space_id": sp.id if sp else None,
            "initiative_count": len(initiatives),
            "system_count": len(system_ids),
            "kpi_count": kpi_count,
            "worst_rag": worst_rag,
            "rag_counts": rag_counts,
            "rag_dots": rags,
            "ini_with_updates": ini_with_updates,
            "ini_with_kpis": ini_with_kpis,
            "coverage_pct": int(ini_with_kpis / len(initiatives) * 100) if initiatives else 0,
            "decision_count": decision_count,
            "impact_level": ch.impact_level,
            "true_importance": true_importance,
        })

    # Stats
    fully_covered = sum(1 for c in challenge_list if c["coverage_pct"] == 100 and c["initiative_count"] > 0)
    partially_covered = sum(1 for c in challenge_list if 0 < c["coverage_pct"] < 100)
    empty = sum(1 for c in challenge_list if c["initiative_count"] == 0)

    governance_bodies = GovernanceBody.query.filter_by(organization_id=org_id, is_active=True).order_by(GovernanceBody.name).all()
    from app.models import UserFilterPreset
    presets = UserFilterPreset.query.filter_by(user_id=current_user.id, organization_id=org_id, feature="challenges_dashboard").order_by(UserFilterPreset.name).all()

    return render_template(
        "workspace/challenges_dashboard.html",
        challenge_list=challenge_list,
        spaces=spaces,
        stats={
            "total": len(challenge_list),
            "fully_covered": fully_covered,
            "partially_covered": partially_covered,
            "empty": empty,
            "red_count": sum(1 for c in challenge_list if c["worst_rag"] == "red"),
            "amber_count": sum(1 for c in challenge_list if c["worst_rag"] == "amber"),
            "green_count": sum(1 for c in challenge_list if c["worst_rag"] == "green"),
        },
        impact_scale=impact_scale,
        presets_list=[{"id": p.id, "name": p.name, "config": p.filters} for p in presets],
        csrf_token=generate_csrf,
    )


@bp.route("/systems-dashboard")
@login_required
@organization_required
def systems_dashboard():
    """Systems Dashboard — system health, reuse, and KPI coverage"""
    from app.models import (
        Challenge, ChallengeInitiativeLink, GovernanceBody, ImpactLevel,
        KPI, KPIValueTypeConfig, Space,
    )
    from app.models.system import InitiativeSystemLink, System
    from app.services.impact_service import compute_true_importance

    org_id = session.get("organization_id")
    org = Organization.query.get(org_id)

    impact_scale = ImpactLevel.get_org_levels(org_id)
    _iw = {lvl: impact_scale[lvl]["weight"] for lvl in impact_scale} if impact_scale else {}
    _im = org.impact_calc_method or "geometric_mean" if org else "geometric_mean"
    _icm = org.impact_qfd_matrix if org else None
    _icr = org.impact_reinforce_weights if org else None

    systems = System.query.filter_by(organization_id=org_id).order_by(System.name).all()

    system_list = []
    for sys in systems:
        # Initiative links
        isl_links = InitiativeSystemLink.query.filter_by(system_id=sys.id).all()
        initiative_names = []
        best_chain = None
        for isl in isl_links:
            ini = isl.initiative
            initiative_names.append(ini.name)
            # Build chain for impact
            ci = ChallengeInitiativeLink.query.filter_by(initiative_id=ini.id).first()
            ch = ci.challenge if ci else None
            sp = ch.space if ch else None
            chain = []
            if sp and sp.impact_level: chain.append(sp.impact_level)
            if ch and ch.impact_level: chain.append(ch.impact_level)
            if ini.impact_level: chain.append(ini.impact_level)
            if sys.impact_level: chain.append(sys.impact_level)
            if len(chain) == 4 and all(chain):
                ti = compute_true_importance(chain, _im, _iw, _icm, _icr)
                if best_chain is None or (ti and (best_chain is None or ti > best_chain)):
                    best_chain = ti

        # KPI count and health
        total_kpis = 0
        kpis_on_track = 0
        kpis_off_track = 0
        for isl in isl_links:
            kpis = KPI.query.filter_by(initiative_system_link_id=isl.id, is_archived=False).all()
            for kpi in kpis:
                total_kpis += 1
                for config in kpi.value_type_configs:
                    if config.target_value is not None:
                        consensus = config.get_consensus_value()
                        if consensus and consensus.get("value") is not None:
                            try:
                                tv = float(config.target_value)
                                cv = float(consensus["value"])
                                td = config.target_direction or "maximize"
                                if td == "minimize":
                                    pct = int((tv / cv) * 100) if cv != 0 else 100
                                else:
                                    pct = int((cv / tv) * 100) if tv != 0 else 0
                                if pct >= 80: kpis_on_track += 1
                                else: kpis_off_track += 1
                            except (ValueError, TypeError, ZeroDivisionError):
                                pass
                        break

        # Portal
        portal = None
        if sys.linked_organization_id and sys.linked_organization:
            lo = sys.linked_organization
            portal = {"id": lo.id, "name": lo.name}

        system_list.append({
            "id": sys.id,
            "name": sys.name,
            "description": sys.description,
            "initiative_count": len(isl_links),
            "initiative_names": initiative_names,
            "kpi_count": total_kpis,
            "kpis_on_track": kpis_on_track,
            "kpis_off_track": kpis_off_track,
            "impact_level": sys.impact_level,
            "true_importance": best_chain,
            "is_portal": portal is not None,
            "portal": portal,
            "is_shared": len(isl_links) > 1,
        })

    # Stats
    portal_count = sum(1 for s in system_list if s["is_portal"])
    shared_count = sum(1 for s in system_list if s["is_shared"])
    orphan_count = sum(1 for s in system_list if s["kpi_count"] == 0)

    from app.models import UserFilterPreset
    presets = UserFilterPreset.query.filter_by(user_id=current_user.id, organization_id=org_id, feature="systems_dashboard").order_by(UserFilterPreset.name).all()

    return render_template(
        "workspace/systems_dashboard.html",
        system_list=system_list,
        stats={
            "total": len(system_list),
            "portal_count": portal_count,
            "shared_count": shared_count,
            "with_kpis": sum(1 for s in system_list if s["kpi_count"] > 0),
            "orphan_count": orphan_count,
        },
        impact_scale=impact_scale,
        presets_list=[{"id": p.id, "name": p.name, "config": p.filters} for p in presets],
        csrf_token=generate_csrf,
    )


@bp.route("/kpi-dashboard")
@login_required
@organization_required
def kpi_dashboard():
    """KPI Dashboard — performance overview across all KPIs"""
    from app.models import (
        ChallengeInitiativeLink, Challenge, Decision, GovernanceBody, ImpactLevel,
        KPI, KPIGovernanceBodyLink, KPIValueTypeConfig, Space, ValueType,
    )
    from app.models.system import InitiativeSystemLink, System
    from app.services.impact_service import compute_true_importance

    org_id = session.get("organization_id")
    org = Organization.query.get(org_id)

    # Impact computation setup
    impact_scale = ImpactLevel.get_org_levels(org_id)
    _iw = {lvl: impact_scale[lvl]["weight"] for lvl in impact_scale} if impact_scale else {}
    _im = org.impact_calc_method or "geometric_mean" if org else "geometric_mean"
    _icm = org.impact_qfd_matrix if org else None
    _icr = org.impact_reinforce_weights if org else None

    # Get all KPIs with eager loading
    kpis = (
        KPI.query
        .join(InitiativeSystemLink)
        .join(Initiative, InitiativeSystemLink.initiative_id == Initiative.id)
        .filter(Initiative.organization_id == org_id)
        .options(
            db.joinedload(KPI.initiative_system_link).joinedload(InitiativeSystemLink.system),
            db.joinedload(KPI.initiative_system_link).joinedload(InitiativeSystemLink.initiative),
            db.joinedload(KPI.value_type_configs).joinedload(KPIValueTypeConfig.value_type),
            db.joinedload(KPI.governance_body_links).joinedload(KPIGovernanceBodyLink.governance_body),
        )
        .order_by(KPI.display_order, KPI.name)
        .all()
    )

    # Build KPI data
    kpi_list = []
    governance_bodies_used = set()
    for kpi in kpis:
        link = kpi.initiative_system_link
        if not link:
            continue
        ini = link.initiative
        sys = link.system

        # Parent chain for impact
        ci = ChallengeInitiativeLink.query.filter_by(initiative_id=ini.id).first()
        ch = ci.challenge if ci else None
        sp = ch.space if ch else None

        # Compute true importance
        chain = []
        if sp and sp.impact_level:
            chain.append(sp.impact_level)
        if ch and ch.impact_level:
            chain.append(ch.impact_level)
        if ini.impact_level:
            chain.append(ini.impact_level)
        if sys and sys.impact_level:
            chain.append(sys.impact_level)
        if kpi.impact_level:
            chain.append(kpi.impact_level)
        true_importance = compute_true_importance(chain, _im, _iw, _icm, _icr) if len(chain) == 5 and all(chain) else None

        # Get primary value (first config with consensus)
        primary_value = None
        primary_formatted = None
        primary_color = None
        target_progress = None
        target_value = None
        target_direction = None
        target_date = None
        consensus_status = None
        value_type_name = None

        for config in kpi.value_type_configs:
            consensus = config.get_consensus_value()
            if consensus and consensus.get("value") is not None:
                vt = config.value_type
                value_type_name = vt.name if vt else None
                primary_value = consensus.get("value")
                consensus_status = consensus.get("status", "no_data")
                try:
                    primary_formatted = current_app.jinja_env.filters["format_value"](primary_value, vt, config)
                except Exception:
                    primary_formatted = str(primary_value)
                primary_color = config.get_value_color(primary_value) if hasattr(config, 'get_value_color') else None

                if config.target_value is not None:
                    target_value = config.target_value
                    target_direction = config.target_direction or "maximize"
                    target_date = config.target_date
                    try:
                        tv = float(config.target_value)
                        cv = float(primary_value)
                        if target_direction == "minimize":
                            target_progress = int((tv / cv) * 100) if cv != 0 else 100
                        elif target_direction == "exact":
                            tol = tv * (config.target_tolerance_pct or 10) / 100
                            diff = abs(cv - tv)
                            target_progress = 100 if diff <= tol else max(0, int(100 - ((diff - tol) / tv * 100)))
                        else:
                            target_progress = int((cv / tv) * 100) if tv != 0 else 0
                    except (ValueError, TypeError, ZeroDivisionError):
                        target_progress = None
                break  # Use first config with data

        # GBs
        gbs = []
        for gbl in kpi.governance_body_links:
            gb = gbl.governance_body
            gbs.append({"id": gb.id, "name": gb.name, "abbreviation": gb.abbreviation, "color": gb.color})
            governance_bodies_used.add(gb.id)

        kpi_list.append({
            "id": kpi.id,
            "name": kpi.name,
            "is_archived": kpi.is_archived,
            "space_name": sp.name if sp else None,
            "challenge_name": ch.name if ch else None,
            "initiative_name": ini.name,
            "system_name": sys.name if sys else None,
            "value_type_name": value_type_name,
            "value": primary_value,
            "formatted_value": primary_formatted,
            "value_color": primary_color,
            "consensus_status": consensus_status,
            "target_value": target_value,
            "target_progress": target_progress,
            "target_direction": target_direction,
            "target_date": target_date.strftime("%Y-%m-%d") if target_date else None,
            "impact_level": kpi.impact_level,
            "true_importance": true_importance,
            "governance_bodies": gbs,
        })

    # Stats
    has_target = [k for k in kpi_list if k["target_progress"] is not None]
    on_track = [k for k in has_target if k["target_progress"] >= 80]
    at_risk = [k for k in has_target if 50 <= k["target_progress"] < 80]
    off_track = [k for k in has_target if k["target_progress"] < 50]

    # Governance bodies for filter
    governance_bodies = GovernanceBody.query.filter_by(organization_id=org_id, is_active=True).order_by(GovernanceBody.name).all()

    # Presets
    from app.models import UserFilterPreset
    presets = UserFilterPreset.query.filter_by(user_id=current_user.id, organization_id=org_id, feature="kpi_dashboard").order_by(UserFilterPreset.name).all()

    return render_template(
        "workspace/kpi_dashboard.html",
        kpi_list=kpi_list,
        stats={
            "total": len(kpi_list),
            "archived": sum(1 for k in kpi_list if k["is_archived"]),
            "with_target": len(has_target),
            "on_track": len(on_track),
            "at_risk": len(at_risk),
            "off_track": len(off_track),
            "no_data": sum(1 for k in kpi_list if k["consensus_status"] in (None, "no_data")),
        },
        governance_bodies=governance_bodies,
        impact_scale=impact_scale,
        presets_list=[{"id": p.id, "name": p.name, "config": p.filters} for p in presets],
        csrf_token=generate_csrf,
    )


@bp.route("/review")
@login_required
@organization_required
def start_review():
    """Start initiatives execution review — redirects to first initiative form.
    Accepts ?ids=1,2,3 for filtered initiatives, otherwise uses all."""
    org_id = session.get("organization_id")

    ids_param = request.args.get("ids", "")
    if ids_param:
        ids = [int(x) for x in ids_param.split(",") if x.strip().isdigit()]
    else:
        initiatives = (
            Initiative.query.filter_by(organization_id=org_id)
            .order_by(Initiative.name)
            .all()
        )
        ids = [i.id for i in initiatives]

    if not ids:
        flash("No initiatives to review.", "warning")
        return redirect(url_for("workspace.index"))

    back = request.args.get("back", url_for("workspace.index"))
    return redirect(
        url_for("organization_admin.initiative_form", initiative_id=ids[0])
        + f"?tab=execution&nav={','.join(str(i) for i in ids)}&nav_pos=0&nav_back={back}"
    )


@bp.route("/governance/<int:gb_id>")
@login_required
@organization_required
def gb_dashboard(gb_id):
    """Governance Body Dashboard — KPIs, actions, decisions for one GB"""
    from app.models import (
        ActionItem, ActionItemMention, ChallengeInitiativeLink, GovernanceBody,
        InitiativeProgressUpdate, InitiativeSystemLink, KPIGovernanceBodyLink,
    )
    from app.services import ConsensusService

    org_id = session.get("organization_id")
    gb = GovernanceBody.query.filter_by(id=gb_id, organization_id=org_id).first_or_404()

    # All GBs for the selector
    all_gbs = GovernanceBody.query.filter_by(organization_id=org_id, is_active=True).order_by(GovernanceBody.display_order).all()

    # ── My KPIs ──
    kpi_links = KPIGovernanceBodyLink.query.filter_by(governance_body_id=gb_id).all()
    my_kpis = []
    initiative_ids = set()
    for link in kpi_links:
        kpi = link.kpi
        if not kpi or not kpi.initiative_system_link:
            continue
        isl = kpi.initiative_system_link
        initiative_ids.add(isl.initiative_id)
        # Get all value type configs with formatted values
        from flask import current_app
        kpi_values = []
        for config in kpi.value_type_configs:
            vt = config.value_type
            cons = ConsensusService.get_cell_value(config)
            raw_val = cons.get("value") if cons else None
            formatted = None
            if raw_val is not None:
                try:
                    formatted = current_app.jinja_env.filters["format_value"](raw_val, vt, config)
                except Exception:
                    formatted = str(raw_val)
            kpi_values.append({
                "name": vt.name,
                "kind": vt.kind,
                "value": raw_val,
                "formatted": formatted,
                "unit": vt.unit_label,
                "color": config.get_value_color(raw_val) if raw_val is not None else None,
                "status": cons.get("status") if cons else "no_data",
            })

        primary_status = kpi_values[0]["status"] if kpi_values else "no_data"
        my_kpis.append({
            "id": kpi.id,
            "name": kpi.name,
            "system_name": isl.system.name,
            "initiative_name": isl.initiative.name,
            "vt_values": kpi_values,
            "status": primary_status,
            "is_archived": kpi.is_archived,
        })

    # ── My Actions ──
    from sqlalchemy import or_ as sql_or
    action_gb_table = db.Table("action_item_governance_body", db.metadata, autoload_with=db.engine)
    action_ids = db.session.query(action_gb_table.c.action_item_id).filter(
        action_gb_table.c.governance_body_id == gb_id
    ).all()
    action_id_list = [a[0] for a in action_ids]
    my_actions = ActionItem.query.filter(
        ActionItem.id.in_(action_id_list), ActionItem.organization_id == org_id
    ).order_by(ActionItem.due_date).all() if action_id_list else []

    # ── My Decisions ──
    all_updates = InitiativeProgressUpdate.query.join(Initiative).filter(
        Initiative.organization_id == org_id, InitiativeProgressUpdate.decisions.isnot(None)
    ).order_by(InitiativeProgressUpdate.created_at.desc()).all()

    from app.models import Decision
    _gb_decisions = Decision.query.filter_by(organization_id=org_id, governance_body_id=gb_id).order_by(Decision.created_at.desc()).all()
    my_decisions = []
    for dd in _gb_decisions:
        ini_name = None
        ini_id = None
        for m in (dd.entity_mentions or []):
            if m.get("entity_type") == "initiative":
                ini_name = m.get("entity_name")
                ini_id = m.get("entity_id")
                break
        my_decisions.append({
            "date": dd.created_at,
            "what": dd.what,
            "who": dd.who or "",
            "tag": ", ".join(dd.tags) if dd.tags else "",
            "initiative_name": ini_name,
            "initiative_id": ini_id,
        })

    # ── My Initiatives (derived from KPIs) ──
    my_initiatives = []
    if initiative_ids:
        inits = Initiative.query.filter(Initiative.id.in_(initiative_ids)).order_by(Initiative.name).all()
        for ini in inits:
            my_initiatives.append({
                "id": ini.id,
                "name": ini.name,
                "rag": ini.execution_rag,
                "impact_level": ini.impact_level,
            })

    # ── Stats ──
    stats = {
        "kpi_count": len(my_kpis),
        "kpi_archived": sum(1 for k in my_kpis if k["is_archived"]),
        "action_count": len(my_actions),
        "action_overdue": sum(1 for a in my_actions if a.is_overdue),
        "action_active": sum(1 for a in my_actions if a.status == "active"),
        "decision_count": len(my_decisions),
        "initiative_count": len(my_initiatives),
    }

    return render_template(
        "workspace/gb_dashboard.html",
        gb=gb,
        all_gbs=all_gbs,
        my_kpis=my_kpis,
        my_actions=my_actions,
        my_decisions=my_decisions,
        my_initiatives=my_initiatives,
        stats=stats,
    )


@bp.route("/impact-docs")
@login_required
def impact_docs():
    """Documentation page for impact compounding methods"""
    return render_template("workspace/impact_docs.html")


@bp.route("/strategy")
@login_required
@organization_required
def strategy_view():
    """Strategy view — card display of strategic pillars"""
    from app.models import Organization, StrategicPillar

    org_id = session.get("organization_id")
    org = Organization.query.get(org_id) if org_id else None
    if not (org and org.strategy_enabled):
        abort(404)

    org_id = session.get("organization_id")
    pillars = (
        StrategicPillar.query.filter_by(organization_id=org_id)
        .order_by(StrategicPillar.display_order)
        .all()
    )
    return render_template("workspace/strategy.html", pillars=pillars)


@bp.route("/dimensions")
@login_required
@organization_required
def dimensions():
    """Value Dimensions — card view of all value types"""
    org_id = session.get("organization_id")
    value_types = ValueType.query.filter_by(
        organization_id=org_id, is_active=True
    ).order_by(ValueType.display_order).all()
    return render_template("workspace/dimensions.html", value_types=value_types)


@bp.route("/dashboard")
@login_required
@organization_required
def dashboard():
    """Dashboard with overview, charts, and recent activity"""
    org_id = session.get("organization_id")
    org_name = session.get("organization_name")

    # Get organization logo
    org = Organization.query.get(org_id)
    org_logo = None
    if org and org.logo_data:
        org_logo = f"data:{org.logo_mime_type};base64,{base64.b64encode(org.logo_data).decode('utf-8')}"

    # Get entity type defaults (for logos in stats cards)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"

        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    # Get statistics
    stats = {
        "spaces": Space.query.filter_by(organization_id=org_id).count(),
        "challenges": Challenge.query.join(Space).filter(Space.organization_id == org_id).count(),
        "initiatives": Initiative.query.filter_by(organization_id=org_id).count(),
        "systems": System.query.filter_by(organization_id=org_id).count(),
        "kpis": db.session.query(KPI)
        .join(InitiativeSystemLink)
        .join(Initiative)
        .filter(Initiative.organization_id == org_id)
        .count(),
        "value_types": ValueType.query.filter_by(organization_id=org_id, is_active=True).count(),
        "governance_bodies": GovernanceBody.query.filter_by(organization_id=org_id, is_active=True).count(),
        "initiatives_no_consensus": Initiative.query.filter_by(
            organization_id=org_id, impact_on_challenge="no_consensus"
        ).count(),
    }

    # Check if organization needs onboarding (empty org + user has admin permissions)
    is_empty_org = stats["spaces"] == 0 and stats["governance_bodies"] == 0 and stats["value_types"] == 0
    has_admin_permissions = current_user.can_manage_spaces(org_id)
    needs_onboarding = is_empty_org and has_admin_permissions

    # Get recent snapshots (last 5) - now with full snapshot info
    recent_snapshots = SnapshotService.get_all_snapshots(org_id, user_id=current_user.id, limit=5)

    # Get recent comments (last 10)
    recent_comments = (
        db.session.query(CellComment)
        .join(KPIValueTypeConfig)
        .join(KPI, KPIValueTypeConfig.kpi_id == KPI.id)
        .join(InitiativeSystemLink)
        .join(Initiative)
        .filter(Initiative.organization_id == org_id)
        .order_by(CellComment.created_at.desc())
        .limit(10)
        .all()
    )

    # Get unread mentions count
    unread_mentions = CommentService.get_unread_mentions_count(current_user.id)

    # Get active announcements for current user
    active_announcements = []
    all_announcements = SystemAnnouncement.query.filter_by(is_active=True).all()
    for ann in all_announcements:
        if ann.is_visible_for_user(current_user.id, org_id):
            # Check if user has already acknowledged (if dismissible)
            if ann.is_dismissible and ann.has_been_acknowledged_by(current_user.id):
                continue
            active_announcements.append(ann)

    # Calculate total action items using centralized service
    from app.services.action_items_service import ActionItemsService

    action_items_count = ActionItemsService.get_action_items_count(org_id)
    total_action_items = action_items_count["total"]

    return render_template(
        "workspace/dashboard.html",
        org_name=org_name,
        org_logo=org_logo,
        entity_defaults=entity_defaults,
        stats=stats,
        recent_snapshots=recent_snapshots,
        recent_comments=recent_comments,
        unread_mentions=unread_mentions,
        needs_onboarding=needs_onboarding,
        active_announcements=active_announcements,
        total_action_items=total_action_items,
        csrf_token=generate_csrf,
    )


@bp.route("/")
@login_required
@organization_required
def index():
    """Workspace - Fully reactive Alpine.js interface"""
    org_id = session.get("organization_id")
    org_name = session.get("organization_name")

    # Get organization for logo and Porter's completion
    org = Organization.query.get(org_id)
    org_logo = None
    porters_completion = None
    if org:
        if org.logo_data:
            org_logo = f"data:{org.logo_mime_type};base64,{base64.b64encode(org.logo_data).decode('utf-8')}"

        # Get Porter's Five Forces completion
        filled, total, status = org.get_porters_completion()
        porters_completion = {"filled": filled, "total": total, "status": status}

    # Get entity type defaults (for logos/icons in tree)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"

        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    # Get filter presets for this user (from database)
    filter_presets_objs = (
        UserFilterPreset.query.filter_by(user_id=current_user.id, organization_id=org_id, feature="workspace")
        .order_by(UserFilterPreset.name)
        .all()
    )

    # Convert to dicts for JSON serialization in template
    filter_presets = filter_presets_objs
    filter_presets_json = json.dumps([preset.to_dict() for preset in filter_presets_objs])

    # Unified preset bar format (config key instead of filters)
    ws_presets_list = [{"id": p.id, "name": p.name, "config": p.filters} for p in filter_presets_objs]

    # Read user preferences for this org
    membership = UserOrganizationMembership.query.filter_by(
        user_id=current_user.id, organization_id=org_id
    ).first()
    prefs = membership.preferences or {} if membership else {}
    ws_focus_mode = bool(prefs.get("ws_focus_mode", True))
    ws_show_badges = bool(prefs.get("ws_show_badges", True))
    ws_badge_mode = int(prefs.get("ws_badge_mode", 1))

    return render_template(
        "workspace/index.html",
        org_name=org_name,
        org_id=org_id,
        organization=org,
        org_logo=org_logo,
        porters_completion=porters_completion,
        entity_defaults=entity_defaults,
        filter_presets=filter_presets,
        filter_presets_json=Markup(filter_presets_json),
        ws_presets_list=ws_presets_list,
        can_contribute=current_user.can_contribute(org_id),
        can_view_snapshots=current_user.can_view_snapshots_for(org_id),
        can_create_snapshots=current_user.can_create_snapshots_for(org_id),
        csrf_token=generate_csrf,
        ws_focus_mode=ws_focus_mode,
        ws_show_badges=ws_show_badges,
        ws_badge_mode=ws_badge_mode,
        strategy_count=StrategicPillar.query.filter_by(organization_id=org_id).count(),
    )


@bp.route("/export-excel")
@login_required
@organization_required
def export_excel():
    """Export workspace to Excel file"""
    org_id = session.get("organization_id")
    org_name = session.get("organization_name")

    # Generate Excel file
    excel_file = ExcelExportService.export_workspace(org_id)

    # Create safe filename
    safe_org_name = "".join(c for c in org_name if c.isalnum() or c in (" ", "-", "_")).strip()
    filename = f"Workspace_{safe_org_name}.xlsx"

    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.route("/kpi/<int:kpi_id>/value-type/<int:vt_id>", methods=["GET", "POST"])
@login_required
@organization_required
def kpi_cell_detail(kpi_id, vt_id):
    """
    Detail page for one KPI cell (KPI + value type).

    Shows:
    - Breadcrumb (org > space > challenge > initiative > system > kpi)
    - Current consensus status
    - List of contributions
    - Form to add/edit contribution
    """
    org_id = session.get("organization_id")

    # Get KPI and value type
    kpi = KPI.query.get_or_404(kpi_id)
    value_type = ValueType.query.get_or_404(vt_id)

    # Security check: ensure KPI belongs to current organization
    is_link = kpi.initiative_system_link
    initiative = is_link.initiative
    if initiative.organization_id != org_id:
        flash("Access denied", "danger")
        return redirect(url_for("workspace.index"))

    # Get KPI-ValueType config
    config = KPIValueTypeConfig.query.filter_by(kpi_id=kpi_id, value_type_id=vt_id).first()
    if not config:
        flash("This KPI does not use this value type", "warning")
        return redirect(url_for("workspace.index"))

    # Get consensus status (handles manual, linked, and formula KPIs)
    consensus = config.get_consensus_value()

    # Get all contributions
    contributions = config.contributions

    # Handle contribution form
    form = ContributionForm()

    # Populate list_value choices for list value types
    if value_type.is_list():
        list_options = value_type.get_list_options()
        form.list_value.choices = [("", "Select...")] + [(opt["key"], opt["label"]) for opt in list_options]

    # Customize qualitative level choices based on value type kind
    if value_type.is_qualitative():
        if value_type.kind == "risk":
            form.qualitative_level.choices = [
                ("", "Select..."),
                ("1", "! (Low)"),
                ("2", "!! (Medium)"),
                ("3", "!!! (High)"),
            ]
        elif value_type.kind == "positive_impact":
            form.qualitative_level.choices = [
                ("", "Select..."),
                ("1", "★ (Low)"),
                ("2", "★★ (Medium)"),
                ("3", "★★★ (High)"),
            ]
        elif value_type.kind == "negative_impact":
            form.qualitative_level.choices = [
                ("", "Select..."),
                ("1", "▼ (Low)"),
                ("2", "▼▼ (Medium)"),
                ("3", "▼▼▼ (High)"),
            ]
        elif value_type.kind == "level":
            form.qualitative_level.choices = [
                ("", "Select..."),
                ("1", "● (Low)"),
                ("2", "●● (Medium)"),
                ("3", "●●● (High)"),
            ]
        elif value_type.kind == "sentiment":
            form.qualitative_level.choices = [
                ("", "Select..."),
                ("1", "☹️ (Negative)"),
                ("2", "😐 (Neutral)"),
                ("3", "😊 (Positive)"),
            ]

    # Check if KPI is archived
    if kpi.is_archived:
        flash(
            "This KPI is archived and cannot accept new contributions. Please unarchive it first if you need to add data.",
            "warning",
        )
        return redirect(url_for("workspace.index"))

    # Check if user has permission to contribute
    if not current_user.can_contribute(org_id):
        if request.method == "POST":
            flash("You do not have permission to contribute values", "danger")
            return redirect(url_for("workspace.index"))
        # If GET request, still allow viewing but disable form

    # Debug POST requests
    if request.method == "POST":
        import logging

        logger = logging.getLogger(__name__)
        logger.info(f"🔍 POST received - All form keys: {list(request.form.keys())}")

    if form.validate_on_submit():
        import logging

        logger = logging.getLogger(__name__)
        logger.info("🔍 POST - Form validation PASSED")
        contributor_name = form.contributor_name.data
        entry_mode = request.form.get("entry_mode", "contributing")  # 'new_data' or 'contributing'

        # Check if this is "new data" mode (time evolved)
        if entry_mode == "new_data":
            # Auto-create snapshot before replacing data
            try:
                snapshot_label = f"Auto: Before update by {contributor_name}"

                # Create snapshot for this specific KPI cell
                # Use allow_duplicates=True so multiple snapshots can be created on same day
                snapshot = SnapshotService.create_kpi_snapshot(
                    config_id=config.id,
                    snapshot_date=date.today(),
                    label=snapshot_label,
                    user_id=current_user.id,
                    allow_duplicates=True,  # Always create new snapshot for auto-snapshots
                )

                if snapshot:
                    flash(f"Snapshot created: {snapshot.snapshot_label} (value: {snapshot.get_value()})", "info")

                # Delete ALL existing contributions for this cell
                Contribution.query.filter_by(kpi_value_type_config_id=config.id).delete()

                # Create new contribution
                contribution = Contribution(
                    kpi_value_type_config_id=config.id, contributor_name=contributor_name, comment=form.comment.data
                )
                if value_type.is_numeric():
                    contribution.numeric_value = form.numeric_value.data
                elif value_type.is_list():
                    contribution.list_value = form.list_value.data or None
                else:
                    contribution.qualitative_level = form.qualitative_level.data

                db.session.add(contribution)
                db.session.commit()

                flash(f"Previous value saved in snapshot. New value entered by {contributor_name}", "success")

                # Preserve filter state when returning to workspace
                from urllib.parse import urlencode

                return_params = [("show_all_columns", "1")]
                processed_keys = set()
                for key in request.form.keys():
                    if key.startswith("workspace_filter_") and key not in processed_keys:
                        filter_key = key.replace("workspace_filter_", "")
                        values = request.form.getlist(key)
                        for value in values:
                            return_params.append((filter_key, value))
                        processed_keys.add(key)

                workspace_url = url_for("workspace.index")
                if return_params:
                    workspace_url = f"{workspace_url}?{urlencode(return_params)}"

                return redirect(workspace_url)

            except Exception as e:
                db.session.rollback()
                flash(f"Error creating snapshot: {str(e)}", "danger")
                return redirect(url_for("workspace.kpi_cell_detail", kpi_id=kpi_id, vt_id=vt_id))

        # Normal mode: contributing to current period
        # Check if this contributor already has a contribution for this cell
        existing = Contribution.query.filter_by(
            kpi_value_type_config_id=config.id, contributor_name=contributor_name
        ).first()

        if existing:
            # Update existing contribution
            existing.numeric_value = None
            existing.qualitative_level = None
            existing.list_value = None
            if value_type.is_numeric():
                existing.numeric_value = form.numeric_value.data
            elif value_type.is_list():
                existing.list_value = form.list_value.data or None
            else:
                existing.qualitative_level = form.qualitative_level.data
            existing.comment = form.comment.data
            flash(f"Contribution from {contributor_name} updated", "success")
        else:
            # Create new contribution
            contribution = Contribution(
                kpi_value_type_config_id=config.id, contributor_name=contributor_name, comment=form.comment.data
            )
            if value_type.is_numeric():
                contribution.numeric_value = form.numeric_value.data
            elif value_type.is_list():
                contribution.list_value = form.list_value.data or None
            else:
                contribution.qualitative_level = form.qualitative_level.data

            db.session.add(contribution)
            flash(f"Contribution from {contributor_name} added", "success")

        db.session.commit()

        # Preserve filter state when returning to workspace
        # Get filters from hidden form fields and build URL manually
        import logging
        import sys
        from urllib.parse import urlencode

        logger = logging.getLogger(__name__)
        logger.info("=" * 80)
        logger.info("🔍 STARTING FILTER PRESERVATION")
        logger.info("=" * 80)
        sys.stdout.flush()  # Force flush

        return_params = [("show_all_columns", "1")]

        # Extract workspace_filters from form data (these were added as hidden fields)
        processed_keys = set()
        logger.info(f"🔍 POST - Form keys: {list(request.form.keys())}")
        for key in request.form.keys():
            if key.startswith("workspace_filter_") and key not in processed_keys:
                filter_key = key.replace("workspace_filter_", "")
                # Get all values for this key (handles multi-select filters)
                values = request.form.getlist(key)
                logger.info(f"🔍 Found filter: {filter_key} = {values}")
                for value in values:
                    return_params.append((filter_key, value))
                processed_keys.add(key)

        # Log what we're redirecting with
        logger.info(f"🔍 Redirecting to workspace with params: {return_params}")

        # Build URL manually to properly handle multiple values for same key
        workspace_url = url_for("workspace.index")
        if return_params:
            workspace_url = f"{workspace_url}?{urlencode(return_params)}"

        logger.info(f"🔍 Final redirect URL: {workspace_url}")
        return redirect(workspace_url)

    # Build breadcrumb
    system = is_link.system
    challenge_names = [ci.challenge.name for ci in initiative.challenge_links]
    space_names = [ci.challenge.space.name for ci in initiative.challenge_links]

    breadcrumb = {
        "organization": session.get("organization_name"),
        "space": space_names[0] if space_names else "N/A",
        "challenge": challenge_names[0] if challenge_names else "N/A",
        "initiative": initiative.name,
        "system": system.name,
        "kpi": kpi.name,
        "value_type": value_type.name,
    }

    # Get formula details if this is a formula KPI
    formula_details = None
    if config.is_formula():
        source_configs = config.get_formula_source_configs()
        mode = config.calculation_config.get("mode", "simple")

        formula_details = {
            "mode": mode,
            "operation": config.calculation_config.get("operation", "sum"),
            "expression": config.calculation_config.get("expression"),
            "expression_evaluated": None,
            "sources": [],
            "values": [],
        }

        # Build namespace for expression evaluation
        namespace = {}

        for source_config in source_configs:
            source_kpi = source_config.kpi
            source_vt = source_config.value_type
            source_initiative = source_kpi.initiative_system_link.initiative
            source_system = source_kpi.initiative_system_link.system
            source_org = source_initiative.organization

            # Get current value
            source_consensus = ConsensusService.get_cell_value(source_config)
            source_value = source_consensus.get("value") if source_consensus else None

            formula_details["sources"].append(
                {
                    "kpi_name": source_kpi.name,
                    "value_type_name": source_vt.name,
                    "organization_name": source_org.name,
                    "path": f"{source_initiative.name} › {source_system.name}",
                    "current_value": source_value,
                    "unit": source_vt.unit_label,
                }
            )

            if source_value is not None:
                float_value = float(source_value)
                formula_details["values"].append(float_value)
                namespace[f"kpi_{source_config.id}"] = float_value

        # For advanced mode, show evaluated expression
        if mode == "advanced" and formula_details["expression"]:
            expr = formula_details["expression"]
            # Replace variables with values for display
            for config_id, value in namespace.items():
                expr = expr.replace(config_id, str(value))
            formula_details["expression_evaluated"] = expr

    # Get entity type defaults (for logos)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"
        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    # Capture workspace filter state from referrer (for preserving on return)
    import logging
    from urllib.parse import parse_qs, urlparse

    logger = logging.getLogger(__name__)
    workspace_filters = {}
    referrer = request.referrer
    logger.info(f"🔍 KPI Detail GET - Referrer URL: {referrer}")
    if referrer and "workspace" in referrer:
        parsed = urlparse(referrer)
        query_params = parse_qs(parsed.query)
        for key, values in query_params.items():
            workspace_filters[key] = values[0] if len(values) == 1 else values
        logger.info(f"🔍 Captured workspace_filters: {workspace_filters}")
    else:
        logger.info("🔍 No workspace in referrer, workspace_filters empty")

    return render_template(
        "workspace/kpi_cell_detail.html",
        kpi=kpi,
        value_type=value_type,
        config=config,
        consensus=consensus,
        contributions=contributions,
        form=form,
        breadcrumb=breadcrumb,
        formula_details=formula_details,
        can_contribute=current_user.can_contribute(org_id),
        entity_defaults=entity_defaults,
        workspace_filters=workspace_filters,
        csrf_token=generate_csrf,
    )


@bp.route("/kpi/<int:kpi_id>/value-type/<int:vt_id>/delete-contribution", methods=["POST"])
@login_required
@organization_required
def delete_contribution(kpi_id, vt_id):
    """
    Delete a contribution from a KPI cell.
    """
    org_id = session.get("organization_id")
    contribution_id = request.form.get("contribution_id")

    if not contribution_id:
        flash("Invalid request", "danger")
        return redirect(url_for("workspace.kpi_cell_detail", kpi_id=kpi_id, vt_id=vt_id))

    # Get contribution and verify ownership
    contribution = Contribution.query.get_or_404(contribution_id)
    config = contribution.kpi_value_type_config
    kpi = config.kpi
    initiative = kpi.initiative_system_link.initiative

    if initiative.organization_id != org_id:
        flash("Access denied", "danger")
        return redirect(url_for("workspace.index"))

    contributor_name = contribution.contributor_name
    db.session.delete(contribution)
    db.session.commit()

    flash(f'Contribution from "{contributor_name}" has been deleted', "success")
    return redirect(url_for("workspace.kpi_cell_detail", kpi_id=kpi_id, vt_id=vt_id))


@bp.route("/api/rollup/<string:entity_type>/<int:entity_id>/<int:value_type_id>")
@login_required
@organization_required
def api_rollup(entity_type, entity_id, value_type_id):
    """
    API endpoint to get roll-up value for a specific entity and value type.

    Used by the tree/grid to display rolled-up values.
    """
    try:
        if entity_type == "system":
            # KPI → System rollup
            from app.models import InitiativeSystemLink

            is_link = InitiativeSystemLink.query.get(entity_id)
            if not is_link:
                return jsonify({"error": "Not found"}), 404

            result = AggregationService.get_kpi_to_system_rollup(is_link, value_type_id)
            return jsonify(result)

        elif entity_type == "initiative":
            result = AggregationService.get_system_to_initiative_rollup(entity_id, value_type_id)
            return jsonify(result)

        elif entity_type == "challenge":
            result = AggregationService.get_initiative_to_challenge_rollup(entity_id, value_type_id)
            return jsonify(result)

        elif entity_type == "space":
            result = AggregationService.get_challenge_to_space_rollup(entity_id, value_type_id)
            return jsonify(result)

        else:
            return jsonify({"error": "Invalid entity type"}), 400

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# SNAPSHOT MANAGEMENT ROUTES (Time-Series Tracking)
# ============================================================================


@bp.route("/snapshots/pivot")
@login_required
@organization_required
def snapshot_pivot():
    """
    Pivot table view of snapshots: KPIs as rows, time periods as columns.

    Allows filtering by year, view type (daily/weekly/monthly/quarterly/yearly), space, and value type.
    """
    org_id = session.get("organization_id")

    # Get available years
    available_years = SnapshotPivotService.get_available_years(org_id)

    if not available_years:
        flash("No snapshots found. Create your first snapshot to start time-series analysis.", "info")
        return redirect(url_for("workspace.list_snapshots"))

    # Get filter parameters
    use_custom_range = request.args.get("use_custom_range") == "true"
    view_type = request.args.get("view_type", "quarterly")
    space_id = request.args.get("space_id", type=int)
    challenge_id = request.args.get("challenge_id", type=int)
    value_type_id = request.args.get("value_type_id", type=int)
    show_targets = request.args.get("show_targets") == "on"
    show_status = request.args.get("show_status") == "on"

    if use_custom_range:
        # Custom date range mode
        start_month = request.args.get("start_month", type=int) or 1
        start_year = request.args.get("start_year_custom", type=int) or available_years[0]
        end_month = request.args.get("end_month", type=int) or 12
        end_year = request.args.get("end_year_custom", type=int) or available_years[-1]

        year_start = start_year
        year_end = end_year

        # Build list of months in range
        custom_months = []
        current_year = start_year
        current_month = start_month

        while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
            custom_months.append((current_year, current_month))
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

        periods = None  # Will filter by custom_months list instead
        quarters = []
        months = []
    else:
        # Simple mode
        year_start = request.args.get("year_start", type=int) or available_years[0]
        year_end = request.args.get("year_end", type=int) or year_start
        quarters = request.args.getlist("quarters", type=int)
        months = request.args.getlist("months", type=int)
        periods = quarters if view_type == "quarterly" else months if view_type == "monthly" else None
        custom_months = None

    # Get pivot data
    pivot_data = SnapshotPivotService.get_pivot_data(
        org_id,
        view_type=view_type,
        space_id=space_id,
        challenge_id=challenge_id,
        value_type_id=value_type_id,
        year_start=year_start,
        year_end=year_end,
        periods=periods,
        custom_months=custom_months,
    )

    # Get filter options - respect private spaces
    spaces_query = Space.query.filter_by(organization_id=org_id)
    # Filter out private spaces unless user is the owner or has admin access
    if not current_user.is_global_admin and not current_user.is_org_admin(org_id):
        spaces_query = spaces_query.filter(
            or_(Space.is_private.is_(False), Space.created_by_user_id == current_user.id)
        )
    spaces = spaces_query.order_by(Space.display_order, Space.name).all()

    # Get all challenges - filtered by accessible spaces only
    from app.models import Challenge

    # Get IDs of accessible spaces
    accessible_space_ids = [s.id for s in spaces]

    all_challenges = (
        Challenge.query.filter_by(organization_id=org_id)
        .filter(Challenge.space_id.in_(accessible_space_ids))
        .order_by(Challenge.display_order, Challenge.name)
        .all()
    )

    value_types = ValueType.query.filter_by(organization_id=org_id, is_active=True).order_by(ValueType.name).all()

    # Prepare template variables
    template_vars = {
        "pivot_data": pivot_data,
        "available_years": available_years,
        "current_year": year_start,
        "year_start": year_start,
        "year_end": year_end,
        "view_type": view_type,
        "spaces": spaces,
        "challenges": all_challenges,
        "value_types": value_types,
        "selected_space_id": space_id,
        "selected_challenge_id": challenge_id,
        "selected_value_type_id": value_type_id,
        "selected_quarters": quarters,
        "selected_months": months,
        "use_custom_range": use_custom_range,
        "show_targets": show_targets,
        "show_status": show_status,
        "csrf_token": generate_csrf,
    }

    # Add custom date values if in custom mode
    if use_custom_range:
        template_vars["start_month"] = request.args.get("start_month", type=int) or 1
        template_vars["start_year_custom"] = request.args.get("start_year_custom", type=int) or available_years[0]
        template_vars["end_month"] = request.args.get("end_month", type=int) or 12
        template_vars["end_year_custom"] = request.args.get("end_year_custom", type=int) or available_years[-1]

    # Get entity type defaults (for logos)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"
        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }
    template_vars["entity_defaults"] = entity_defaults

    # Load saved charts for preset bar
    from sqlalchemy import or_ as sa_or
    chart_presets = (
        SavedChart.query.filter_by(organization_id=org_id)
        .filter(sa_or(SavedChart.created_by_user_id == current_user.id, SavedChart.is_shared.is_(True)))
        .order_by(SavedChart.updated_at.desc())
        .all()
    )
    template_vars["chart_presets_list"] = [
        {"id": c.id, "name": c.name, "config": {
            "year_start": c.year_start, "year_end": c.year_end,
            "view_type": c.view_type, "chart_type": c.chart_type,
            "space_id": c.space_id, "value_type_id": c.value_type_id,
            "period_filter": c.period_filter,
            "config_ids_colors": c.config_ids_colors,
        }}
        for c in chart_presets
    ]

    return render_template("workspace/snapshot_pivot.html", **template_vars)


@bp.route("/snapshots/chart-data")
@login_required
@organization_required
def snapshot_chart_data():
    """
    API endpoint to get chart data for selected KPIs.

    Returns JSON formatted for Chart.js.
    """
    org_id = session.get("organization_id")

    # Get parameters - support both old (year) and new (date range) format
    year = request.args.get("year", type=int)
    start_year = request.args.get("start_year_custom", type=int)
    end_year = request.args.get("end_year_custom", type=int)
    start_month = request.args.get("start_month", type=int, default=1)
    end_month = request.args.get("end_month", type=int, default=12)

    view_type = request.args.get("view_type", "quarterly")
    config_ids = request.args.getlist("config_ids", type=int)
    show_targets = request.args.get("show_targets") == "true"

    # Determine which format to use
    if start_year and end_year:
        # New format: use date range
        year_param = start_year  # Pass start year for now
    elif year:
        # Old format: single year
        year_param = year
    else:
        return jsonify({"error": "Missing required parameters"}), 400

    if not config_ids:
        return jsonify({"error": "No KPIs selected"}), 400

    try:
        chart_data = SnapshotPivotService.get_chart_data(
            org_id,
            year_param,
            config_ids,
            view_type,
            show_targets,
            start_year=start_year,
            end_year=end_year,
            start_month=start_month,
            end_month=end_month,
        )
        return jsonify(chart_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/snapshots/charts/save", methods=["POST"])
@login_required
@organization_required
def save_chart():
    """Save a chart configuration"""
    org_id = session.get("organization_id")
    data = request.get_json()

    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Chart name is required"}), 400

    config_colors = data.get("config_colors", {})
    if not config_colors:
        return jsonify({"error": "No KPIs selected"}), 400

    # Get year range - use current year as fallback
    from datetime import datetime

    current_year = datetime.now().year
    year_start = data.get("year_start") or data.get("start_year_custom") or current_year
    year_end = data.get("year_end") or data.get("end_year_custom") or current_year

    chart = SavedChart(
        organization_id=org_id,
        created_by_user_id=current_user.id,
        name=name,
        description=data.get("description", ""),
        year_start=year_start,
        year_end=year_end,
        view_type=data.get("view_type", "quarterly"),
        chart_type=data.get("chart_type", "line"),
        space_id=data.get("space_id"),
        value_type_id=data.get("value_type_id"),
        period_filter=data.get("period_filter"),
        is_shared=data.get("is_shared", False),
    )
    chart.set_config_colors(config_colors)

    db.session.add(chart)
    db.session.commit()

    return jsonify({"success": True, "id": chart.id, "name": chart.name, "is_shared": chart.is_shared})


@bp.route("/snapshots/charts/search")
@login_required
@organization_required
def search_charts():
    """Search saved charts (instant search) - shows user's private charts + public charts"""
    org_id = session.get("organization_id")
    query = request.args.get("q", "").strip()

    # Show: user's own charts (private or public) OR public charts from others
    charts_query = (
        SavedChart.query.filter_by(organization_id=org_id)
        .filter(
            or_(
                SavedChart.created_by_user_id == current_user.id,  # User's own charts
                SavedChart.is_shared.is_(True),  # Public charts from others
            )
        )
        .order_by(SavedChart.updated_at.desc())
    )

    if query:
        charts_query = charts_query.filter(
            or_(
                SavedChart.name.ilike(f"%{query}%"),
                SavedChart.description.ilike(f"%{query}%"),
            )
        )

    charts = charts_query.limit(20).all()

    return jsonify(
        {
            "charts": [
                {
                    "id": c.id,
                    "name": c.name,
                    "description": c.description,
                    "year_range": f"{c.year_start}-{c.year_end}" if c.year_start != c.year_end else str(c.year_start),
                    "view_type": c.view_type,
                    "chart_type": c.chart_type,
                    "kpi_count": len(c.get_config_colors()),
                    "is_shared": c.is_shared,
                    "is_owner": c.created_by_user_id == current_user.id,
                    "created_by": c.created_by.display_name,
                }
                for c in charts
            ]
        }
    )


@bp.route("/snapshots/charts/<int:chart_id>")
@login_required
@organization_required
def load_chart(chart_id):
    """Load a saved chart configuration"""
    org_id = session.get("organization_id")
    chart = SavedChart.query.filter_by(id=chart_id, organization_id=org_id).first_or_404()

    return jsonify(
        {
            "id": chart.id,
            "name": chart.name,
            "description": chart.description,
            "year_start": chart.year_start,
            "year_end": chart.year_end,
            "view_type": chart.view_type,
            "chart_type": chart.chart_type,
            "space_id": chart.space_id,
            "value_type_id": chart.value_type_id,
            "period_filter": chart.period_filter,
            "config_colors": chart.get_config_colors(),
        }
    )


@bp.route("/snapshots/pivot/export")
@login_required
@organization_required
def export_pivot_excel():
    """Export pivot table to Excel"""
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    org_id = session.get("organization_id")

    # Get available years for fallback
    available_years = SnapshotPivotService.get_available_years(org_id)
    if not available_years:
        flash("No data to export", "warning")
        return redirect(url_for("workspace.list_snapshots"))

    # Get same filter parameters as pivot view
    use_custom_range = request.args.get("use_custom_range") == "true"
    view_type = request.args.get("view_type", "quarterly")
    space_id = request.args.get("space_id", type=int)
    value_type_id = request.args.get("value_type_id", type=int)

    if use_custom_range:
        start_month = request.args.get("start_month", type=int) or 1
        start_year = request.args.get("start_year_custom", type=int) or available_years[0]
        end_month = request.args.get("end_month", type=int) or 12
        end_year = request.args.get("end_year_custom", type=int) or available_years[-1]

        year_start = start_year
        year_end = end_year

        custom_months = []
        current_year = start_year
        current_month = start_month
        while (current_year < end_year) or (current_year == end_year and current_month <= end_month):
            custom_months.append((current_year, current_month))
            current_month += 1
            if current_month > 12:
                current_month = 1
                current_year += 1

        periods = None
    else:
        year_start = request.args.get("year_start", type=int) or available_years[0]
        year_end = request.args.get("year_end", type=int) or year_start
        quarters = request.args.getlist("quarters", type=int)
        months = request.args.getlist("months", type=int)
        periods = quarters if view_type == "quarterly" else months if view_type == "monthly" else None
        custom_months = None

    # Get data
    pivot_data = SnapshotPivotService.get_pivot_data(
        org_id,
        view_type=view_type,
        space_id=space_id,
        value_type_id=value_type_id,
        year_start=year_start,
        year_end=year_end,
        periods=periods,
        custom_months=custom_months,
    )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snapshot Analysis"

    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers - metadata + target info + periods
    headers = [
        "Organization",
        "Space",
        "Challenge",
        "Initiative",
        "System",
        "KPI",
        "Value Type",
        "Target Value",
        "Target Date",
        "Target Direction",
        "Target Tolerance %",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Period headers start after the metadata columns
    for idx, period in enumerate(pivot_data["periods"], start=len(headers) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.value = period
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Get organization name
    from app.models import Organization

    org = Organization.query.get(org_id)
    org_name = org.name if org else "Unknown"

    # Write data
    row_num = 2
    for kpi in pivot_data["kpis"]:
        # Get KPI relationships for metadata
        kpi_obj = KPI.query.get(kpi["kpi_id"])
        if not kpi_obj:
            continue

        # Get Initiative and System through InitiativeSystemLink
        init_sys_link = InitiativeSystemLink.query.get(kpi_obj.initiative_system_link_id)
        if not init_sys_link:
            continue

        initiative = Initiative.query.get(init_sys_link.initiative_id)
        system = System.query.get(init_sys_link.system_id)

        # Get Challenge(s) through ChallengeInitiativeLink
        challenge_links = ChallengeInitiativeLink.query.filter_by(initiative_id=initiative.id).all()
        challenges = [Challenge.query.get(link.challenge_id) for link in challenge_links if link]
        challenge_names = ", ".join([c.name for c in challenges if c])

        # Get Space(s) from challenges
        spaces = list(set([c.space for c in challenges if c and c.space]))
        space_names = ", ".join([s.name for s in spaces if s])

        # Get config for target info
        config = kpi.get("config")

        # Write row data
        ws.cell(row=row_num, column=1, value=org_name)
        ws.cell(row=row_num, column=2, value=space_names)
        ws.cell(row=row_num, column=3, value=challenge_names)
        ws.cell(row=row_num, column=4, value=initiative.name if initiative else "")
        ws.cell(row=row_num, column=5, value=system.name if system else "")
        ws.cell(row=row_num, column=6, value=kpi["kpi_name"])
        ws.cell(row=row_num, column=7, value=kpi["value_type_name"])

        # Target information
        if config and config.target_value:
            ws.cell(row=row_num, column=8, value=float(config.target_value))
            ws.cell(row=row_num, column=9, value=config.target_date.strftime("%Y-%m-%d") if config.target_date else "")
            ws.cell(row=row_num, column=10, value=config.target_direction or "maximize")
            ws.cell(
                row=row_num, column=11, value=config.target_tolerance_pct if config.target_direction == "exact" else ""
            )
        else:
            ws.cell(row=row_num, column=8, value="")
            ws.cell(row=row_num, column=9, value="")
            ws.cell(row=row_num, column=10, value="")
            ws.cell(row=row_num, column=11, value="")

        # Write period values starting from column 12
        for idx, period in enumerate(pivot_data["periods"], start=12):
            if period in kpi["values"]:
                value_data = kpi["values"][period]
                if value_data["value"] is not None:
                    ws.cell(row=row_num, column=idx, value=float(value_data["value"]))

        row_num += 1

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"snapshot_analysis_{view_type}_{year_start}-{year_end}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@bp.route("/snapshots/create", methods=["POST"])
@login_required
@organization_required
def create_snapshot():
    """
    Create a snapshot of current workspace state.

    Captures all KPI values and rollups for the organization.
    Accepts optional period tag overrides (year, quarter, month).
    """
    org_id = session.get("organization_id")

    if not current_user.can_create_snapshots_for(org_id):
        flash("You don't have permission to create snapshots.", "danger")
        return redirect(url_for("workspace.index"))

    snapshot_date_str = request.form.get("snapshot_date")
    label = request.form.get("label", "").strip()
    is_public = request.form.get("is_public") == "true"

    # Parse date
    if snapshot_date_str:
        try:
            snapshot_date = date.fromisoformat(snapshot_date_str)
        except ValueError:
            flash("Invalid date format", "danger")
            return redirect(url_for("workspace.index"))
    else:
        snapshot_date = date.today()

    # Get optional period tag overrides
    year_override = request.form.get("year")
    quarter_override = request.form.get("quarter")
    month_override = request.form.get("month")

    # Convert to integers if provided
    try:
        year = int(year_override) if year_override else None
        quarter = int(quarter_override) if quarter_override else None
        month = int(month_override) if month_override else None
    except ValueError:
        flash("Invalid period tag values", "danger")
        return redirect(url_for("workspace.list_snapshots"))

    # Create snapshots
    try:
        result = SnapshotService.create_organization_snapshot(
            org_id,
            snapshot_date=snapshot_date,
            label=label or None,
            user_id=current_user.id,
            is_public=is_public,
            year_override=year,
            quarter_override=quarter,
            month_override=month,
        )

        visibility = "Public" if is_public else "Private"
        flash(
            f'{visibility} snapshot created: {result["kpi_snapshots"]} KPI values, '
            f'{result["rollup_snapshots"]} rollup values captured for {snapshot_date.isoformat()}',
            "success",
        )

        if result["skipped"] > 0:
            flash(f'{result["skipped"]} KPIs skipped (no consensus data)', "info")

    except Exception as e:
        db.session.rollback()
        flash(f"Error creating snapshot: {str(e)}", "danger")

    return redirect(url_for("workspace.list_snapshots"))


@bp.route("/snapshots/list")
@login_required
@organization_required
def list_snapshots():
    """List all available snapshots for the organization"""
    org_id = session.get("organization_id")

    if not current_user.can_view_snapshots_for(org_id):
        flash("You don't have permission to view snapshots.", "danger")
        return redirect(url_for("workspace.index"))

    # Get filter parameters
    show_private = request.args.get("show_private", "1") == "1"
    show_public = request.args.get("show_public", "1") == "1"

    try:
        # Get all snapshots with full details
        snapshots = SnapshotService.get_all_snapshots(
            org_id, user_id=current_user.id, show_private=show_private, show_public=show_public
        )

        # Format for template
        snapshots_info = []
        for snap in snapshots:
            snapshots_info.append(
                {
                    "batch_id": snap["snapshot_batch_id"],
                    "date": snap["snapshot_date"].isoformat(),
                    "created_at": snap["created_at"].strftime("%Y-%m-%d %H:%M:%S"),
                    "timestamp": snap["created_at"].isoformat(),
                    "label": snap["snapshot_label"],
                    "kpi_count": snap["kpi_count"],
                    "formatted_date": snap["snapshot_date"].strftime("%Y-%m-%d (%A)"),
                    "formatted_time": snap["created_at"].strftime("%H:%M:%S"),
                    "is_public": snap["is_public"],
                    "owner_user_id": snap["owner_user_id"],
                    "owner_name": snap["owner_name"],
                    "is_owner": snap["owner_user_id"] == current_user.id,
                }
            )

        return render_template(
            "workspace/snapshots.html",
            snapshots=snapshots_info,
            organization_name=session.get("organization_name"),
            show_private=show_private,
            show_public=show_public,
            current_user_id=current_user.id,
            csrf_token=generate_csrf,
            can_create_snapshots=current_user.can_create_snapshots_for(org_id),
        )

    except Exception as e:
        flash(f"Error loading snapshots: {str(e)}", "danger")
        return redirect(url_for("workspace.index"))


@bp.route("/snapshots/view/<batch_id>")
@login_required
@organization_required
def view_snapshot(batch_id):
    """
    View workspace state as of a specific snapshot batch.

    Shows historical values instead of current values.
    """
    org_id = session.get("organization_id")

    # Get snapshot info from batch
    sample = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).first()
    if not sample:
        flash("Snapshot not found", "danger")
        return redirect(url_for("workspace.list_snapshots"))

    view_date = sample.snapshot_date

    # Get spaces and value types (current structure)
    spaces = Space.query.filter_by(organization_id=org_id).order_by(Space.display_order, Space.name).all()

    value_types = (
        ValueType.query.filter_by(organization_id=org_id, is_active=True).order_by(ValueType.display_order).all()
    )

    # Get governance bodies for filtering
    governance_bodies = (
        GovernanceBody.query.filter_by(organization_id=org_id, is_active=True)
        .order_by(GovernanceBody.display_order)
        .all()
    )

    # Get level visibility controls (default all visible)
    show_levels = {
        "spaces": request.args.get("show_spaces", "1") == "1",
        "challenges": request.args.get("show_challenges", "1") == "1",
        "initiatives": request.args.get("show_initiatives", "1") == "1",
        "systems": request.args.get("show_systems", "1") == "1",
        "kpis": request.args.get("show_kpis", "1") == "1",
    }

    return render_template(
        "workspace/index.html",
        spaces=spaces,
        value_types=value_types,
        governance_bodies=governance_bodies,
        selected_governance_body_ids=[],
        show_archived=False,
        show_levels=show_levels,
        organization_name=session.get("organization_name"),
        snapshot_date=view_date,
        is_historical_view=True,
        csrf_token=generate_csrf,
    )


@bp.route("/snapshots/compare")
@login_required
@organization_required
def compare_snapshots():
    """Compare two snapshots side-by-side"""
    org_id = session.get("organization_id")

    # Get batch_id parameters
    batch_id1 = request.args.get("batch_id1")
    batch_id2 = request.args.get("batch_id2", "current")

    if not batch_id1:
        flash("Please select a snapshot to compare", "warning")
        return redirect(url_for("workspace.list_snapshots"))

    # Get first snapshot info
    sample1 = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id1).first()
    if not sample1:
        flash("Snapshot not found", "danger")
        return redirect(url_for("workspace.list_snapshots"))

    date1 = sample1.snapshot_date
    datetime1 = sample1.created_at
    label1 = sample1.snapshot_label

    # Get second snapshot info (or use current)
    if batch_id2 != "current":
        sample2 = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id2).first()
        if not sample2:
            flash("Second snapshot not found", "danger")
            return redirect(url_for("workspace.list_snapshots"))
        date2 = sample2.snapshot_date
        datetime2 = sample2.created_at
        label2 = sample2.snapshot_label
    else:
        date2 = None
        datetime2 = None
        label2 = "Current"

    # Get all KPI configs for this organization
    configs = (
        db.session.query(KPIValueTypeConfig)
        .join(KPI, KPIValueTypeConfig.kpi_id == KPI.id)
        .join(InitiativeSystemLink)
        .join(Initiative)
        .filter(Initiative.organization_id == org_id)
        .all()
    )

    # Build comparison data
    comparisons = []
    for config in configs:
        # Get snapshot 1 value - match by batch_id
        snapshot1 = KPISnapshot.query.filter_by(kpi_value_type_config_id=config.id, snapshot_batch_id=batch_id1).first()

        # Get snapshot 2 value (or current consensus)
        if batch_id2 != "current":
            snapshot2 = KPISnapshot.query.filter_by(
                kpi_value_type_config_id=config.id, snapshot_batch_id=batch_id2
            ).first()
            value2 = snapshot2.get_value() if snapshot2 else None
        else:
            # Use current consensus - get contributions for this config
            contributions = Contribution.query.filter_by(kpi_value_type_config_id=config.id).all()
            consensus = ConsensusService.calculate_consensus(contributions)
            value2 = consensus.get("value")

        value1 = snapshot1.get_value() if snapshot1 else None

        # Calculate change
        change = None
        percent_change = None
        if value1 is not None and value2 is not None:
            change = float(value2) - float(value1)
            if value1 != 0:
                percent_change = (change / float(value1)) * 100

        comparisons.append(
            {
                "config": config,
                "kpi": config.kpi,
                "value_type": config.value_type,
                "value1": value1,
                "value2": value2,
                "change": change,
                "percent_change": percent_change,
            }
        )

    # Get entity type defaults (for logos)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"
        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    return render_template(
        "workspace/compare_snapshots.html",
        comparisons=comparisons,
        date1=date1,
        datetime1=datetime1,
        date2=date2,
        datetime2=datetime2,
        label1=label1,
        label2=label2,
        organization_name=session.get("organization_name"),
        entity_defaults=entity_defaults,
        csrf_token=generate_csrf,
    )


@bp.route("/snapshots/<batch_id>/toggle-privacy", methods=["POST"])
@login_required
@organization_required
def toggle_snapshot_privacy(batch_id):
    """Toggle privacy status of a snapshot batch (private <-> public)"""
    try:
        print(f"[DEBUG] Toggling privacy for batch_id: {batch_id}")

        # Get one snapshot from this batch to check ownership
        sample_snapshot = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).first()

        if not sample_snapshot:
            print(f"[DEBUG] Snapshot not found for batch_id: {batch_id}")
            return jsonify({"error": "Snapshot not found"}), 404

        print(
            f"[DEBUG] Current is_public: {sample_snapshot.is_public}, owner: {sample_snapshot.owner_user_id}, current_user: {current_user.id}"
        )

        # Check ownership
        if sample_snapshot.owner_user_id != current_user.id:
            print(f"[DEBUG] Ownership check failed: {sample_snapshot.owner_user_id} != {current_user.id}")
            return jsonify({"error": "Only the snapshot owner can change privacy settings"}), 403

        # Toggle all KPI snapshots in this batch
        new_status = not sample_snapshot.is_public
        kpi_count = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).update({"is_public": new_status})

        # Toggle all rollup snapshots in this batch
        rollup_count = RollupSnapshot.query.filter_by(snapshot_batch_id=batch_id).update({"is_public": new_status})

        db.session.commit()

        print(
            f"[DEBUG] Toggled {kpi_count} KPI snapshots and {rollup_count} rollup snapshots to is_public={new_status}"
        )

        return jsonify(
            {
                "success": True,
                "is_public": new_status,
                "message": f'Snapshot is now {"public" if new_status else "private"}',
            }
        )

    except Exception:
        db.session.rollback()


@bp.route("/snapshots/<batch_id>/delete", methods=["POST"])
@login_required
@organization_required
def delete_snapshot(batch_id):
    """Delete all snapshots in a batch"""
    try:
        # Get one snapshot from this batch to check ownership
        sample_snapshot = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).first()

        if not sample_snapshot:
            return jsonify({"error": "Snapshot not found"}), 404

        # Check ownership
        if sample_snapshot.owner_user_id != current_user.id:
            return jsonify({"error": "Only the snapshot owner can delete snapshots"}), 403

        # Delete all KPI snapshots in this batch
        kpi_count = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).delete()

        # Delete all rollup snapshots in this batch
        rollup_count = RollupSnapshot.query.filter_by(snapshot_batch_id=batch_id).delete()

        db.session.commit()

        return jsonify(
            {
                "success": True,
                "message": f"Deleted snapshot batch ({kpi_count} KPI snapshots, {rollup_count} rollup snapshots)",
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/snapshots/bulk/toggle-privacy", methods=["POST"])
@login_required
@organization_required
def bulk_toggle_snapshot_privacy():
    """Toggle privacy status for multiple snapshot batches"""
    try:
        data = request.get_json()
        batch_ids = data.get("batch_ids", [])
        make_public = data.get("make_public", True)

        if not batch_ids:
            return jsonify({"error": "No batch IDs provided"}), 400

        success_count = 0
        error_count = 0

        for batch_id in batch_ids:
            # Get one snapshot from this batch to check ownership
            sample_snapshot = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).first()

            if not sample_snapshot:
                error_count += 1
                continue

            # Check ownership
            if sample_snapshot.owner_user_id != current_user.id:
                error_count += 1
                continue

            # Update all KPI snapshots in this batch
            KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).update({"is_public": make_public})

            # Update all rollup snapshots in this batch
            RollupSnapshot.query.filter_by(snapshot_batch_id=batch_id).update({"is_public": make_public})

            success_count += 1

        db.session.commit()

        return jsonify(
            {
                "success": True,
                "success_count": success_count,
                "error_count": error_count,
                "message": f'Updated {success_count} snapshot(s) to {"public" if make_public else "private"}',
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/snapshots/bulk/delete", methods=["POST"])
@login_required
@organization_required
def bulk_delete_snapshots():
    """Delete multiple snapshot batches"""
    try:
        data = request.get_json()
        batch_ids = data.get("batch_ids", [])

        if not batch_ids:
            return jsonify({"error": "No batch IDs provided"}), 400

        success_count = 0
        error_count = 0

        for batch_id in batch_ids:
            # Get one snapshot from this batch to check ownership
            sample_snapshot = KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).first()

            if not sample_snapshot:
                error_count += 1
                continue

            # Check ownership
            if sample_snapshot.owner_user_id != current_user.id:
                error_count += 1
                continue

            # Delete all KPI snapshots in this batch
            KPISnapshot.query.filter_by(snapshot_batch_id=batch_id).delete()

            # Delete all rollup snapshots in this batch
            RollupSnapshot.query.filter_by(snapshot_batch_id=batch_id).delete()

            success_count += 1

        db.session.commit()

        return jsonify(
            {
                "success": True,
                "success_count": success_count,
                "error_count": error_count,
                "message": f"Deleted {success_count} snapshot batch(es)",
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
        return jsonify({"error": str(e)}), 500
        print(f"[DEBUG] Error toggling privacy: {str(e)}")
        return jsonify({"error": str(e)}), 500


@bp.route("/api/kpi/<int:config_id>/trend")
@login_required
@organization_required
def get_kpi_trend(config_id):
    """
    Get trend information for a KPI.

    Returns: {'direction': 'up'|'down'|'stable', 'change': value, 'percent_change': percent}
    """
    try:
        trend = SnapshotService.calculate_trend(config_id, periods=2)

        if trend is None:
            return jsonify({"error": "Insufficient historical data"}), 404

        return jsonify(trend)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/kpi/<int:config_id>/history")
@login_required
@organization_required
def get_kpi_history(config_id):
    """
    Get historical snapshots for a KPI.

    Returns array of snapshots with dates and values.
    """
    try:
        limit = request.args.get("limit", 50, type=int)
        snapshots = SnapshotService.get_kpi_history(config_id, limit=limit)

        # Get the config and current consensus value
        config = KPIValueTypeConfig.query.get_or_404(config_id)
        consensus = config.get_consensus_value()  # Use model method that handles formula/linked/manual

        # Format for chart: array of {date, value} objects
        # Reverse so oldest is first (better for chart display)
        # Use created_at timestamp to distinguish snapshots on the same day
        history = []
        for snapshot in reversed(snapshots):
            value = snapshot.get_value()
            if value is not None:  # Only include snapshots with actual values
                # Use full timestamp for snapshots on same day
                date_label = snapshot.created_at.strftime("%Y-%m-%d %H:%M:%S")
                history.append({"date": date_label, "value": float(value), "label": snapshot.snapshot_label})

        # Add current value as the latest point (if it exists and differs from last snapshot)
        if consensus and consensus.get("status") != "no_data":
            current_value = consensus.get("value")
            if current_value is not None:
                # Use current timestamp for current value
                from datetime import datetime

                current_time = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
                history.append({"date": current_time, "value": float(current_value), "label": "Current"})

        return jsonify({"history": history, "count": len(history)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================================
# COMMENTS & COLLABORATION ROUTES (@mentions support)
# ============================================================================


@bp.route("/api/cell/<int:config_id>/comments", methods=["GET"])
@login_required
@organization_required
def get_cell_comments(config_id):
    """Get all comments for a KPI cell"""
    try:
        org_id = session.get("organization_id")

        # Check permission to view comments
        if not current_user.can_view_comments(org_id):
            return jsonify({"error": "You do not have permission to view comments"}), 403

        include_resolved = request.args.get("include_resolved", "true").lower() == "true"
        comments = CommentService.get_comments_for_cell(config_id, include_resolved=include_resolved)

        def render_comment_tree(comment):
            """Recursively render comment with replies"""
            result = {
                **comment.to_dict(),
                "rendered_text": CommentService.render_comment_with_mentions(
                    comment.comment_text, org_id, comment_id=comment.id
                ),
                "replies": [],
            }

            # Add replies
            for reply in comment.replies:
                result["replies"].append(render_comment_tree(reply))

            return result

        return jsonify(
            {
                "comments": [c.to_dict() for c in comments],
                "count": len(comments),
                "rendered_comments": [render_comment_tree(c) for c in comments],
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/cell/<int:config_id>/comments", methods=["POST"])
@login_required
@organization_required
def create_cell_comment(config_id):
    """Create a new comment on a KPI cell"""
    try:
        org_id = session.get("organization_id")

        # Check permission to add comments
        if not current_user.can_add_comments(org_id):
            return jsonify({"error": "You do not have permission to add comments"}), 403

        data = request.get_json()
        comment_text = data.get("comment_text", "").strip()
        parent_comment_id = data.get("parent_comment_id")

        if not comment_text:
            return jsonify({"error": "Comment text is required"}), 400

        comment = CommentService.create_comment(
            config_id=config_id,
            user_id=current_user.id,
            comment_text=comment_text,
            parent_comment_id=parent_comment_id,
            organization_id=org_id,
        )

        return jsonify(
            {
                "success": True,
                "comment": comment.to_dict(),
                "rendered_text": CommentService.render_comment_with_mentions(
                    comment.comment_text, org_id, comment_id=comment.id
                ),
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/comments/<int:comment_id>", methods=["PUT"])
@login_required
@organization_required
def update_cell_comment(comment_id):
    """Update an existing comment"""
    try:
        comment = CellComment.query.get(comment_id)
        if not comment:
            return jsonify({"error": "Comment not found"}), 404

        # Check ownership
        if comment.user_id != current_user.id:
            return jsonify({"error": "Unauthorized"}), 403

        data = request.get_json()
        comment_text = data.get("comment_text", "").strip()

        if not comment_text:
            return jsonify({"error": "Comment text is required"}), 400

        org_id = session.get("organization_id")

        updated_comment = CommentService.update_comment(
            comment_id=comment_id, comment_text=comment_text, organization_id=org_id
        )

        return jsonify(
            {
                "success": True,
                "comment": updated_comment.to_dict(),
                "rendered_text": CommentService.render_comment_with_mentions(
                    updated_comment.comment_text, org_id, comment_id=updated_comment.id
                ),
            }
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/comments/<int:comment_id>", methods=["DELETE"])
@login_required
@organization_required
def delete_cell_comment(comment_id):
    """Delete a comment"""
    try:
        comment = CellComment.query.get(comment_id)
        if not comment:
            return jsonify({"error": "Comment not found"}), 404

        # Check ownership or admin
        if comment.user_id != current_user.id and not current_user.is_global_admin:
            return jsonify({"error": "Unauthorized"}), 403

        success = CommentService.delete_comment(comment_id)

        return jsonify({"success": success})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/comments/<int:comment_id>/resolve", methods=["POST"])
@login_required
@organization_required
def resolve_comment(comment_id):
    """Mark a comment as resolved"""
    try:
        comment = CommentService.resolve_comment(comment_id, current_user.id)

        return jsonify({"success": True, "comment": comment.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/comments/<int:comment_id>/unresolve", methods=["POST"])
@login_required
@organization_required
def unresolve_comment(comment_id):
    """Mark a comment as unresolved"""
    try:
        comment = CommentService.unresolve_comment(comment_id)

        return jsonify({"success": True, "comment": comment.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/mentions/unread")
@login_required
@organization_required
def get_unread_mentions():
    """Get unread mentions for current user"""
    try:
        limit = request.args.get("limit", 20, type=int)
        mentions = CommentService.get_unread_mentions(current_user.id, limit=limit)
        total_count = CommentService.get_unread_mentions_count(current_user.id)

        return jsonify(
            {
                "mentions": [m.to_dict() for m in mentions],
                "count": len(mentions),  # Number of mentions returned (limited)
                "total_count": total_count,  # Total unread mentions count
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/mentions/<int:notification_id>/read", methods=["POST"])
@login_required
@organization_required
def mark_mention_read(notification_id):
    """Mark a mention as read"""
    try:
        notification = CommentService.mark_mention_read(notification_id)

        return jsonify({"success": True, "notification": notification.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/mentions/mark-all-read", methods=["POST"])
@login_required
@organization_required
def mark_all_mentions_read():
    """Mark all mentions as read for current user"""
    try:
        count = CommentService.mark_all_mentions_read(current_user.id)

        return jsonify({"success": True, "count": count})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/my-comments")
@login_required
@organization_required
def get_my_comments():
    """Get comments authored by current user"""
    try:
        org_id = session.get("organization_id")
        limit = request.args.get("limit", 50, type=int)

        # Get comments authored by current user (join through config -> kpi to filter by organization)
        comments = (
            CellComment.query.filter_by(user_id=current_user.id)
            .join(KPIValueTypeConfig, CellComment.kpi_value_type_config_id == KPIValueTypeConfig.id)
            .join(KPI, KPIValueTypeConfig.kpi_id == KPI.id)
            .join(InitiativeSystemLink, KPI.initiative_system_link_id == InitiativeSystemLink.id)
            .join(Initiative, InitiativeSystemLink.initiative_id == Initiative.id)
            .filter(Initiative.organization_id == org_id)
            .order_by(CellComment.created_at.desc())
            .limit(limit)
            .all()
        )

        def comment_with_context(comment):
            """Add KPI and cell context to comment"""
            config = comment.config
            kpi = config.kpi if config else None
            value_type = config.value_type if config else None

            return {
                **comment.to_dict(),
                "rendered_text": CommentService.render_comment_with_mentions(
                    comment.comment_text, org_id, comment_id=comment.id
                ),
                "kpi_name": kpi.name if kpi else "Unknown KPI",
                "value_type_name": value_type.name if value_type else "Unknown Type",
                "config_id": config.id if config else None,
                "reply_to": comment.parent.user.display_name if comment.parent else None,
            }

        return jsonify({"comments": [comment_with_context(c) for c in comments], "count": len(comments)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/org/users/search")
@login_required
@organization_required
def search_org_users():
    """
    Search users in current organization for @mention autocomplete.

    Query param: q=search_term
    """
    try:
        org_id = session.get("organization_id")
        search_term = request.args.get("q", "").strip().lower()

        # Build query
        query = (
            db.session.query(User)
            .join(UserOrganizationMembership)
            .filter(UserOrganizationMembership.organization_id == org_id)
        )

        # Filter by search term if provided
        if search_term:
            query = query.filter(
                db.or_(User.login.ilike(f"%{search_term}%"), User.display_name.ilike(f"%{search_term}%"))
            )

        # Get results (limit 10)
        users = query.order_by(User.display_name).limit(10).all()

        return jsonify(
            {"users": [{"id": u.id, "login": u.login, "display_name": u.display_name, "email": u.email} for u in users]}
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/mentions/search")
@login_required
@organization_required
def search_mentions():
    """
    Search for users and entities to mention (unified autocomplete).

    Query param: q=search_term
    Returns: {"users": [...], "entities": [...]}
    """
    from app.services.mention_service import MentionService

    try:
        org_id = session.get("organization_id")
        search_term = request.args.get("q", "").strip()
        limit = request.args.get("limit", 10, type=int)

        results = MentionService.search_all(search_term, org_id, limit)

        return jsonify(results)

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/search")
@login_required
@organization_required
def search_page():
    """Search results page - Enhanced with SearchService (fuzzy matching, modifiers, filters)"""
    org_id = session.get("organization_id")
    org_name = session.get("organization_name")
    query = request.args.get("q", "").strip()

    # Read filter parameters from URL
    entity_types_param = request.args.get("entity_types", "")
    date_range = request.args.get("date_range", "")
    status_param = request.args.get("status", "")
    exact_match = request.args.get("exact", "") == "1"

    # Parse filters for SearchService and template
    filters = {}
    entity_types_list = []
    status_list = []

    if entity_types_param:
        entity_types_list = entity_types_param.split(",")
        # Only add entity_types filter if it's not "all types" (less than 6 means actual filtering)
        if len(entity_types_list) < 6:
            filters["entity_types"] = entity_types_list
    if date_range:
        filters["date_range"] = date_range
    if status_param:
        status_list = status_param.split(",")
        filters["status"] = status_list

    if exact_match:
        filters["exact"] = True

    if not query:
        return render_template(
            "workspace/search.html",
            organization_name=org_name,
            query="",
            results={},
            total=0,
            entity_types=entity_types_list,
            date_range=date_range,
            status=status_list,
            exact_match=exact_match,
            csrf_token=generate_csrf,
        )

    # Use SearchService for fuzzy matching, modifiers, and filters
    # This gives the same powerful search experience as the navbar
    search_results = SearchService.search_all(query, filters=filters, organization_id=org_id)

    # Transform SearchService results to match template expectations
    # SearchService returns more detailed results (match_score, updated_at, etc.)
    results = {
        "spaces": search_results.get("spaces", []),
        "challenges": search_results.get("challenges", []),
        "initiatives": search_results.get("initiatives", []),
        "systems": search_results.get("systems", []),
        "kpis": search_results.get("kpis", []),
        "value_types": search_results.get("value_types", []),
        "comments": search_results.get("comments", []),
        "action_items": search_results.get("action_items", []),
        "entity_links": search_results.get("entity_links", []),
    }

    # Add space_name to challenges if missing (template expects it)
    for challenge in results["challenges"]:
        if "space" not in challenge and "space_name" in challenge:
            challenge["space"] = challenge["space_name"]

    # Add initiative/system names to KPIs if missing
    for kpi in results["kpis"]:
        if "initiative" not in kpi and "initiative_name" in kpi:
            kpi["initiative"] = kpi["initiative_name"]
        if "system" not in kpi and "system_name" in kpi:
            kpi["system"] = kpi["system_name"]

    # Count totals
    total = sum(len(results[key]) for key in results)

    # Get entity type defaults (for logos)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"
        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    return render_template(
        "workspace/search.html",
        organization_name=org_name,
        query=query,
        results=results,
        total=total,
        entity_defaults=entity_defaults,
        entity_types=entity_types_list,
        date_range=date_range,
        status=status_list,
        exact_match=exact_match,
        csrf_token=generate_csrf,
    )


@bp.route("/api/search/live")
@login_required
@organization_required
def live_search():
    """Live search API endpoint - returns JSON results as user types"""
    org_id = session.get("organization_id")
    query = request.args.get("q", "").strip()

    if not query or len(query) < 2:
        return jsonify({"results": []})

    # Get entity type defaults (for icons/logos)
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"
        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    search_pattern = f"%{query}%"
    results = []

    # Limit to top 5 of each type for quick display
    limit = 5

    # Search Spaces
    spaces = (
        Space.query.filter(
            Space.organization_id == org_id,
            db.or_(Space.name.ilike(search_pattern), Space.description.ilike(search_pattern)),
        )
        .limit(limit)
        .all()
    )

    for s in spaces:
        results.append(
            {
                "type": "space",
                "id": s.id,
                "name": s.name,
                "description": s.description[:100] if s.description else None,
                "url": url_for("workspace.index", _anchor=f"space-{s.id}"),
                "edit_url": url_for("organization_admin.edit_space", space_id=s.id),
                "icon": entity_defaults.get("space", {}).get("icon", "🏢"),
                "logo": entity_defaults.get("space", {}).get("logo"),
            }
        )

    # Search Challenges
    challenges = (
        Challenge.query.join(Space)
        .filter(
            Space.organization_id == org_id,
            db.or_(Challenge.name.ilike(search_pattern), Challenge.description.ilike(search_pattern)),
        )
        .limit(limit)
        .all()
    )

    for c in challenges:
        results.append(
            {
                "type": "challenge",
                "id": c.id,
                "name": c.name,
                "description": c.description[:100] if c.description else None,
                "space": c.space.name,
                "url": url_for("workspace.index", _anchor=f"challenge-{c.id}"),
                "edit_url": url_for("organization_admin.edit_challenge", challenge_id=c.id),
                "icon": entity_defaults.get("challenge", {}).get("icon", "ƒ"),
                "logo": entity_defaults.get("challenge", {}).get("logo"),
            }
        )

    # Search Initiatives
    initiatives = (
        Initiative.query.filter(
            Initiative.organization_id == org_id,
            db.or_(Initiative.name.ilike(search_pattern), Initiative.description.ilike(search_pattern)),
        )
        .limit(limit)
        .all()
    )

    for i in initiatives:
        results.append(
            {
                "type": "initiative",
                "id": i.id,
                "name": i.name,
                "description": i.description[:100] if i.description else None,
                "url": url_for("workspace.index", _anchor=f"initiative-{i.id}"),
                "edit_url": url_for("organization_admin.edit_initiative", initiative_id=i.id),
                "icon": entity_defaults.get("initiative", {}).get("icon", "δ"),
                "logo": entity_defaults.get("initiative", {}).get("logo"),
            }
        )

    # Search Systems
    systems = (
        System.query.filter(
            System.organization_id == org_id,
            db.or_(System.name.ilike(search_pattern), System.description.ilike(search_pattern)),
        )
        .limit(limit)
        .all()
    )

    for s in systems:
        results.append(
            {
                "type": "system",
                "id": s.id,
                "name": s.name,
                "description": s.description[:100] if s.description else None,
                "url": url_for("workspace.index", _anchor=f"system-{s.id}"),
                "edit_url": url_for("organization_admin.edit_system", system_id=s.id),
                "icon": entity_defaults.get("system", {}).get("icon", "Φ"),
                "logo": entity_defaults.get("system", {}).get("logo"),
            }
        )

    # Search KPIs
    kpis = (
        db.session.query(KPI)
        .join(InitiativeSystemLink)
        .join(Initiative)
        .filter(
            Initiative.organization_id == org_id,
            db.or_(KPI.name.ilike(search_pattern), KPI.description.ilike(search_pattern)),
        )
        .limit(limit)
        .all()
    )

    for k in kpis:
        results.append(
            {
                "type": "kpi",
                "id": k.id,
                "name": k.name,
                "description": k.description[:100] if k.description else None,
                "initiative": k.initiative_system_link.initiative.name if k.initiative_system_link else None,
                "system": k.initiative_system_link.system.name if k.initiative_system_link else None,
                "url": url_for("workspace.index", _anchor=f"kpi-{k.id}"),
                "edit_url": url_for("organization_admin.edit_kpi", kpi_id=k.id),
                "icon": entity_defaults.get("kpi", {}).get("icon", "Ψ"),
                "logo": entity_defaults.get("kpi", {}).get("logo"),
            }
        )

    return jsonify({"results": results, "total": len(results), "limit_per_type": limit})


@bp.route("/api/search/advanced", methods=["POST"])
@login_required
@organization_required
def advanced_search():
    """
    Advanced search API endpoint with fuzzy matching, query parsing, and filters.

    Accepts JSON POST body:
    {
        "query": "search text",
        "filters": {
            "entity_types": ["kpis", "systems", "initiatives", "challenges", "spaces"],
            "date_range": "last_week",
            "status": ["at_risk", "incomplete"],
            ... (additional filters)
        }
    }

    Returns JSON with categorized results and query info.
    """
    org_id = session.get("organization_id")
    data = request.get_json()

    if not data or "query" not in data:
        return jsonify({"error": "Query parameter required"}), 400

    query = data.get("query", "").strip()
    filters = data.get("filters", {})

    if not query or len(query) < 2:
        return jsonify(SearchService._empty_results())

    # Execute search using SearchService
    results = SearchService.search_all(query, filters, org_id)

    # Enhance results with URLs and entity defaults
    entity_defaults_raw = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    entity_defaults = {}
    for default in entity_defaults_raw:
        logo_url = None
        if default.default_logo_data and default.default_logo_mime_type:
            logo_url = f"data:{default.default_logo_mime_type};base64,{base64.b64encode(default.default_logo_data).decode('utf-8')}"
        entity_defaults[default.entity_type] = {
            "color": default.default_color,
            "icon": default.default_icon,
            "logo": logo_url,
        }

    # Add URLs and icons to results
    for kpi_result in results.get("kpis", []):
        kpi_result["url"] = url_for("workspace.index", _anchor=f"kpi-{kpi_result['id']}")
        kpi_result["edit_url"] = url_for("organization_admin.edit_kpi", kpi_id=kpi_result["id"])
        kpi_result["icon"] = entity_defaults.get("kpi", {}).get("icon", "Ψ")
        kpi_result["logo"] = entity_defaults.get("kpi", {}).get("logo")

    for system_result in results.get("systems", []):
        system_result["url"] = url_for("workspace.index", _anchor=f"system-{system_result['id']}")
        system_result["edit_url"] = url_for("organization_admin.edit_system", system_id=system_result["id"])
        system_result["icon"] = entity_defaults.get("system", {}).get("icon", "Φ")
        system_result["logo"] = entity_defaults.get("system", {}).get("logo")

    for initiative_result in results.get("initiatives", []):
        initiative_result["url"] = url_for("workspace.index", _anchor=f"initiative-{initiative_result['id']}")
        initiative_result["edit_url"] = url_for(
            "organization_admin.edit_initiative", initiative_id=initiative_result["id"]
        )
        initiative_result["icon"] = entity_defaults.get("initiative", {}).get("icon", "δ")
        initiative_result["logo"] = entity_defaults.get("initiative", {}).get("logo")

    for challenge_result in results.get("challenges", []):
        challenge_result["url"] = url_for("workspace.index", _anchor=f"challenge-{challenge_result['id']}")
        challenge_result["edit_url"] = url_for("organization_admin.edit_challenge", challenge_id=challenge_result["id"])
        challenge_result["icon"] = entity_defaults.get("challenge", {}).get("icon", "ƒ")
        challenge_result["logo"] = entity_defaults.get("challenge", {}).get("logo")

    for space_result in results.get("spaces", []):
        space_result["url"] = url_for("workspace.index", _anchor=f"space-{space_result['id']}")
        space_result["edit_url"] = url_for("organization_admin.edit_space", space_id=space_result["id"])
        space_result["icon"] = entity_defaults.get("space", {}).get("icon", "🏢")
        space_result["logo"] = entity_defaults.get("space", {}).get("logo")

    for ai_result in results.get("action_items", []):
        ai_result["url"] = url_for("action_items.view", item_id=ai_result["id"])
        ai_result["edit_url"] = url_for("action_items.edit", item_id=ai_result["id"])
        ai_result["icon"] = "bi-check2-square"
        ai_result["logo"] = None

    # Calculate totals
    total_results = (
        len(results.get("kpis", []))
        + len(results.get("systems", []))
        + len(results.get("initiatives", []))
        + len(results.get("challenges", []))
        + len(results.get("spaces", []))
        + len(results.get("entity_links", []))
        + len(results.get("action_items", []))
    )

    results["total_results"] = total_results

    return jsonify(results)


# ============================================================================
# Saved Searches API Endpoints
# ============================================================================


@bp.route("/api/saved-searches", methods=["GET"])
@login_required
@organization_required
def get_saved_searches():
    """
    Get all saved searches for the current user in the current organization.

    Returns:
        JSON array of saved search objects
    """
    org_id = session.get("organization_id")
    user_id = current_user.id

    searches = SavedSearch.get_user_searches(user_id, org_id)
    return jsonify({"searches": [s.to_dict() for s in searches]})


@bp.route("/api/saved-searches", methods=["POST"])
@login_required
@organization_required
def create_saved_search():
    """
    Create a new saved search.

    Expects JSON body:
    {
        "name": "Search name",
        "query": "search text",
        "filters": {"entity_types": [...], ...},
        "is_default": false
    }

    Returns:
        JSON with created search object
    """
    org_id = session.get("organization_id")
    user_id = current_user.id
    data = request.get_json()

    if not data or "name" not in data or "query" not in data:
        return jsonify({"error": "Name and query are required"}), 400

    # Validate name length
    if len(data["name"]) > 200:
        return jsonify({"error": "Name must be 200 characters or less"}), 400

    # Check for duplicate names
    existing = (
        db.session.query(SavedSearch).filter_by(user_id=user_id, organization_id=org_id, name=data["name"]).first()
    )
    if existing:
        return jsonify({"error": "A saved search with this name already exists"}), 400

    # Create new saved search
    saved_search = SavedSearch(
        user_id=user_id,
        organization_id=org_id,
        name=data["name"],
        query=data["query"],
        filters=data.get("filters"),
        is_default=data.get("is_default", False),
    )

    db.session.add(saved_search)
    db.session.commit()

    return jsonify({"search": saved_search.to_dict()}), 201


@bp.route("/api/saved-searches/<int:search_id>", methods=["GET"])
@login_required
@organization_required
def get_saved_search(search_id):
    """
    Get a specific saved search by ID.

    Returns:
        JSON with search object or 404 if not found
    """
    org_id = session.get("organization_id")
    user_id = current_user.id

    saved_search = (
        db.session.query(SavedSearch).filter_by(id=search_id, user_id=user_id, organization_id=org_id).first()
    )

    if not saved_search:
        return jsonify({"error": "Saved search not found"}), 404

    return jsonify({"search": saved_search.to_dict()})


@bp.route("/api/saved-searches/<int:search_id>", methods=["PUT"])
@login_required
@organization_required
def update_saved_search(search_id):
    """
    Update a saved search.

    Expects JSON body with fields to update:
    {
        "name": "New name",
        "query": "new search text",
        "filters": {...},
        "is_default": true
    }

    Returns:
        JSON with updated search object
    """
    org_id = session.get("organization_id")
    user_id = current_user.id
    data = request.get_json()

    if not data:
        return jsonify({"error": "No data provided"}), 400

    # Find the saved search
    saved_search = (
        db.session.query(SavedSearch).filter_by(id=search_id, user_id=user_id, organization_id=org_id).first()
    )

    if not saved_search:
        return jsonify({"error": "Saved search not found"}), 404

    # Update fields
    if "name" in data:
        if len(data["name"]) > 200:
            return jsonify({"error": "Name must be 200 characters or less"}), 400

        # Check for duplicate names (excluding current search)
        existing = (
            db.session.query(SavedSearch)
            .filter_by(user_id=user_id, organization_id=org_id, name=data["name"])
            .filter(SavedSearch.id != search_id)
            .first()
        )
        if existing:
            return jsonify({"error": "A saved search with this name already exists"}), 400

        saved_search.name = data["name"]

    if "query" in data:
        saved_search.query = data["query"]

    if "filters" in data:
        saved_search.filters = data["filters"]

    if "is_default" in data and data["is_default"]:
        saved_search.set_as_default()
    elif "is_default" in data and not data["is_default"]:
        saved_search.is_default = False

    db.session.commit()

    return jsonify({"search": saved_search.to_dict()})


@bp.route("/api/saved-searches/<int:search_id>", methods=["DELETE"])
@login_required
@organization_required
def delete_saved_search(search_id):
    """
    Delete a saved search.

    Returns:
        JSON success message or 404 if not found
    """
    org_id = session.get("organization_id")
    user_id = current_user.id

    saved_search = (
        db.session.query(SavedSearch).filter_by(id=search_id, user_id=user_id, organization_id=org_id).first()
    )

    if not saved_search:
        return jsonify({"error": "Saved search not found"}), 404

    db.session.delete(saved_search)
    db.session.commit()

    return jsonify({"message": "Saved search deleted successfully"})


@bp.route("/api/saved-searches/<int:search_id>/set-default", methods=["POST"])
@login_required
@organization_required
def set_default_search(search_id):
    """
    Set a saved search as the default.

    Returns:
        JSON success message or 404 if not found
    """
    org_id = session.get("organization_id")
    user_id = current_user.id

    saved_search = (
        db.session.query(SavedSearch).filter_by(id=search_id, user_id=user_id, organization_id=org_id).first()
    )

    if not saved_search:
        return jsonify({"error": "Saved search not found"}), 404

    saved_search.set_as_default()

    return jsonify({"message": "Default search set successfully", "search": saved_search.to_dict()})


@bp.route("/api/kpi/<int:kpi_id>/status")
@login_required
@organization_required
def get_kpi_status(kpi_id):
    """
    Get traffic light status for a KPI.

    Returns JSON with status (green/yellow/red), reason, and details.
    """
    try:
        # Verify KPI belongs to current organization
        kpi = (
            KPI.query.join(InitiativeSystemLink)
            .join(System)
            .filter(KPI.id == kpi_id, System.organization_id == session.get("organization_id"))
            .first_or_404()
        )

        status_data = kpi.get_status()

        return jsonify(
            {
                "kpi_id": kpi.id,
                "kpi_name": kpi.name,
                "status": status_data["status"],
                "reason": status_data["reason"],
                "details": status_data["details"],
            }
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/kpis/statuses")
@login_required
@organization_required
def get_all_kpi_statuses():
    """
    Get traffic light statuses for all KPIs in the organization.

    Returns JSON with array of {kpi_id, kpi_name, status, reason, details}.
    Useful for dashboard displays.
    """
    try:
        org_id = session.get("organization_id")

        # Get all KPIs for this organization
        kpis = (
            KPI.query.join(InitiativeSystemLink)
            .join(System)
            .filter(System.organization_id == org_id)
            .order_by(KPI.name)
            .all()
        )

        results = []
        for kpi in kpis:
            status_data = kpi.get_status()
            results.append(
                {
                    "kpi_id": kpi.id,
                    "kpi_name": kpi.name,
                    "status": status_data["status"],
                    "reason": status_data["reason"],
                    "details": status_data["details"],
                    "is_archived": kpi.is_archived,
                }
            )

        # Count by status
        status_counts = {"green": 0, "yellow": 0, "red": 0}
        for result in results:
            status_counts[result["status"]] += 1

        return jsonify({"kpis": results, "summary": status_counts, "total": len(results)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@bp.route("/api/value-types/reorder", methods=["POST"])
@login_required
@organization_required
def reorder_value_types():
    """Update display order of value types via drag-and-drop"""
    try:
        org_id = session.get("organization_id")
        data = request.get_json()

        if not data or "value_type_ids" not in data:
            return jsonify({"error": "Missing value_type_ids"}), 400

        value_type_ids = data["value_type_ids"]

        # Update display_order for each value type
        for index, vt_id in enumerate(value_type_ids):
            vt = ValueType.query.filter_by(id=vt_id, organization_id=org_id).first()
            if vt:
                vt.display_order = index

        db.session.commit()

        return jsonify({"success": True, "message": "Value type order updated"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@bp.route("/api/reorder/<entity_type>", methods=["POST"])
@login_required
@organization_required
def reorder_rows(entity_type):
    """Update display order of rows (spaces, challenges, initiatives, systems, KPIs) via drag-and-drop"""
    try:
        org_id = session.get("organization_id")
        data = request.get_json()

        if not data or "ids" not in data:
            return jsonify({"error": "Missing ids"}), 400

        ids = data["ids"]
        parent_id = data.get("parent_id")

        # Route to appropriate handler based on entity type
        if entity_type == "space":
            _reorder_spaces(org_id, ids)
        elif entity_type == "challenge":
            _reorder_challenges(org_id, ids, parent_id)
        elif entity_type == "initiative":
            _reorder_initiatives(org_id, ids, parent_id)
        elif entity_type == "system":
            _reorder_systems(org_id, ids, parent_id)
        elif entity_type == "kpi":
            _reorder_kpis(org_id, ids, parent_id)
        else:
            return jsonify({"error": f"Unknown entity type: {entity_type}"}), 400

        db.session.commit()
        return jsonify({"success": True, "message": f"{entity_type.capitalize()} order updated"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


def _reorder_spaces(org_id, space_ids):
    """Reorder spaces within an organization"""
    from app.models import Space

    for index, space_id in enumerate(space_ids):
        space = Space.query.filter_by(id=space_id, organization_id=org_id).first()
        if space:
            space.display_order = index


def _reorder_challenges(org_id, challenge_ids, parent_space_id):
    """Reorder challenges within a space"""
    from app.models import Challenge

    for index, challenge_id in enumerate(challenge_ids):
        challenge = Challenge.query.filter_by(id=challenge_id, space_id=parent_space_id, organization_id=org_id).first()
        if challenge:
            challenge.display_order = index


def _reorder_initiatives(org_id, link_ids, parent_challenge_id):
    """Reorder initiatives within a challenge (by updating ChallengeInitiativeLink)"""
    from app.models import ChallengeInitiativeLink

    for index, link_id in enumerate(link_ids):
        link = ChallengeInitiativeLink.query.filter_by(id=link_id, challenge_id=parent_challenge_id).first()
        if link:
            link.display_order = index


def _reorder_systems(org_id, link_ids, parent_initiative_id):
    """Reorder systems within an initiative (by updating InitiativeSystemLink)"""
    from app.models import InitiativeSystemLink

    for index, link_id in enumerate(link_ids):
        link = InitiativeSystemLink.query.filter_by(id=link_id, initiative_id=parent_initiative_id).first()
        if link:
            link.display_order = index


def _reorder_kpis(org_id, kpi_ids, parent_link_id):
    """Reorder KPIs within a system (within an InitiativeSystemLink)"""
    from app.models import KPI

    for index, kpi_id in enumerate(kpi_ids):
        kpi = KPI.query.filter_by(id=kpi_id, initiative_system_link_id=parent_link_id).first()
        if kpi:
            kpi.display_order = index


@bp.route("/api/linked-kpi/organizations")
@login_required
def api_get_organizations_for_linking():
    """Get list of organizations user has access to for linking KPIs.

    Optionally filters to only organizations that have KPIs with specific value type kind.

    Query params:
        kind: Value type kind to filter by (numeric, sentiment, risk, etc.)
    """
    # Get optional kind filter
    required_kind = request.args.get("kind")

    # Get all orgs where user has membership
    user_org_ids = [m.organization_id for m in current_user.organization_memberships]

    # If kind filter specified, only get orgs that have KPIs with that kind
    if required_kind:
        # Query for org IDs that have at least one KPI with the required kind
        org_ids_with_kind = (
            db.session.query(Initiative.organization_id)
            .join(InitiativeSystemLink)
            .join(KPI)
            .join(KPIValueTypeConfig, KPIValueTypeConfig.kpi_id == KPI.id)
            .join(ValueType, ValueType.id == KPIValueTypeConfig.value_type_id)
            .filter(
                Initiative.organization_id.in_(user_org_ids),
                KPI.is_archived.is_(False),
                (ValueType.kind == "numeric" if required_kind == "numeric" else ValueType.kind == required_kind),
            )
            .distinct()
            .all()
        )
        # Extract org IDs from query result tuples
        filtered_org_ids = [org_id for (org_id,) in org_ids_with_kind]
        orgs = Organization.query.filter(Organization.id.in_(filtered_org_ids)).filter_by(is_deleted=False).all()
    else:
        orgs = Organization.query.filter(Organization.id.in_(user_org_ids)).filter_by(is_deleted=False).all()

    return jsonify(
        [
            {
                "id": org.id,
                "name": org.name,
            }
            for org in orgs
        ]
    )


@bp.route("/api/linked-kpi/kpis/<int:org_id>")
@login_required
def api_get_kpis_for_linking(org_id):
    """Get list of KPIs from an organization for linking"""
    # Verify user has access to this org
    has_access = any(m.organization_id == org_id for m in current_user.organization_memberships)
    if not has_access:
        return jsonify({"error": "Access denied"}), 403

    # Get all KPIs from this org with full hierarchy context (not archived)
    kpis = (
        db.session.query(KPI)
        .join(InitiativeSystemLink)
        .join(Initiative)
        .filter(Initiative.organization_id == org_id)
        .filter(~KPI.is_archived)
        .join(ChallengeInitiativeLink)
        .join(Challenge)
        .join(Space)
        .all()
    )

    result = []
    for kpi in kpis:
        is_link = kpi.initiative_system_link
        initiative = is_link.initiative
        system = is_link.system

        # Get challenges
        challenges = [ci_link.challenge.name for ci_link in initiative.challenge_links]
        # Get space
        spaces = list(set([ci_link.challenge.space.name for ci_link in initiative.challenge_links]))

        result.append(
            {
                "id": kpi.id,
                "name": kpi.name,
                "system": system.name,
                "initiative": initiative.name,
                "challenges": ", ".join(challenges),
                "spaces": ", ".join(spaces),
                "full_path": f"{', '.join(spaces)} → {', '.join(challenges)} → {initiative.name} → {system.name} → {kpi.name}",
            }
        )

    return jsonify(result)


@bp.route("/api/linked-kpi/value-types/<int:kpi_id>")
@login_required
def api_get_value_types_for_linking(kpi_id):
    """Get list of value types from a KPI for linking"""
    kpi = KPI.query.get_or_404(kpi_id)

    # Verify user has access to this KPI's org
    org_id = kpi.initiative_system_link.initiative.organization_id
    has_access = any(m.organization_id == org_id for m in current_user.organization_memberships)
    if not has_access:
        return jsonify({"error": "Access denied"}), 403

    result = []
    for config in kpi.value_type_configs:
        vt = config.value_type
        result.append(
            {
                "id": vt.id,
                "name": vt.name,
                "kind": vt.kind,
                "unit_label": vt.unit_label,
                "display": f"{vt.name}" + (f" ({vt.unit_label})" if vt.unit_label else ""),
            }
        )

    return jsonify(result)


# =============================================================================
# FILTER PRESET API ENDPOINTS
# =============================================================================


@bp.route("/api/preferences", methods=["POST"])
@login_required
@organization_required
def save_preferences():
    """Save a user preference key/value for the current org membership."""
    org_id = session.get("organization_id")
    data = request.get_json() or {}

    membership = UserOrganizationMembership.query.filter_by(
        user_id=current_user.id, organization_id=org_id
    ).first()
    if not membership:
        return jsonify({"error": "Membership not found"}), 404

    prefs = dict(membership.preferences or {})
    prefs.update(data)
    membership.preferences = prefs
    db.session.commit()
    return jsonify({"ok": True})


@bp.route("/api/impact", methods=["POST"])
@login_required
@organization_required
def update_impact():
    """Update impact_level for an entity (AJAX)."""
    from app.models import Challenge, KPI, Space, System

    data = request.get_json() or {}
    entity_type = data.get("entity_type")
    entity_id = data.get("entity_id")
    impact_level = data.get("impact_level")  # 1, 2, 3, or None

    org_id = session.get("organization_id")
    if not entity_type or not entity_id:
        return jsonify({"error": "Missing entity_type or entity_id"}), 400

    model_map = {"space": Space, "challenge": Challenge, "initiative": Initiative, "system": System, "kpi": KPI}
    model = model_map.get(entity_type)
    if not model:
        return jsonify({"error": "Invalid entity_type"}), 400

    entity = model.query.get(entity_id)
    if not entity:
        return jsonify({"error": "Entity not found"}), 404

    entity.impact_level = int(impact_level) if impact_level else None
    db.session.commit()
    try:
        localStorage_key = "ws_dirty"
    except Exception:
        pass
    return jsonify({"ok": True, "impact_level": entity.impact_level})


@bp.route("/api/filter-presets")
@login_required
@organization_required
def get_filter_presets():
    """Get all filter presets for current user in current organization"""
    org_id = session.get("organization_id")

    presets = (
        UserFilterPreset.query.filter_by(user_id=current_user.id, organization_id=org_id, feature="workspace")
        .order_by(UserFilterPreset.name)
        .all()
    )

    return jsonify([preset.to_dict() for preset in presets])


@bp.route("/api/filter-presets", methods=["POST"])
@login_required
@organization_required
def save_filter_preset():
    """Save a new filter preset"""
    org_id = session.get("organization_id")

    data = request.get_json()
    name = data.get("name", "").strip()
    filters = data.get("filters", {})

    if not name:
        return jsonify({"error": "Preset name is required"}), 400

    # Check if name already exists (within workspace feature)
    existing = UserFilterPreset.query.filter_by(
        user_id=current_user.id, organization_id=org_id, feature="workspace", name=name
    ).first()

    if existing:
        return jsonify({"error": f"A preset named '{name}' already exists"}), 400

    # Create new preset
    preset = UserFilterPreset(user_id=current_user.id, organization_id=org_id, feature="workspace", name=name, filters=filters)

    db.session.add(preset)
    db.session.flush()  # Flush to get the preset.id

    # Immediately set this as the last used preset
    membership = UserOrganizationMembership.query.filter_by(user_id=current_user.id, organization_id=org_id).first()
    if membership:
        membership.last_workspace_preset_id = preset.id

    db.session.commit()

    return jsonify(preset.to_dict()), 201


@bp.route("/api/filter-presets/<int:preset_id>", methods=["PUT"])
@login_required
@organization_required
def update_filter_preset(preset_id):
    """Update an existing filter preset"""
    org_id = session.get("organization_id")

    preset = UserFilterPreset.query.get(preset_id)
    if not preset:
        return jsonify({"error": "Preset not found"}), 404

    # Verify ownership
    if preset.user_id != current_user.id or preset.organization_id != org_id:
        return jsonify({"error": "Access denied"}), 403

    data = request.get_json()

    # Update name if provided
    if "name" in data:
        new_name = data["name"].strip()
        if not new_name:
            return jsonify({"error": "Preset name cannot be empty"}), 400

        # Check if new name conflicts with another preset (within same feature)
        existing = (
            UserFilterPreset.query.filter_by(
                user_id=current_user.id, organization_id=org_id, feature=preset.feature, name=new_name
            )
            .filter(UserFilterPreset.id != preset_id)
            .first()
        )

        if existing:
            return jsonify({"error": f"A preset named '{new_name}' already exists"}), 400

        preset.name = new_name

    # Update filters if provided
    if "filters" in data:
        preset.filters = data["filters"]

    db.session.commit()

    return jsonify(preset.to_dict())


@bp.route("/api/filter-presets/<int:preset_id>", methods=["DELETE"])
@login_required
@organization_required
def delete_filter_preset(preset_id):
    """Delete a filter preset"""
    org_id = session.get("organization_id")

    preset = UserFilterPreset.query.get(preset_id)
    if not preset:
        return jsonify({"error": "Preset not found"}), 404

    # Verify ownership
    if preset.user_id != current_user.id or preset.organization_id != org_id:
        return jsonify({"error": "Access denied"}), 403

    db.session.delete(preset)
    db.session.commit()

    return jsonify({"success": True, "message": f"Preset '{preset.name}' deleted"})


@bp.route("/api/filter-presets/<int:preset_id>/set-last-used", methods=["POST"])
@login_required
@organization_required
def set_last_used_filter_preset(preset_id):
    """Mark this preset as the last one used by this user in this organization"""
    org_id = session.get("organization_id")

    # Verify the preset exists and belongs to this user/org
    preset = UserFilterPreset.query.get(preset_id)
    if not preset:
        return jsonify({"error": "Preset not found"}), 404

    if preset.user_id != current_user.id or preset.organization_id != org_id:
        return jsonify({"error": "Access denied"}), 403

    # Update the user's membership to track this as the last used preset
    membership = UserOrganizationMembership.query.filter_by(user_id=current_user.id, organization_id=org_id).first()

    if membership:
        membership.last_workspace_preset_id = preset_id
        db.session.commit()
        return jsonify({"success": True, "preset_id": preset_id})
    else:
        return jsonify({"error": "Membership not found"}), 404


@bp.route("/api/clear-last-preset", methods=["POST"])
@login_required
@organization_required
def clear_last_preset():
    """Clear the last used preset for this user in this organization"""
    org_id = session.get("organization_id")

    membership = UserOrganizationMembership.query.filter_by(user_id=current_user.id, organization_id=org_id).first()

    if membership:
        membership.last_workspace_preset_id = None
        db.session.commit()
        return jsonify({"success": True})
    else:
        return jsonify({"error": "Membership not found"}), 404


@bp.route("/api/kpis-for-formula/<int:org_id>")
@login_required
@organization_required
def get_kpis_for_formula(org_id):
    """
    Get all available KPIs that can be used in formula calculations.
    Returns KPI configs with their current values from all accessible organizations.
    """
    # Verify access to organization
    if session.get("organization_id") != org_id:
        return jsonify({"error": "Access denied"}), 403

    from app.services import ConsensusService

    # Get all organizations the user has access to
    user_org_ids = [membership.organization_id for membership in current_user.organization_memberships]

    # Get all KPIs from ALL accessible organizations with their value type configs
    kpis = (
        db.session.query(KPI, KPIValueTypeConfig, ValueType, Initiative, System, Organization)
        .join(KPIValueTypeConfig, KPI.id == KPIValueTypeConfig.kpi_id)
        .join(ValueType, KPIValueTypeConfig.value_type_id == ValueType.id)
        .join(InitiativeSystemLink, KPI.initiative_system_link_id == InitiativeSystemLink.id)
        .join(Initiative, InitiativeSystemLink.initiative_id == Initiative.id)
        .join(System, InitiativeSystemLink.system_id == System.id)
        .join(Organization, Initiative.organization_id == Organization.id)
        .filter(Initiative.organization_id.in_(user_org_ids))
        .filter(KPI.is_archived.is_(False))
        .filter(ValueType.is_active.is_(True))
        .order_by(Organization.name, Initiative.name, System.name, KPI.name, ValueType.name)
        .all()
    )

    result = []
    for kpi, config, value_type, initiative, system, organization in kpis:
        # ONLY include numeric value types - formulas can't work with qualitative types
        if not value_type.is_numeric():
            continue

        # Get current consensus value
        consensus = ConsensusService.get_cell_value(config)
        current_value = consensus.get("value") if consensus else None

        # Format display value
        if current_value is not None:
            try:
                display_value = f"{float(current_value):.2f} {value_type.unit_label or ''}".strip()
            except (ValueError, TypeError):
                display_value = str(current_value)
        else:
            display_value = "—"

        # Highlight if from current org
        is_current_org = organization.id == org_id

        result.append(
            {
                "id": config.id,
                "name": f"{kpi.name} - {value_type.name}",
                "kpi_name": kpi.name,
                "value_type_name": value_type.name,
                "value_type_kind": value_type.kind,
                "organization_name": organization.name,
                "organization_id": organization.id,
                "is_current_org": is_current_org,
                "path": f"{initiative.name} › {system.name}",
                "currentValue": display_value,
                "icon": "💶" if value_type.unit_label in ["€", "$", "USD"] else "📊",
                "variable_name": kpi.get_variable_name(),  # Sanitized Python variable name
            }
        )

    return jsonify({"kpis": result})


@bp.route("/kpi-config/<int:config_id>/calculation", methods=["POST"])
@login_required
@organization_required
def update_calculation_config(config_id):
    """
    Update the calculation configuration for a KPI config.
    Supports manual, linked, and formula calculation types.
    """
    org_id = session.get("organization_id")

    # Get the config and verify access
    config = KPIValueTypeConfig.query.get_or_404(config_id)
    # Get organization through KPI -> InitiativeSystemLink -> Initiative
    kpi_org_id = config.kpi.initiative_system_link.initiative.organization_id
    if kpi_org_id != org_id:
        return jsonify({"error": "Access denied"}), 403

    # Check permissions
    if not current_user.can_manage_kpis(org_id):
        return jsonify({"error": "Permission denied"}), 403

    data = request.get_json()
    calculation_type = data.get("calculation_type")
    calculation_config = data.get("calculation_config")
    clear_contributions = data.get("clear_contributions", False)

    # Validate calculation type
    if calculation_type not in [
        KPIValueTypeConfig.CALC_TYPE_MANUAL,
        KPIValueTypeConfig.CALC_TYPE_LINKED,
        KPIValueTypeConfig.CALC_TYPE_FORMULA,
    ]:
        return jsonify({"error": "Invalid calculation type"}), 400

    # For formula type, validate the configuration
    if calculation_type == KPIValueTypeConfig.CALC_TYPE_FORMULA:
        if not calculation_config or not calculation_config.get("kpi_config_ids"):
            return jsonify({"error": "Formula configuration must include kpi_config_ids"}), 400

        # Validate that source configs exist and are accessible
        source_ids = calculation_config.get("kpi_config_ids", [])
        source_configs = KPIValueTypeConfig.query.filter(KPIValueTypeConfig.id.in_(source_ids)).all()

        if len(source_configs) != len(source_ids):
            return jsonify({"error": "Some source KPIs not found"}), 404

        # Check for circular dependencies (simple check)
        if config_id in source_ids:
            return jsonify({"error": "Cannot reference self in formula"}), 400

        # Validate that all source KPIs are numeric (formulas can't use qualitative values)
        for source_config in source_configs:
            if not source_config.value_type.is_numeric():
                return (
                    jsonify(
                        {
                            "error": f"Formula cannot use qualitative value type '{source_config.value_type.name}' ({source_config.value_type.kind}). Only numeric value types can be used in formulas."
                        }
                    ),
                    400,
                )

        # Validate based on mode
        mode = calculation_config.get("mode", "simple")

        if mode == "simple":
            # Validate operation for simple mode
            valid_operations = ["sum", "avg", "min", "max", "multiply", "subtract", "divide"]
            if calculation_config.get("operation") not in valid_operations:
                return jsonify({"error": f"Invalid operation. Must be one of: {', '.join(valid_operations)}"}), 400
        elif mode == "advanced":
            # Validate expression exists for advanced mode
            if not calculation_config.get("expression"):
                return jsonify({"error": "Advanced mode requires an expression"}), 400

            # Basic validation: check that expression only contains valid KPI references
            expression = calculation_config.get("expression")
            import re

            # Find all kpi_* references in expression
            kpi_refs = re.findall(r"kpi_(\d+)", expression)
            referenced_ids = [int(ref) for ref in kpi_refs]

            # Ensure all referenced KPIs are in the source list
            for ref_id in referenced_ids:
                if ref_id not in source_ids:
                    return jsonify({"error": f"Expression references kpi_{ref_id} which is not in selected KPIs"}), 400
        else:
            return jsonify({"error": "Invalid mode. Must be 'simple' or 'advanced'"}), 400

    # For linked type, validate compatibility
    if calculation_type == KPIValueTypeConfig.CALC_TYPE_LINKED and calculation_config:
        source_value_type_id = calculation_config.get("linked_value_type_id")
        if source_value_type_id:
            source_vt = ValueType.query.get(source_value_type_id)
            current_vt = config.value_type

            if not source_vt:
                return jsonify({"error": "Source value type not found"}), 404

            # Check compatibility: numeric to numeric, or same qualitative type
            if current_vt.kind == "numeric":
                if source_vt.kind != "numeric":
                    return (
                        jsonify(
                            {
                                "error": f"Cannot link numeric value type to {source_vt.kind}. Only numeric to numeric is allowed."
                            }
                        ),
                        400,
                    )
            else:
                # Qualitative types must match exactly
                if source_vt.kind != current_vt.kind:
                    return (
                        jsonify(
                            {
                                "error": f"Cannot link {current_vt.kind} to {source_vt.kind}. Qualitative types must match exactly."
                            }
                        ),
                        400,
                    )

    # Update the configuration
    config.calculation_type = calculation_type
    config.calculation_config = calculation_config

    # For linked type, also set the linked_source_* fields
    if calculation_type == KPIValueTypeConfig.CALC_TYPE_LINKED and calculation_config:
        config.linked_source_org_id = calculation_config.get("linked_org_id")
        config.linked_source_kpi_id = calculation_config.get("linked_kpi_id")
        config.linked_source_value_type_id = calculation_config.get("linked_value_type_id")
    elif calculation_type != KPIValueTypeConfig.CALC_TYPE_LINKED:
        # Clear linked fields if switching away from linked mode
        config.linked_source_org_id = None
        config.linked_source_kpi_id = None
        config.linked_source_value_type_id = None

    # Clear contributions if requested
    if clear_contributions:
        from app.models.contribution import Contribution

        deleted_count = Contribution.query.filter_by(kpi_value_type_config_id=config.id).delete()
        if deleted_count > 0:
            flash(f"Cleared {deleted_count} old contribution{'s' if deleted_count > 1 else ''}", "info")

    db.session.commit()

    flash("Calculation configuration updated successfully", "success")
    return jsonify({"success": True, "message": "Configuration updated"})


@bp.route("/api/organizations-for-linking")
@login_required
def organizations_for_linking():
    """Get list of organizations user has access to for linking KPIs"""
    # Get all organizations the user is a member of
    orgs = []
    for membership in current_user.organization_memberships:
        if membership.organization.is_active:
            orgs.append({"id": membership.organization_id, "name": membership.organization.name})
    return jsonify({"organizations": orgs})


@bp.route("/api/kpis-for-linking/<int:org_id>")
@login_required
def kpis_for_linking(org_id):
    """Get list of KPIs from a specific organization for linking.

    Optionally filters by value type kind - only returns KPIs that have at least one
    value type matching the required kind (for linking compatibility).

    Query params:
        kind: Value type kind to filter by (numeric, sentiment, risk, etc.)
    """
    # Verify user has access to this organization
    membership = UserOrganizationMembership.query.filter_by(user_id=current_user.id, organization_id=org_id).first()
    if not membership:
        return jsonify({"error": "Access denied"}), 403

    # Get optional kind filter
    required_kind = request.args.get("kind")

    # Get all KPIs in this organization with their value type configs
    query = (
        db.session.query(KPI)
        .join(InitiativeSystemLink)
        .join(Initiative)
        .filter(Initiative.organization_id == org_id, KPI.is_archived.is_(False))
    )

    # If kind filter specified, only get KPIs that have at least one matching value type
    if required_kind:
        query = (
            query.join(KPIValueTypeConfig, KPIValueTypeConfig.kpi_id == KPI.id)
            .join(ValueType, ValueType.id == KPIValueTypeConfig.value_type_id)
            .filter((ValueType.kind == "numeric" if required_kind == "numeric" else ValueType.kind == required_kind))
        )

    kpis = query.distinct().all()

    result = [{"id": kpi.id, "name": kpi.name} for kpi in kpis]
    return jsonify({"kpis": result})


@bp.route("/api/kpi/<int:kpi_id>/value-types")
@login_required
def kpi_value_types(kpi_id):
    """Get value types configured for a specific KPI"""
    kpi = KPI.query.get_or_404(kpi_id)

    # Verify user has access (through organization membership)
    org_id = kpi.initiative_system_link.initiative.organization_id
    membership = UserOrganizationMembership.query.filter_by(user_id=current_user.id, organization_id=org_id).first()
    if not membership:
        return jsonify({"error": "Access denied"}), 403

    # Get value types for this KPI
    configs = KPIValueTypeConfig.query.filter_by(kpi_id=kpi_id).all()
    result = [
        {"id": config.value_type_id, "name": config.value_type.name, "unit_label": config.value_type.unit_label or ""}
        for config in configs
    ]
    return jsonify({"value_types": result})


@bp.route("/api/announcement/<int:announcement_id>/acknowledge", methods=["POST"])
@login_required
def acknowledge_announcement(announcement_id):
    """Acknowledge an announcement (mark as dismissed/read)"""
    # Check if already acknowledged
    existing = UserAnnouncementAcknowledgment.query.filter_by(
        announcement_id=announcement_id, user_id=current_user.id
    ).first()

    if not existing:
        acknowledgment = UserAnnouncementAcknowledgment(announcement_id=announcement_id, user_id=current_user.id)
        db.session.add(acknowledgment)
        db.session.commit()

    return jsonify({"success": True})


@bp.route("/data")
@login_required
def get_data():
    """
    API endpoint that returns ALL workspace data as JSON for Alpine.js to handle

    Returns:
    {
        "spaces": [...],
        "valueTypes": [...],
        "governanceBodies": [...],
        "groups": [...],
        "impactLevels": [...]
    }
    """
    org_id = session.get("organization_id")

    try:
        return _build_workspace_data(org_id)
    except Exception as e:
        import traceback
        current_app.logger.error(f"Workspace get_data error: {traceback.format_exc()}")
        return jsonify({"error": str(e), "spaces": [], "valueTypes": [], "governanceBodies": [], "groups": [], "impactLevels": [], "impactScale": {}}), 500


def _build_workspace_data(org_id):
    """Build and return workspace data JSON. Extracted for error handling."""
    import time as _perf_time, logging as _perf_log  # [PERF_TRACE] removable
    _pt = _perf_log.getLogger("perf_trace")  # [PERF_TRACE]
    import os as _perf_os  # [PERF_TRACE]
    _perf_logpath = _perf_os.path.join(_perf_os.path.dirname(_perf_os.path.dirname(_perf_os.path.dirname(__file__))), "perf_trace.log")  # [PERF_TRACE]
    if not _pt.handlers: _pt.addHandler(_perf_log.FileHandler(_perf_logpath)); _pt.setLevel(_perf_log.INFO)  # [PERF_TRACE]
    _perf_start = _perf_time.time()  # [PERF_TRACE]
    _pt.info(f"[PERF_TRACE] _build_workspace_data START org={org_id}")  # [PERF_TRACE]
    # Get spaces with privacy filtering
    spaces_query = Space.query.filter_by(organization_id=org_id)
    if not current_user.is_global_admin and not current_user.is_super_admin and not current_user.is_org_admin(org_id):
        spaces_query = spaces_query.filter(or_(Space.is_private.is_(False), Space.created_by == current_user.id))

    # Eager-load the entire hierarchy to avoid N+1 queries
    spaces = (
        spaces_query
        .options(
            selectinload(Space.challenges)
            .selectinload(Challenge.initiative_links)
            .options(
                selectinload(ChallengeInitiativeLink.initiative)
                .selectinload(Initiative.system_links)
                .options(
                    selectinload(InitiativeSystemLink.system),
                    selectinload(InitiativeSystemLink.kpis)
                    .options(
                        selectinload(KPI.value_type_configs),
                        selectinload(KPI.governance_body_links)
                        .selectinload(KPIGovernanceBodyLink.governance_body),
                    ),
                ),
            ),
        )
        .order_by(Space.display_order, Space.name)
        .all()
    )

    # Get value types
    value_types = (
        ValueType.query.filter_by(organization_id=org_id, is_active=True).order_by(ValueType.display_order).all()
    )

    # Pre-computed rollup cache (if enabled)
    from app.models import RollupCacheEntry
    from app.models.system_setting import SystemSetting as _SS
    _use_precomputed = _SS.is_precompute_rollups_enabled()
    _tree_cache_on = _SS.is_tree_cache_enabled()  # [PERF_TRACE]
    _pt.info(f"[PERF_TRACE] Settings: precompute={_use_precomputed}, tree_cache={_tree_cache_on}")  # [PERF_TRACE]
    _rollup_cache = {}  # (entity_type, entity_id, vt_id) → RollupCacheEntry
    if _use_precomputed:
        _stale_info = _SS.get_rollup_stale_info(org_id)  # [PERF_TRACE]
        _is_stale = _stale_info is not None  # [PERF_TRACE]
        _pt.info(f"[PERF_TRACE] Rollup cache stale={_is_stale} info={type(_stale_info).__name__}({_stale_info!r:.100s}...)" if _stale_info else f"[PERF_TRACE] Rollup cache FRESH for org={org_id}")  # [PERF_TRACE]
        # Auto-recompute if cache is stale
        if _is_stale:
            try:
                from app.services.rollup_compute_service import RollupComputeService
                if isinstance(_stale_info, list):
                    _rc_result = RollupComputeService.recompute_incremental(org_id, _stale_info)
                    _pt.info(f"[PERF_TRACE] INCREMENTAL RECOMPUTE: {_rc_result['entities_computed']} entities, {_rc_result['values_cached']} cached in {_rc_result['duration_ms']}ms")  # [PERF_TRACE]
                else:
                    _rc_result = RollupComputeService.recompute_organization(org_id)
                    _pt.info(f"[PERF_TRACE] FULL RECOMPUTE: {_rc_result['entities_computed']} entities, {_rc_result['values_cached']} values in {_rc_result['duration_ms']}ms")  # [PERF_TRACE]
            except Exception as _rc_err:
                _pt.error(f"[PERF_TRACE] ROLLUP RECOMPUTE FAILED: {_rc_err}")  # [PERF_TRACE]
                _use_precomputed = False  # Fall back to live
        else:
            _pt.info(f"[PERF_TRACE] Rollup cache FRESH — skipping recompute")  # [PERF_TRACE]

        if _use_precomputed:
            _cache_load_start = _perf_time.time()  # [PERF_TRACE]
            _cache_entries = RollupCacheEntry.query.filter_by(organization_id=org_id).all()
            for _ce in _cache_entries:
                _rollup_cache[(_ce.entity_type, _ce.entity_id, _ce.value_type_id)] = _ce
            _pt.info(f"[PERF_TRACE] Cache loaded: {len(_cache_entries)} entries in {int((_perf_time.time() - _cache_load_start) * 1000)}ms")  # [PERF_TRACE]
    else:
        _pt.info(f"[PERF_TRACE] Precompute OFF — computing rollups LIVE")  # [PERF_TRACE]

    def _get_cached_rollup(entity_type, entity_id, vt_id):
        """Get pre-computed rollup value from cache, or None."""
        ce = _rollup_cache.get((entity_type, entity_id, vt_id))
        if not ce:
            return None
        return {
            "value": ce.value,
            "formatted_value": ce.formatted_value,
            "unit_label": ce.unit_label,
            "color": ce.color or "#6c757d",
            "formula": ce.formula,
            "is_complete": ce.is_complete,
            "count_total": ce.count_total or 0,
            "count_included": ce.count_included or 0,
            "list_label": ce.list_label,
            "list_color": ce.list_color,
        }

    def _get_cached_kpi_value(kpi_id, vt_id):
        """Get pre-computed KPI consensus value from cache."""
        ce = _rollup_cache.get(("kpi", kpi_id, vt_id))
        if not ce:
            return None
        return {
            "config_id": None,  # Not available from cache
            "value": ce.value,
            "formatted_value": ce.formatted_value,
            "unit_label": ce.unit_label,
            "color": ce.color,
            "calculation_type": ce.calculation_type or "manual",
            "consensus_status": ce.consensus_status or "no_data",
            "consensus_count": ce.consensus_count or 0,
            "comments_tooltip": ce.comments_tooltip or "",
            "has_target": ce.has_target or False,
            "list_label": ce.list_label,
            "list_color": ce.list_color,
            "target_value": ce.target_value_formatted,
            "target_date": ce.target_date,
            "target_direction": ce.target_direction,
            "target_progress": ce.target_progress,
            "target_color": ce.target_color,
        }

    # Get governance bodies
    governance_bodies = (
        GovernanceBody.query.filter_by(organization_id=org_id, is_active=True)
        .order_by(GovernanceBody.display_order)
        .all()
    )

    # Get entity type defaults for logo/icon fallbacks
    entity_defaults = EntityTypeDefault.query.filter_by(organization_id=org_id).all()
    default_logos = {}
    default_icons = {}
    for default in entity_defaults:
        # Store default logo if exists
        if default.default_logo_data and default.default_logo_mime_type:
            default_logos[default.entity_type] = (
                f"data:{default.default_logo_mime_type};base64,"
                f"{base64.b64encode(default.default_logo_data).decode('utf-8')}"
            )
        # Store default icon (text/emoji)
        if default.default_icon:
            default_icons[default.entity_type] = default.default_icon

    # Helper function to get logo URL for an entity
    def get_logo_url(entity, entity_type):
        """Get logo URL - entity's own logo or default logo for the type"""
        if (
            hasattr(entity, "logo_data")
            and entity.logo_data
            and hasattr(entity, "logo_mime_type")
            and entity.logo_mime_type
        ):
            return f"data:{entity.logo_mime_type};base64,{base64.b64encode(entity.logo_data).decode('utf-8')}"
        return default_logos.get(entity_type)

    # Helper function to get icon for an entity
    def get_icon(entity, entity_type):
        """Get icon - entity's own icon or default icon for the type"""
        if hasattr(entity, "icon") and entity.icon:
            return entity.icon
        return default_icons.get(entity_type)

    # Batch-load ALL entity links for this org in ONE query
    from app.models import EntityLink

    # Collect entity IDs by type
    _ids_by_type = {"organization": [org_id], "space": [], "challenge": [], "initiative": [], "system": [], "kpi": []}
    for _sp in spaces:
        _ids_by_type["space"].append(_sp.id)
        for _ch in _sp.challenges:
            _ids_by_type["challenge"].append(_ch.id)
            for _ci in _ch.initiative_links:
                _ids_by_type["initiative"].append(_ci.initiative.id)
                for _sl in _ci.initiative.system_links:
                    _ids_by_type["system"].append(_sl.system.id)
                    for _kpi in _sl.kpis:
                        _ids_by_type["kpi"].append(_kpi.id)

    _or_conditions = []
    for _etype, _eids in _ids_by_type.items():
        if _eids:
            _or_conditions.append(
                db.and_(EntityLink.entity_type == _etype, EntityLink.entity_id.in_(_eids))
            )
    _all_links = (
        EntityLink.query.filter(or_(*_or_conditions)).order_by(EntityLink.display_order).all()
        if _or_conditions else []
    )
    # Build lookup: (entity_type, entity_id) → [serialized links]
    _links_cache = {}
    for _link in _all_links:
        _type_info = _link.get_type_info()
        _key = (_link.entity_type, _link.entity_id)
        if _key not in _links_cache:
            _links_cache[_key] = []
        _links_cache[_key].append({
            "id": _link.id,
            "title": _link.title,
            "url": _link.url,
            "bs_icon": _type_info["bs_icon"],
            "icon_color": _type_info["color"],
        })

    def get_entity_links(entity_type, entity_id):
        """Get links from pre-loaded cache — O(1) dict lookup, no DB query."""
        return _links_cache.get((entity_type, entity_id), [])

    # Helper function to calculate formula value types from rollup values
    def calculate_formula_value_types(rollup_values_dict, all_value_types, color_config_getter=None):
        """
        Calculate formula value types from existing rollup values.

        Args:
            rollup_values_dict: dict of {value_type_id: {value, formatted_value, ...}}
            all_value_types: list of all ValueType objects
            color_config_getter: optional function to get color config for a value type

        Returns:
            Updated rollup_values_dict with formula value types added
        """
        from flask import current_app
        from simpleeval import FunctionNotDefined, InvalidExpression, simple_eval

        for vt in all_value_types:
            if not vt.is_formula() or vt.id in rollup_values_dict:
                continue  # Skip non-formula or already calculated

            if not vt.calculation_config:
                continue

            mode = vt.calculation_config.get("mode", "simple")

            try:
                if mode == "advanced":
                    # Advanced mode: Python expression with variable names
                    expression = vt.calculation_config.get("expression")
                    if not expression:
                        continue

                    # Build context from rollup values
                    context = {}
                    for other_vt in all_value_types:
                        if other_vt.id in rollup_values_dict and not other_vt.is_formula():
                            var_name = other_vt.name.lower().replace(" ", "_").replace("-", "_")
                            value = rollup_values_dict[other_vt.id].get("value")
                            if value is not None:
                                context[var_name] = float(value)

                    # Evaluate expression
                    result = simple_eval(expression, names=context)
                    if result is not None:
                        calculated_value = float(result)
                    else:
                        continue

                else:
                    # Simple mode: operation on source value types
                    operation = vt.calculation_config.get("operation")
                    source_value_type_ids = vt.calculation_config.get("source_value_type_ids", [])

                    if not operation or not source_value_type_ids:
                        continue

                    # Check if all source value types have rollup values
                    source_values = []
                    for source_vt_id in source_value_type_ids:
                        if source_vt_id in rollup_values_dict:
                            value = rollup_values_dict[source_vt_id].get("value")
                            if value is not None:
                                source_values.append(float(value))
                            else:
                                break  # Missing value
                        else:
                            break  # Source not in rollup

                    if len(source_values) != len(source_value_type_ids):
                        continue  # Not all sources available

                    # Apply operation
                    if operation == "add":
                        calculated_value = sum(source_values)
                    elif operation == "subtract":
                        calculated_value = source_values[0]
                        for v in source_values[1:]:
                            calculated_value -= v
                    elif operation == "multiply":
                        calculated_value = 1.0
                        for v in source_values:
                            calculated_value *= v
                    elif operation == "divide":
                        if len(source_values) < 2 or source_values[1] == 0:
                            continue
                        calculated_value = source_values[0] / source_values[1]
                    else:
                        continue

                # Format and add to rollup values
                color_config = color_config_getter(vt.id) if color_config_getter else None
                formatted_value = current_app.jinja_env.filters["format_value"](calculated_value, vt, color_config)

                if color_config and hasattr(color_config, "get_value_color"):
                    color = color_config.get_value_color(calculated_value)
                else:
                    color = current_app.jinja_env.filters["default_value_color"](calculated_value)

                rollup_values_dict[vt.id] = {
                    "value": calculated_value,
                    "formatted_value": formatted_value,
                    "unit_label": vt.unit_label,
                    "color": color or "#6c757d",
                    "formula": vt.get_formula_display(),
                    "is_complete": True,
                }

            except (FunctionNotDefined, InvalidExpression, ZeroDivisionError, Exception):
                continue  # Skip on any error

        return rollup_values_dict

    # Build full hierarchical tree: Spaces → Challenges → Initiatives → Systems → KPIs
    spaces_data = []
    # Helper: compute rollup values for an entity (live or from cache)
    def _get_entity_rollups(entity_type, entity_id, entity_obj, color_config_getter):
        """Get rollup values for an entity — from cache if precomputed, else live."""
        rollup_values = {}
        if _use_precomputed:
            for vt in value_types:
                cached = _get_cached_rollup(entity_type, entity_id, vt.id)
                if cached:
                    rollup_values[vt.id] = cached
            return rollup_values

        # Live computation (current behavior)
        for vt in value_types:
            if vt.is_formula():
                continue
            rollup_data = entity_obj.get_rollup_value(vt.id)
            if rollup_data and rollup_data.get("value") is not None:
                color_config = color_config_getter(vt.id)
                formatted_value = current_app.jinja_env.filters["format_value"](
                    rollup_data.get("value"), vt, color_config
                )
                if color_config and hasattr(color_config, "get_value_color"):
                    color = color_config.get_value_color(rollup_data.get("value"))
                else:
                    color = current_app.jinja_env.filters["default_value_color"](rollup_data.get("value"))
                rollup_values[vt.id] = {
                    "value": rollup_data.get("value"),
                    "formatted_value": formatted_value,
                    "unit_label": vt.unit_label,
                    "color": color or "#6c757d",
                    "formula": rollup_data.get("formula"),
                    "is_complete": rollup_data.get("is_complete", False),
                    "count_total": rollup_data.get("count_total", 0),
                    "count_included": rollup_data.get("count_included", 0),
                    "list_label": vt.get_list_option_label(rollup_data.get("value")) if vt.is_list() else None,
                    "list_color": vt.get_list_option_color(rollup_data.get("value")) if vt.is_list() else None,
                }
        # Formula pass (only for live mode — cache already includes formulas)
        if not _use_precomputed:
            rollup_values = calculate_formula_value_types(rollup_values, value_types, color_config_getter)
        return rollup_values

    for space in spaces:
        # Get space rollup values
        space_rollup_values = _get_entity_rollups("space", space.id, space, lambda vt_id: space.get_color_config(vt_id))

        # Get space SWOT completion
        swot_filled, swot_total, swot_status = space.get_swot_completion()
        swot_completion = {
            "filled": swot_filled,
            "total": swot_total,
            "status": swot_status,  # 'empty', 'partial', 'complete'
        }

        # Get space entity links
        space_entity_links = get_entity_links("space", space.id)

        challenges_data = []
        for challenge in space.challenges:
            challenge_rollup_values = _get_entity_rollups("challenge", challenge.id, challenge, lambda vt_id, ch=challenge: ch.get_color_config(vt_id)
            )

            # Get challenge entity links
            challenge_entity_links = get_entity_links("challenge", challenge.id)

            # Get initiatives under this challenge
            initiatives_data = []
            for link in challenge.initiative_links:
                initiative = link.initiative

                initiative_rollup_values = _get_entity_rollups("initiative", initiative.id, initiative, lambda vt_id, ini=initiative: ini.get_color_config(vt_id))

                # Get initiative form completion
                form_filled, form_total, form_status = initiative.get_form_completion()
                form_completion = {
                    "filled": form_filled,
                    "total": form_total,
                    "status": form_status,  # 'empty', 'partial', 'complete'
                }

                # Get initiative entity links
                initiative_entity_links = get_entity_links("initiative", initiative.id)

                # Get systems under this initiative
                systems_data = []
                for sys_link in initiative.system_links:
                    system = sys_link.system

                    system_rollup_values = _get_entity_rollups("system", system.id, sys_link, lambda vt_id, sl=sys_link: sl.get_color_config(vt_id))

                    # Get system entity links
                    system_entity_links = get_entity_links("system", system.id)

                    # Get KPIs under this system
                    kpis_data = []
                    for kpi in sys_link.kpis:
                        # Get KPI values with full details for rendering
                        kpi_values = {}
                        for vt in value_types:
                            # Pre-computed path: read from cache
                            if _use_precomputed:
                                cached_kpi = _get_cached_kpi_value(kpi.id, vt.id)
                                if cached_kpi:
                                    kpi_values[vt.id] = cached_kpi
                                continue

                            # Find config for this value type
                            config = next((c for c in kpi.value_type_configs if c.value_type_id == vt.id), None)

                            # For formula value types: calculate on-the-fly if no config exists but source value types are available
                            if not config and vt.is_formula():
                                # Check if this KPI has all required source value types
                                source_value_type_ids = (
                                    vt.calculation_config.get("source_value_type_ids", [])
                                    if vt.calculation_config.get("mode") == "simple"
                                    else []
                                )

                                if vt.calculation_config.get("mode") == "advanced":
                                    # For advanced mode, extract variable names from configs and check if they exist
                                    can_calculate = True
                                else:
                                    # For simple mode, check if all source value types have configs on this KPI
                                    source_configs = [
                                        c for c in kpi.value_type_configs if c.value_type_id in source_value_type_ids
                                    ]
                                    can_calculate = len(source_configs) == len(source_value_type_ids)

                                if can_calculate:
                                    # Create temporary config for calculation (not persisted to DB)
                                    from app.models import KPIValueTypeConfig

                                    temp_config = KPIValueTypeConfig(
                                        kpi_id=kpi.id,
                                        value_type_id=vt.id,
                                        calculation_type="manual",  # Not used for formula value types
                                    )
                                    temp_config.value_type = vt
                                    temp_config.kpi = kpi
                                    config = temp_config

                            if config:
                                consensus = config.get_consensus_value()

                                # Calculate target progress if target exists
                                target_progress = None
                                target_color = None
                                if config.target_value is not None and consensus and consensus.get("value") is not None:
                                    target_dir = config.target_direction or "maximize"
                                    target_val = float(config.target_value)
                                    current_val = float(consensus.get("value"))

                                    if target_dir == "minimize":
                                        progress = int((target_val / current_val) * 100) if current_val != 0 else 100
                                    elif target_dir == "exact":
                                        tolerance = target_val * (config.target_tolerance_pct or 10) / 100
                                        diff = abs(current_val - target_val)
                                        if diff <= tolerance:
                                            progress = 100
                                        else:
                                            progress = max(0, int(100 - ((diff - tolerance) / target_val * 100)))
                                    else:  # maximize
                                        progress = int((current_val / target_val) * 100)

                                    target_progress = progress
                                    if progress >= 90:
                                        target_color = "#28a745"
                                    elif progress >= 60:
                                        target_color = "#ffc107"
                                    else:
                                        target_color = "#dc3545"

                                # Format the value using config settings
                                formatted_value = None
                                if consensus and consensus.get("value") is not None:
                                    formatted_value = current_app.jinja_env.filters["format_value"](
                                        consensus.get("value"), vt, config
                                    )

                                _list_val = consensus.get("value") if (vt.is_list() and consensus) else None

                                # Collect non-empty contribution comments for tooltip
                                _comment_lines = [
                                    f"{c.contributor_name}: {c.comment.strip()}"
                                    for c in getattr(config, "contributions", [])
                                    if c.comment and c.comment.strip()
                                ]
                                _comments_tooltip = ("\n\nComments:\n" + "\n".join(_comment_lines)) if _comment_lines else ""

                                kpi_values[vt.id] = {
                                    "config_id": config.id,
                                    "value": consensus.get("value") if consensus else None,
                                    "formatted_value": formatted_value,
                                    "unit_label": vt.unit_label,
                                    "color": config.get_value_color(consensus.get("value")) if consensus else None,
                                    "calculation_type": config.calculation_type,
                                    "consensus_status": consensus.get("status") if consensus else "no_data",
                                    "consensus_count": consensus.get("count") if consensus else 0,
                                    "comments_tooltip": _comments_tooltip,
                                    "has_target": config.target_value is not None or bool(getattr(config, "target_list_value", None)),
                                    "list_label": vt.get_list_option_label(_list_val) if _list_val else None,
                                    "list_color": vt.get_list_option_color(_list_val) if _list_val else None,
                                    "target_value": current_app.jinja_env.filters["format_value"](config.target_value, vt, config) if config.target_value is not None else None,
                                    "target_date": (
                                        config.target_date.strftime("%Y-%m-%d") if config.target_date else None
                                    ),
                                    "target_direction": (
                                        config.target_direction or "maximize" if config.target_value else None
                                    ),
                                    "target_progress": target_progress,
                                    "target_color": target_color,
                                }

                        # Get governance body info (full details for badges)
                        governance_bodies_data = []
                        for gb_link in kpi.governance_body_links:
                            governance_bodies_data.append(
                                {
                                    "id": gb_link.governance_body.id,
                                    "name": gb_link.governance_body.name,
                                    "abbreviation": gb_link.governance_body.abbreviation,
                                    "color": gb_link.governance_body.color,
                                }
                            )

                        # Get target direction from configs (if any has a target)
                        target_direction = None
                        for config in kpi.value_type_configs:
                            if config.target_value is not None:
                                target_direction = config.target_direction or "maximize"
                                break

                        # Check if KPI has linked sources
                        has_linked_sources = any(config.linked_source_kpi_id for config in kpi.value_type_configs)

                        # Get KPI entity links
                        kpi_entity_links = get_entity_links("kpi", kpi.id)

                        kpis_data.append(
                            {
                                "id": kpi.id,
                                "name": kpi.name,
                                "description": kpi.description if kpi.description else None,
                                "display_order": kpi.display_order,
                                "logo_url": get_logo_url(kpi, "kpi"),
                                "icon": get_icon(kpi, "kpi"),
                                "values": kpi_values,
                                "is_archived": kpi.is_archived,
                                "archived_at": kpi.archived_at.strftime("%Y-%m-%d") if kpi.archived_at else None,
                                "governance_bodies": governance_bodies_data,
                                "target_direction": target_direction,
                                "has_linked_sources": has_linked_sources,
                                "entity_links": kpi_entity_links,
                                "impact_level": kpi.impact_level,
                            }
                        )

                    # Inherited links: KPIs → System (deduplicated by URL, all sources tracked)
                    _url_map = {}
                    for kd in kpis_data:
                        for lnk in kd["entity_links"]:
                            label = f"KPI: {kd['name']}"
                            if lnk["url"] in _url_map:
                                _url_map[lnk["url"]]["from_sources"].append(label)
                            else:
                                _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
                    system_inherited = list(_url_map.values())

                    # Portal: linked CISK org info
                    _portal = None
                    if system.linked_organization_id:
                        _lo = system.linked_organization
                        if _lo:
                            _lo_logo = None
                            if _lo.logo_data and _lo.logo_mime_type:
                                _lo_logo = f"data:{_lo.logo_mime_type};base64,{base64.b64encode(_lo.logo_data).decode('utf-8')}"
                            _portal = {
                                "id": _lo.id,
                                "name": _lo.name,
                                "logo_url": _lo_logo,
                            }

                    systems_data.append(
                        {
                            "id": system.id,
                            "link_id": sys_link.id,  # For parent change operations
                            "name": system.name,
                            "description": system.description if system.description else None,
                            "edit_url": url_for("organization_admin.edit_system", system_id=system.id),
                            "logo_url": get_logo_url(system, "system"),
                            "icon": get_icon(system, "system"),
                            "rollup_values": system_rollup_values,
                            "entity_links": system_entity_links,
                            "inherited_links": system_inherited,
                            "kpis": kpis_data,
                            "impact_level": system.impact_level,
                            "portal": _portal,
                        }
                    )

                # Inherited links: Systems + KPIs → Initiative (deduplicated by URL, all sources tracked)
                _url_map = {}
                for sd in systems_data:
                    for lnk in sd["entity_links"]:
                        label = f"System: {sd['name']}"
                        if lnk["url"] in _url_map:
                            _url_map[lnk["url"]]["from_sources"].append(label)
                        else:
                            _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
                    for lnk in sd["inherited_links"]:
                        label = f"{lnk['from_sources'][0]} via {sd['name']}" if lnk.get("from_sources") else lnk["from_label"]
                        if lnk["url"] in _url_map:
                            _url_map[lnk["url"]]["from_sources"].append(label)
                        else:
                            _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
                initiative_inherited = list(_url_map.values())

                initiatives_data.append(
                    {
                        "id": initiative.id,
                        "link_id": link.id,  # For parent change operations
                        "name": initiative.name,
                        "description": initiative.description if initiative.description else None,
                        "logo_url": get_logo_url(initiative, "initiative"),
                        "icon": get_icon(initiative, "initiative"),
                        "group_label": initiative.group_label,
                        "impact_on_challenge": initiative.impact_on_challenge,
                        "rollup_values": initiative_rollup_values,
                        "form_completion": form_completion,
                        "entity_links": initiative_entity_links,
                        "inherited_links": initiative_inherited,
                        "systems": systems_data,
                        "execution_rag": initiative.execution_rag,
                        "impact_level": initiative.impact_level,
                    }
                )

            # Inherited links: Initiatives + Systems + KPIs → Challenge (deduplicated, all sources tracked)
            _url_map = {}
            for ind in initiatives_data:
                for lnk in ind["entity_links"]:
                    label = f"Initiative: {ind['name']}"
                    if lnk["url"] in _url_map:
                        _url_map[lnk["url"]]["from_sources"].append(label)
                    else:
                        _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
                for lnk in ind["inherited_links"]:
                    label = f"{lnk['from_sources'][0]} via {ind['name']}" if lnk.get("from_sources") else lnk["from_label"]
                    if lnk["url"] in _url_map:
                        _url_map[lnk["url"]]["from_sources"].append(label)
                    else:
                        _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
            challenge_inherited = list(_url_map.values())

            challenges_data.append(
                {
                    "id": challenge.id,
                    "name": challenge.name,
                    "description": challenge.description if challenge.description else None,
                    "logo_url": get_logo_url(challenge, "challenge"),
                    "icon": get_icon(challenge, "challenge"),
                    "display_order": challenge.display_order,
                    "rollup_values": challenge_rollup_values,
                    "entity_links": challenge_entity_links,
                    "inherited_links": challenge_inherited,
                    "initiatives": initiatives_data,
                    "impact_level": challenge.impact_level,
                }
            )

        # Inherited links: Challenges + everything below → Space (deduplicated, all sources tracked)
        _url_map = {}
        for cd in challenges_data:
            for lnk in cd["entity_links"]:
                label = f"Challenge: {cd['name']}"
                if lnk["url"] in _url_map:
                    _url_map[lnk["url"]]["from_sources"].append(label)
                else:
                    _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
            for lnk in cd["inherited_links"]:
                label = f"{lnk['from_sources'][0]} via {cd['name']}" if lnk.get("from_sources") else lnk["from_label"]
                if lnk["url"] in _url_map:
                    _url_map[lnk["url"]]["from_sources"].append(label)
                else:
                    _url_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
        space_inherited = list(_url_map.values())

        spaces_data.append(
            {
                "id": space.id,
                "name": space.name,
                "description": space.description if space.description else None,
                "logo_url": get_logo_url(space, "space"),
                "icon": get_icon(space, "space"),
                "display_order": space.display_order,
                "is_private": space.is_private,
                "space_label": space.space_label,
                "rollup_values": space_rollup_values,
                "swot_completion": swot_completion,
                "entity_links": space_entity_links,
                "inherited_links": space_inherited,
                "challenges": challenges_data,
                "impact_level": space.impact_level,
            }
        )

    # Build value types data
    value_types_data = [
        {
            "id": vt.id,
            "name": vt.name,
            "display_order": vt.display_order,
            "unit_label": vt.unit_label,
            "kind": vt.kind,
            "calculation_type": vt.calculation_type,
            "formula_display": vt.get_formula_display() if vt.is_formula() else None,
            "list_options": vt.list_options if vt.is_list() else None,
        }
        for vt in value_types
    ]

    # Build governance bodies data
    governance_bodies_data = [{"id": gb.id, "name": gb.name} for gb in governance_bodies]

    # Get unique initiative groups
    groups = (
        db.session.query(Initiative.group_label)
        .filter(Initiative.organization_id == org_id, Initiative.group_label.isnot(None))
        .distinct()
        .all()
    )
    groups_data = [g[0] for g in groups]

    # Impact levels (legacy + new configurable system)
    impact_levels_data = [
        {"value": "not_assessed", "label": "Not Assessed"},
        {"value": "low", "label": "Low"},
        {"value": "medium", "label": "Medium"},
        {"value": "high", "label": "High"},
        {"value": "no_consensus", "label": "No Consensus"},
    ]
    # New configurable impact scale
    impact_scale = ImpactLevel.get_org_levels(org_id)

    # Build decision mention counts per entity (from Decision model)
    from app.models import Decision
    _all_decisions = Decision.query.filter_by(organization_id=org_id).all()
    _dec_counts = {}  # "entity_type:entity_id" → count
    for _dd in _all_decisions:
        for _dm in (_dd.entity_mentions or []):
            key = f"{_dm.get('entity_type')}:{_dm.get('entity_id')}"
            _dec_counts[key] = _dec_counts.get(key, 0) + 1

    # Inject decision_count into entity dicts
    for space in spaces_data:
        space["decision_count"] = _dec_counts.get(f"space:{space['id']}", 0)
        for challenge in space.get("challenges", []):
            challenge["decision_count"] = _dec_counts.get(f"challenge:{challenge['id']}", 0)
            for ini in challenge.get("initiatives", []):
                ini["decision_count"] = _dec_counts.get(f"initiative:{ini['id']}", 0)
                for sys in ini.get("systems", []):
                    sys["decision_count"] = _dec_counts.get(f"system:{sys['id']}", 0)
                    for kpi in sys.get("kpis", []):
                        kpi["decision_count"] = _dec_counts.get(f"kpi:{kpi['id']}", 0)

    # Org-level entity links (direct)
    org_entity_links = get_entity_links("organization", org_id)

    # Org inherited = all unique links from every space + their inherited children
    _org_inh_map = {}
    for sd in spaces_data:
        for lnk in sd["entity_links"]:
            label = f"Space: {sd['name']}"
            if lnk["url"] in _org_inh_map:
                _org_inh_map[lnk["url"]]["from_sources"].append(label)
            else:
                _org_inh_map[lnk["url"]] = {**lnk, "from_label": label, "from_sources": [label]}
        for lnk in sd["inherited_links"]:
            if lnk["url"] in _org_inh_map:
                _org_inh_map[lnk["url"]]["from_sources"].extend(lnk.get("from_sources", [lnk.get("from_label", "")]))
            else:
                _org_inh_map[lnk["url"]] = {**lnk}
    org_inherited_links = list(_org_inh_map.values())

    # ── Compute true importance (product of weights through the chain) ──
    # ── Compute true importance using configured method ──
    if impact_scale:
        from app.services.impact_service import compute_true_importance

        _org = Organization.query.get(org_id)
        _method = _org.impact_calc_method or "geometric_mean" if _org else "geometric_mean"
        _weights = {lvl: impact_scale[lvl]["weight"] for lvl in impact_scale}
        _custom_matrix = _org.impact_qfd_matrix if _org else None
        _custom_reinforce = _org.impact_reinforce_weights if _org else None

        def _ti(chain):
            """Compute true importance for a chain of impact levels."""
            if not chain or not all(chain):
                return None
            return compute_true_importance(chain, _method, _weights, _custom_matrix, _custom_reinforce)

        for space in spaces_data:
            s_il = space.get("impact_level")
            space["true_importance_level"] = _ti([s_il]) if s_il else None
            for challenge in space.get("challenges", []):
                c_il = challenge.get("impact_level")
                challenge["true_importance_level"] = _ti([s_il, c_il])
                for initiative in challenge.get("initiatives", []):
                    i_il = initiative.get("impact_level")
                    initiative["true_importance_level"] = _ti([s_il, c_il, i_il])
                    for system in initiative.get("systems", []):
                        sy_il = system.get("impact_level")
                        system["true_importance_level"] = _ti([s_il, c_il, i_il, sy_il])
                        for kpi in system.get("kpis", []):
                            k_il = kpi.get("impact_level")
                            kpi["true_importance_level"] = _ti([s_il, c_il, i_il, sy_il, k_il])

    _perf_total = int((_perf_time.time() - _perf_start) * 1000)  # [PERF_TRACE]
    _pt.info(f"[PERF_TRACE] _build_workspace_data END — TOTAL {_perf_total}ms (precompute={'ON' if _use_precomputed else 'OFF'})")  # [PERF_TRACE]

    return jsonify(
        {
            "spaces": spaces_data,
            "valueTypes": value_types_data,
            "governanceBodies": governance_bodies_data,
            "groups": groups_data,
            "impactLevels": impact_levels_data,
            "impactScale": impact_scale,
            "orgEntityLinks": org_entity_links,
            "orgInheritedLinks": org_inherited_links,
        }
    )


@bp.route("/api/action-items-count")
@login_required
def get_action_items_count():
    """Get count of action items requiring attention"""
    from app.services.action_items_service import ActionItemsService

    org_id = session.get("organization_id")

    # Get action items count from centralized service
    action_items_count = ActionItemsService.get_action_items_count(org_id)

    return jsonify({"total_issues": action_items_count["total"]})


@bp.route("/api/change-parent/<entity_type>", methods=["POST"])
@login_required
def change_parent(entity_type):
    """Change the parent of an entity (move to different parent)"""
    try:
        org_id = session.get("organization_id")
        data = request.get_json()

        entity_id = data.get("entity_id")
        new_parent_id = data.get("new_parent_id")

        if not entity_id or not new_parent_id:
            return jsonify({"error": "Missing entity_id or new_parent_id"}), 400

        if entity_type == "challenge":
            # Move challenge to different space
            challenge = Challenge.query.filter_by(id=entity_id, organization_id=org_id).first()
            if not challenge:
                return jsonify({"error": "Challenge not found"}), 404

            challenge.space_id = new_parent_id

        elif entity_type == "initiative":
            # Move initiative to different challenge
            link = ChallengeInitiativeLink.query.filter_by(id=entity_id).first()
            if not link or link.initiative.organization_id != org_id:
                return jsonify({"error": "Initiative link not found"}), 404

            link.challenge_id = new_parent_id

        elif entity_type == "system":
            # Move system to different initiative
            link = InitiativeSystemLink.query.filter_by(id=entity_id).first()
            if not link or link.system.organization_id != org_id:
                return jsonify({"error": "System link not found"}), 404

            link.initiative_id = new_parent_id

        elif entity_type == "kpi":
            # Move KPI to different system (system is a link)
            kpi = KPI.query.filter_by(id=entity_id).first()
            if not kpi:
                return jsonify({"error": "KPI not found"}), 404

            # Verify ownership through system link
            old_link = InitiativeSystemLink.query.get(kpi.initiative_system_link_id)
            if not old_link or old_link.system.organization_id != org_id:
                return jsonify({"error": "Unauthorized"}), 403

            # new_parent_id is the InitiativeSystemLink id
            kpi.initiative_system_link_id = new_parent_id

        else:
            return jsonify({"error": f"Unknown entity type: {entity_type}"}), 400

        db.session.commit()
        return jsonify({"success": True, "message": f"{entity_type.capitalize()} parent changed"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
