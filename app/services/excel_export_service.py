"""
Excel Export Service — multi-sheet workspace export.

Sheets:
  1. Overview     — workspace identity, structure counts, KPI RAG distribution
  2. Tree         — hierarchical Space → Challenge → Initiative → System → KPI
                    with Excel outline (expand/collapse) + RAG conditional formatting
  3. KPIs         — every KPI as one flat row with target/current/Δ/RAG
  4. Action Items — every action / memo with priority, status, mentions
  5. Settings     — value types, impact levels, pillars, governance, geography

Numeric KPI cells are stored as numbers (not strings) so Excel can sort/filter.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models import (
    ActionItem,
    EntityTypeDefault,
    GeographyRegion,
    GovernanceBody,
    ImpactLevel,
    Space,
    StrategicPillar,
    ValueType,
)


# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_BG = "1F2937"
HEADER_FG = "FFFFFF"
ACCENT = "0D6EFD"
SUBTLE_BG = "F8FAFC"
BORDER_COLOR = "E5E7EB"

LEVEL_TINTS = {
    "space":      "DBEAFE",
    "challenge":  "E0E7FF",
    "initiative": "EDE9FE",
    "system":     "FCE7F3",
    "kpi":        "F1F5F9",
}

LEVEL_FONT_SIZE = {
    "space": 12, "challenge": 11, "initiative": 11, "system": 10, "kpi": 10,
}

RAG_GREEN = "D1FADF"
RAG_AMBER = "FEF3C7"
RAG_RED = "FEE2E2"
RAG_NONE = "F1F5F9"

DEFAULT_ICON = {
    "space": "🌳", "challenge": "🎯", "initiative": "🚀",
    "system": "⚙️", "kpi": "📊",
}


# ── Helpers ──────────────────────────────────────────────────────────────────
THIN = Side(border_style="thin", color=BORDER_COLOR)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _header_cell(cell, text, *, bg=HEADER_BG, fg=HEADER_FG, size=11, align="left"):
    cell.value = text
    cell.font = Font(bold=True, size=size, color=fg)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER_ALL


def _value_cell(cell, value, *, bg=None, bold=False, italic=False, align="left",
                color=None, number_format=None, hyperlink=None):
    cell.value = value
    cell.font = Font(bold=bold, italic=italic, size=10, color=color or "111827")
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER_ALL
    if number_format:
        cell.number_format = number_format
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.font = Font(bold=bold, italic=italic, size=10, color="0563C1", underline="single")


def _compute_rag(value, target, direction, tolerance_pct):
    """Return ('green'|'amber'|'red', progress_pct) or (None, None)."""
    if value is None or target is None:
        return (None, None)
    try:
        cv = float(value)
        tv = float(target)
    except (TypeError, ValueError):
        return (None, None)
    direction = direction or "maximize"
    tol = tolerance_pct or 10
    try:
        if direction == "minimize":
            progress = int((tv / cv) * 100) if cv != 0 else 100
        elif direction == "exact":
            tol_abs = abs(tv) * tol / 100
            diff = abs(cv - tv)
            progress = 100 if diff <= tol_abs else max(0, int(100 - ((diff - tol_abs) / abs(tv) * 100))) if tv != 0 else 0
        else:  # maximize
            progress = int((cv / tv) * 100) if tv != 0 else 0
    except (ZeroDivisionError, ValueError):
        return (None, None)
    if progress >= 90:
        return ("green", progress)
    if progress >= 60:
        return ("amber", progress)
    return ("red", progress)


def _rag_fill_color(rag):
    return {"green": RAG_GREEN, "amber": RAG_AMBER, "red": RAG_RED}.get(rag)


def _kpi_value_for_export(consensus_value, vt):
    """Return a value suitable for a numeric Excel cell, or string for non-numeric kinds."""
    if consensus_value is None:
        return None
    if vt.kind == "numeric":
        try:
            return float(consensus_value)
        except (TypeError, ValueError):
            return None
    if vt.kind in ("risk", "positive_impact", "negative_impact", "level", "sentiment"):
        try:
            return int(consensus_value)
        except (TypeError, ValueError):
            return None
    return str(consensus_value)


def _numeric_format_for_vt(vt):
    if vt.kind != "numeric":
        return None
    if vt.numeric_format == "integer":
        base = "0"
    else:
        decimals = vt.decimal_places if vt.decimal_places is not None else 2
        base = "0." + ("0" * decimals)
    if vt.unit_label:
        # Excel format: 0.00 "kg"
        return f'{base} "{vt.unit_label}"'
    return base


# ═══════════════════════════════════════════════════════════════════════════════
class ExcelExportService:

    @staticmethod
    def export_workspace(organization_id, *, base_url=None, generated_by=None):
        """
        Export the full workspace to a multi-sheet Excel workbook.

        Args:
            organization_id: target organization
            base_url: e.g. "https://app.cisk.fr" — used for hyperlinks back to the app.
            generated_by: optional user-friendly identifier for the Overview sheet.

        Returns:
            BytesIO of the .xlsx file.
        """
        from app.models import Organization

        org = Organization.query.get(organization_id)
        spaces = Space.query.filter_by(organization_id=organization_id).order_by(Space.display_order, Space.name).all()
        value_types = (
            ValueType.query.filter_by(organization_id=organization_id, is_active=True)
            .order_by(ValueType.display_order, ValueType.name).all()
        )
        branding = {d.entity_type: d for d in EntityTypeDefault.query.filter_by(organization_id=organization_id).all()}

        ctx = {
            "org": org,
            "org_id": organization_id,
            "spaces": spaces,
            "value_types": value_types,
            "branding": branding,
            "base_url": (base_url or "").rstrip("/"),
            "generated_by": generated_by,
            "rollup_cache": ExcelExportService._load_rollup_cache(organization_id),
        }

        wb = Workbook()
        wb.remove(wb.active)

        ExcelExportService._build_overview(wb, ctx)
        ExcelExportService._build_tree(wb, ctx)
        ExcelExportService._build_kpis_flat(wb, ctx)
        ExcelExportService._build_action_items(wb, ctx)
        ExcelExportService._build_settings(wb, ctx)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    @staticmethod
    def _load_rollup_cache(organization_id):
        """Load all RollupCacheEntry rows for the org keyed by (entity_type, entity_id, value_type_id).

        Returns {} when pre-compute is off, so callers fall back to live aggregation.
        """
        from app.models import RollupCacheEntry
        from app.models.system_setting import SystemSetting

        if not SystemSetting.is_precompute_rollups_enabled():
            return {}
        entries = RollupCacheEntry.query.filter_by(organization_id=organization_id).all()
        return {(e.entity_type, e.entity_id, e.value_type_id): e for e in entries}

    @staticmethod
    def _kpi_value(ctx, kpi_id, vt_id, cfg):
        """Return the consensus numeric value for a KPI/value-type pair.

        Reads from the pre-computed RollupCacheEntry when present; falls back
        to cfg.get_consensus_value() (which lazy-loads contributions) only on
        miss. This avoids per-cell consensus recomputation during big exports.
        """
        cache = ctx.get("rollup_cache") or {}
        ce = cache.get(("kpi", kpi_id, vt_id))
        if ce is not None:
            return ce.value
        try:
            cons = cfg.get_consensus_value()
        except Exception:
            cons = None
        return (cons or {}).get("value")

    # ── 1. Overview ───────────────────────────────────────────────────────────
    @staticmethod
    def _build_overview(wb, ctx):
        ws = wb.create_sheet("Overview")
        org = ctx["org"]
        spaces = ctx["spaces"]

        # Counts
        n_spaces = len(spaces)
        n_challenges = sum(len(sp.challenges) for sp in spaces)
        n_initiatives = sum(len(ch.initiative_links) for sp in spaces for ch in sp.challenges)
        seen_systems, seen_kpis = set(), 0
        for sp in spaces:
            for ch in sp.challenges:
                for il in ch.initiative_links:
                    for sl in il.initiative.system_links:
                        seen_systems.add(sl.system_id)
                        seen_kpis += len(sl.kpis)
        n_systems = len(seen_systems)
        n_kpis = seen_kpis

        # KPI RAG distribution
        rag_counts = {"green": 0, "amber": 0, "red": 0, "none": 0}
        for sp in spaces:
            for ch in sp.challenges:
                for il in ch.initiative_links:
                    for sl in il.initiative.system_links:
                        for kpi in sl.kpis:
                            for cfg in kpi.value_type_configs:
                                value = ExcelExportService._kpi_value(ctx, kpi.id, cfg.value_type_id, cfg)
                                rag, _ = _compute_rag(value, cfg.target_value, cfg.target_direction, cfg.target_tolerance_pct)
                                rag_counts[rag or "none"] += 1

        action_total = ActionItem.query.filter_by(organization_id=ctx["org_id"]).count()
        gb_total = GovernanceBody.query.filter_by(organization_id=ctx["org_id"]).count()
        pillar_total = StrategicPillar.query.filter_by(organization_id=ctx["org_id"]).count()
        vt_total = len(ctx["value_types"])

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 60
        for col in ("C", "D", "E", "F"):
            ws.column_dimensions[col].width = 16

        # Title block
        ws.merge_cells("A1:F1")
        _header_cell(ws["A1"], f"{org.name} — Workspace Export",
                     bg=HEADER_BG, size=18, align="left")
        ws.row_dimensions[1].height = 36
        ws.merge_cells("A2:F2")
        gen = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
        sub = f"Generated {gen}"
        if ctx["generated_by"]:
            sub += f" · by {ctx['generated_by']}"
        try:
            from app import __version__ as app_version
            sub += f" · CISK Navigator v{app_version}"
        except Exception:
            pass
        _value_cell(ws["A2"], sub, italic=True, color="6B7280")

        if org.description:
            ws.merge_cells("A3:F3")
            _value_cell(ws["A3"], org.description, italic=True, color="374151")

        # Counts table
        ws["A5"] = ""
        _header_cell(ws["A5"], "STRUCTURE")
        _header_cell(ws["B5"], "COUNT", align="right")
        for r, (label, val) in enumerate([
            ("Spaces", n_spaces),
            ("Challenges", n_challenges),
            ("Initiatives (unique)", n_initiatives),
            ("Systems (unique)", n_systems),
            ("KPIs", n_kpis),
        ], start=6):
            _value_cell(ws.cell(r, 1), label, bold=True)
            _value_cell(ws.cell(r, 2), val, align="right", number_format="#,##0")

        # KPI RAG distribution
        _header_cell(ws["A12"], "KPI HEALTH")
        _header_cell(ws["B12"], "COUNT", align="right")
        for r, (label, key, fill) in enumerate([
            ("🟢 On track (≥90% to target)", "green", RAG_GREEN),
            ("🟡 At risk (60–89%)", "amber", RAG_AMBER),
            ("🔴 Off track (<60%)", "red", RAG_RED),
            ("⚪ No target / no data", "none", RAG_NONE),
        ], start=13):
            _value_cell(ws.cell(r, 1), label, bold=True, bg=fill)
            _value_cell(ws.cell(r, 2), rag_counts[key], align="right", number_format="#,##0", bg=fill)

        # Org-level counts
        _header_cell(ws["A19"], "OTHER")
        _header_cell(ws["B19"], "COUNT", align="right")
        for r, (label, val) in enumerate([
            ("Action items", action_total),
            ("Governance bodies", gb_total),
            ("Strategic pillars", pillar_total),
            ("Value types", vt_total),
        ], start=20):
            _value_cell(ws.cell(r, 1), label, bold=True)
            _value_cell(ws.cell(r, 2), val, align="right", number_format="#,##0")

        # Sheet hint
        ws.merge_cells("A26:F26")
        _value_cell(
            ws["A26"],
            "→ Use the bottom tabs: Tree (hierarchical, expand/collapse), KPIs (flat), Action Items, Settings.",
            italic=True, color="6B7280",
        )

    # ── 2. Tree ───────────────────────────────────────────────────────────────
    @staticmethod
    def _build_tree(wb, ctx):
        ws = wb.create_sheet("Tree")
        value_types = ctx["value_types"]
        spaces = ctx["spaces"]
        branding = ctx["branding"]

        # Outline +/- toggle on the PARENT row, not below the block.
        ws.sheet_properties.outlinePr.summaryBelow = False
        ws.sheet_properties.outlinePr.summaryRight = False

        # Columns: A=Structure | B=Owner/Notes | C..=value types
        ws.column_dimensions["A"].width = 56
        ws.column_dimensions["B"].width = 22
        for i, vt in enumerate(value_types, start=3):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Header
        _header_cell(ws.cell(1, 1), "Structure")
        _header_cell(ws.cell(1, 2), "Owner / Notes")
        for i, vt in enumerate(value_types, start=3):
            label = vt.name + (f"\n({vt.unit_label})" if vt.unit_label else "")
            _header_cell(ws.cell(1, i), label, align="center")
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "C2"

        row = 2
        for sp in spaces:
            row = ExcelExportService._tree_space(ws, sp, value_types, branding, ctx, row)

        # Bottom totals row (KPI count)
        ws.cell(row + 1, 1, "")
        last_row = row - 1
        # Autofilter on header
        ws.auto_filter.ref = f"A1:{get_column_letter(2 + len(value_types))}{last_row}"

    @staticmethod
    def _icon(branding, kind):
        d = branding.get(kind)
        return (d.default_icon if d and d.default_icon else DEFAULT_ICON.get(kind, "•"))

    @staticmethod
    def _row_url(ctx, kind, entity_id):
        if not ctx["base_url"]:
            return None
        # All entities live under /workspace; deep links can be improved later.
        return f"{ctx['base_url']}/workspace/?focus={kind}:{entity_id}"

    @staticmethod
    def _tree_space(ws, sp, value_types, branding, ctx, row):
        start = row
        icon = ExcelExportService._icon(branding, "space")
        bg = LEVEL_TINTS["space"]
        _value_cell(ws.cell(row, 1), f"{icon}  {sp.name}", bold=True, bg=bg, color="0D47A1")
        ws.cell(row, 1).font = Font(bold=True, size=LEVEL_FONT_SIZE["space"], color="0D47A1")
        _value_cell(ws.cell(row, 2), sp.space_label or "", italic=True, bg=bg, color="475569")
        ExcelExportService._write_rollup_row(ws, row, "space", sp.id, sp, value_types, ctx, bg, bold=True)
        ws.row_dimensions[row].height = 22
        row += 1

        for ch in sp.challenges:
            row = ExcelExportService._tree_challenge(ws, ch, value_types, branding, ctx, row)

        if row > start + 1:
            ws.row_dimensions.group(start + 1, row - 1, outline_level=1, hidden=False)
        return row

    @staticmethod
    def _tree_challenge(ws, ch, value_types, branding, ctx, row):
        start = row
        icon = ExcelExportService._icon(branding, "challenge")
        bg = LEVEL_TINTS["challenge"]
        _value_cell(ws.cell(row, 1), f"  {icon}  {ch.name}", bold=True, bg=bg, color="3730A3")
        ws.cell(row, 1).font = Font(bold=True, size=LEVEL_FONT_SIZE["challenge"], color="3730A3")
        _value_cell(ws.cell(row, 2), "", bg=bg)
        ExcelExportService._write_rollup_row(ws, row, "challenge", ch.id, ch, value_types, ctx, bg, bold=True)
        row += 1

        for il in ch.initiative_links:
            row = ExcelExportService._tree_initiative(ws, il.initiative, value_types, branding, ctx, row)

        if row > start + 1:
            ws.row_dimensions.group(start + 1, row - 1, outline_level=2, hidden=False)
        return row

    @staticmethod
    def _tree_initiative(ws, ini, value_types, branding, ctx, row):
        start = row
        icon = ExcelExportService._icon(branding, "initiative")
        bg = LEVEL_TINTS["initiative"]
        _value_cell(ws.cell(row, 1), f"    {icon}  {ini.name}", bold=True, bg=bg, color="5B21B6")
        ws.cell(row, 1).font = Font(bold=True, size=LEVEL_FONT_SIZE["initiative"], color="5B21B6")
        owner = ini.responsible_person or ""
        _value_cell(ws.cell(row, 2), owner, italic=True, bg=bg, color="475569")
        ExcelExportService._write_rollup_row(ws, row, "initiative", ini.id, ini, value_types, ctx, bg)
        row += 1

        for sl in ini.system_links:
            row = ExcelExportService._tree_system(ws, sl, value_types, branding, ctx, row)

        if row > start + 1:
            ws.row_dimensions.group(start + 1, row - 1, outline_level=3, hidden=False)
        return row

    @staticmethod
    def _tree_system(ws, sl, value_types, branding, ctx, row):
        start = row
        icon = ExcelExportService._icon(branding, "system")
        bg = LEVEL_TINTS["system"]
        sys_obj = sl.system
        _value_cell(ws.cell(row, 1), f"      {icon}  {sys_obj.name}", bg=bg, color="9D174D")
        ws.cell(row, 1).font = Font(size=LEVEL_FONT_SIZE["system"], color="9D174D")
        _value_cell(ws.cell(row, 2), "", bg=bg)
        ExcelExportService._write_rollup_row(ws, row, "system", sl.system_id, sl, value_types, ctx, bg)
        row += 1

        for kpi in sl.kpis:
            row = ExcelExportService._tree_kpi(ws, kpi, value_types, branding, ctx, row)

        if row > start + 1:
            ws.row_dimensions.group(start + 1, row - 1, outline_level=4, hidden=False)
        return row

    @staticmethod
    def _tree_kpi(ws, kpi, value_types, branding, ctx, row):
        icon = ExcelExportService._icon(branding, "kpi")
        bg = LEVEL_TINTS["kpi"]
        _value_cell(ws.cell(row, 1), f"        {icon}  {kpi.name}", italic=True, bg=bg, color="334155")
        _value_cell(ws.cell(row, 2), "", bg=bg)

        # Each value-type cell stores a NUMBER (when numeric) and gets RAG fill if target exists.
        for i, vt in enumerate(value_types, start=3):
            cfg = next((c for c in kpi.value_type_configs if c.value_type_id == vt.id), None)
            if not cfg:
                _value_cell(ws.cell(row, i), None, bg=bg)
                continue
            value = ExcelExportService._kpi_value(ctx, kpi.id, vt.id, cfg)
            cell_value = _kpi_value_for_export(value, vt)
            rag, _ = _compute_rag(value, cfg.target_value, cfg.target_direction, cfg.target_tolerance_pct)
            cell_bg = _rag_fill_color(rag) or bg
            _value_cell(
                ws.cell(row, i),
                cell_value,
                bg=cell_bg,
                align="center",
                number_format=_numeric_format_for_vt(vt),
                bold=True if rag else False,
            )
        return row + 1

    @staticmethod
    def _write_rollup_row(ws, row, entity_type, entity_id, entity, value_types, ctx, bg, *, bold=False):
        """Write rollup values for non-leaf entities.

        Reads from ctx["rollup_cache"] when available; falls back to a live
        entity.get_rollup_value() call only on cache miss (precompute disabled
        or entry not yet computed for this entity/value-type pair).
        """
        cache = ctx.get("rollup_cache") or {}
        for i, vt in enumerate(value_types, start=3):
            ce = cache.get((entity_type, entity_id, vt.id))
            if ce is not None:
                value = ce.value
                formatted = ce.formatted_value
            else:
                try:
                    rollup = entity.get_rollup_value(vt.id) or {}
                except Exception:
                    rollup = {}
                value = rollup.get("value")
                formatted = rollup.get("formatted_value")
            if value is not None:
                fmt = _numeric_format_for_vt(vt)
                if vt.kind == "numeric":
                    try:
                        v = float(value)
                    except (TypeError, ValueError):
                        v = None
                else:
                    v = formatted or str(value)
                _value_cell(ws.cell(row, i), v, bg=bg, align="center",
                            number_format=fmt, bold=bold)
            else:
                _value_cell(ws.cell(row, i), None, bg=bg)

    # ── 3. KPIs flat ──────────────────────────────────────────────────────────
    @staticmethod
    def _build_kpis_flat(wb, ctx):
        ws = wb.create_sheet("KPIs")
        headers = [
            "Space", "Challenge", "Initiative", "System", "KPI",
            "Value Type", "Current", "Target", "Δ %", "RAG",
            "Direction", "Target Date", "Tolerance %", "Last Update", "Last Contributor",
        ]
        for i, h in enumerate(headers, start=1):
            _header_cell(ws.cell(1, i), h, align="center")
            ws.column_dimensions[get_column_letter(i)].width = 18
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["E"].width = 28
        ws.column_dimensions["F"].width = 18
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "F2"

        row = 2
        for sp in ctx["spaces"]:
            for ch in sp.challenges:
                for il in ch.initiative_links:
                    ini = il.initiative
                    for sl in ini.system_links:
                        sys_obj = sl.system
                        for kpi in sl.kpis:
                            for cfg in kpi.value_type_configs:
                                vt = cfg.value_type
                                if not vt:
                                    continue
                                value = ExcelExportService._kpi_value(ctx, kpi.id, vt.id, cfg)
                                cell_value = _kpi_value_for_export(value, vt)
                                rag, progress = _compute_rag(value, cfg.target_value, cfg.target_direction, cfg.target_tolerance_pct)
                                rag_bg = _rag_fill_color(rag) or RAG_NONE

                                # Last contribution
                                last_contrib = None
                                if hasattr(cfg, "contributions"):
                                    sorted_c = sorted(
                                        [c for c in cfg.contributions if getattr(c, "created_at", None)],
                                        key=lambda c: c.created_at, reverse=True,
                                    )
                                    if sorted_c:
                                        last_contrib = sorted_c[0]

                                _value_cell(ws.cell(row, 1), sp.name)
                                _value_cell(ws.cell(row, 2), ch.name)
                                _value_cell(ws.cell(row, 3), ini.name)
                                _value_cell(ws.cell(row, 4), sys_obj.name)
                                _value_cell(ws.cell(row, 5), kpi.name, bold=True)
                                _value_cell(ws.cell(row, 6), vt.name)
                                _value_cell(ws.cell(row, 7), cell_value, align="right",
                                            number_format=_numeric_format_for_vt(vt))
                                _value_cell(ws.cell(row, 8),
                                            float(cfg.target_value) if cfg.target_value is not None else None,
                                            align="right", number_format=_numeric_format_for_vt(vt))
                                _value_cell(ws.cell(row, 9),
                                            (progress / 100.0) if progress is not None else None,
                                            align="right", number_format="0%", bg=rag_bg, bold=bool(rag))
                                _value_cell(ws.cell(row, 10),
                                            {"green": "🟢 On track", "amber": "🟡 At risk",
                                             "red": "🔴 Off track"}.get(rag, "—"),
                                            align="center", bg=rag_bg, bold=bool(rag))
                                _value_cell(ws.cell(row, 11), cfg.target_direction or "—", align="center")
                                _value_cell(ws.cell(row, 12),
                                            cfg.target_date.strftime("%Y-%m-%d") if cfg.target_date else None,
                                            align="center")
                                _value_cell(ws.cell(row, 13),
                                            int(cfg.target_tolerance_pct) if cfg.target_tolerance_pct is not None else None,
                                            align="right", number_format="0\"%\"")
                                _value_cell(ws.cell(row, 14),
                                            last_contrib.created_at.strftime("%Y-%m-%d") if last_contrib else None,
                                            align="center")
                                _value_cell(ws.cell(row, 15),
                                            getattr(last_contrib, "contributor_name", None) if last_contrib else None)
                                row += 1

        if row > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"

    # ── 4. Action Items ───────────────────────────────────────────────────────
    @staticmethod
    def _build_action_items(wb, ctx):
        ws = wb.create_sheet("Action Items")
        headers = ["Type", "Title", "Description", "Status", "Priority",
                   "Due Date", "Completed", "Owner", "Created By", "Visibility",
                   "Mentions", "Governance Bodies", "Created"]
        for i, h in enumerate(headers, start=1):
            _header_cell(ws.cell(1, i), h, align="center")
            ws.column_dimensions[get_column_letter(i)].width = 16
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["K"].width = 30
        ws.column_dimensions["L"].width = 22
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "C2"

        priority_bg = {
            "urgent": "FECACA", "high": "FED7AA", "medium": "FDE68A", "low": "E5E7EB",
        }
        status_bg = {
            "completed": RAG_GREEN, "active": "DBEAFE", "draft": "F1F5F9", "cancelled": "FEE2E2",
        }

        items = (ActionItem.query.filter_by(organization_id=ctx["org_id"])
                 .order_by(ActionItem.created_at.desc()).all())

        for r, ai in enumerate(items, start=2):
            mentions = ", ".join(
                f"{m.entity_type}:{(m.mention_text or '').lstrip('@')}"
                for m in (ai.mentions or [])
            )
            gbs = ", ".join(gb.name for gb in (ai.governance_bodies or []))
            _value_cell(ws.cell(r, 1), ai.type or "action", align="center")
            _value_cell(ws.cell(r, 2), ai.title, bold=True)
            _value_cell(ws.cell(r, 3), ai.description or "")
            _value_cell(ws.cell(r, 4), ai.status or "", align="center",
                        bg=status_bg.get(ai.status), bold=True)
            _value_cell(ws.cell(r, 5), ai.priority or "", align="center",
                        bg=priority_bg.get(ai.priority), bold=True)
            _value_cell(ws.cell(r, 6),
                        ai.due_date.strftime("%Y-%m-%d") if ai.due_date else None,
                        align="center")
            _value_cell(ws.cell(r, 7),
                        ai.completed_at.strftime("%Y-%m-%d") if ai.completed_at else None,
                        align="center")
            _value_cell(ws.cell(r, 8), ai.owner_user.login if ai.owner_user else "")
            _value_cell(ws.cell(r, 9), ai.created_by_user.login if ai.created_by_user else "")
            _value_cell(ws.cell(r, 10), ai.visibility or "", align="center")
            _value_cell(ws.cell(r, 11), mentions)
            _value_cell(ws.cell(r, 12), gbs)
            _value_cell(ws.cell(r, 13),
                        ai.created_at.strftime("%Y-%m-%d") if ai.created_at else None,
                        align="center")

        if items:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(items) + 1}"

    # ── 5. Settings ───────────────────────────────────────────────────────────
    @staticmethod
    def _build_settings(wb, ctx):
        ws = wb.create_sheet("Settings")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 24
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 60
        ws.column_dimensions["E"].width = 18

        row = 1

        def section(title):
            nonlocal row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            _header_cell(ws.cell(row, 1), title, bg=ACCENT, fg="FFFFFF", size=12)
            ws.row_dimensions[row].height = 22
            row += 1

        def headers(cols):
            nonlocal row
            for i, label in enumerate(cols, start=1):
                _header_cell(ws.cell(row, i), label, align="left")
            row += 1

        # Value Types
        section("VALUE TYPES")
        headers(["Name", "Kind", "Numeric Format", "Unit", "Default Aggregation"])
        for vt in ctx["value_types"]:
            _value_cell(ws.cell(row, 1), vt.name, bold=True)
            _value_cell(ws.cell(row, 2), vt.kind)
            _value_cell(ws.cell(row, 3), vt.numeric_format or "")
            _value_cell(ws.cell(row, 4), vt.unit_label or "")
            _value_cell(ws.cell(row, 5), getattr(vt, "default_aggregation_formula", "") or "")
            row += 1
        row += 1

        # Impact levels
        section("IMPACT LEVELS")
        headers(["Level", "Label", "Icon", "Weight", "Color"])
        levels = ImpactLevel.query.filter_by(organization_id=ctx["org_id"]).order_by(ImpactLevel.level).all()
        for lv in levels:
            _value_cell(ws.cell(row, 1), lv.level, align="center")
            _value_cell(ws.cell(row, 2), lv.label, bold=True)
            _value_cell(ws.cell(row, 3), lv.icon or "", align="center")
            _value_cell(ws.cell(row, 4), float(lv.weight) if lv.weight is not None else None,
                        align="right", number_format="0.00")
            _value_cell(ws.cell(row, 5), lv.color or "")
            if lv.color:
                ws.cell(row, 5).fill = _fill(lv.color.lstrip("#") or "FFFFFF")
            row += 1
        row += 1

        # Strategic pillars
        section("STRATEGIC PILLARS")
        headers(["Name", "Accent", "Icon", "Description", "Order"])
        for p in StrategicPillar.query.filter_by(organization_id=ctx["org_id"]).order_by(StrategicPillar.display_order).all():
            _value_cell(ws.cell(row, 1), p.name, bold=True)
            _value_cell(ws.cell(row, 2), p.accent_color or "")
            if p.accent_color:
                ws.cell(row, 2).fill = _fill(p.accent_color.lstrip("#") or "FFFFFF")
            _value_cell(ws.cell(row, 3), p.bs_icon or "", align="center")
            _value_cell(ws.cell(row, 4), p.description or "")
            _value_cell(ws.cell(row, 5), p.display_order, align="right")
            row += 1
        row += 1

        # Governance bodies
        section("GOVERNANCE BODIES")
        headers(["Name", "Abbreviation", "Color", "Description", "Active"])
        for gb in GovernanceBody.query.filter_by(organization_id=ctx["org_id"]).order_by(GovernanceBody.display_order, GovernanceBody.name).all():
            _value_cell(ws.cell(row, 1), gb.name, bold=True)
            _value_cell(ws.cell(row, 2), gb.abbreviation or "", align="center")
            _value_cell(ws.cell(row, 3), gb.color or "")
            if gb.color:
                ws.cell(row, 3).fill = _fill(gb.color.lstrip("#") or "FFFFFF")
            _value_cell(ws.cell(row, 4), gb.description or "")
            _value_cell(ws.cell(row, 5), "yes" if gb.is_active else "no", align="center")
            row += 1
        row += 1

        # Geography
        section("GEOGRAPHY")
        headers(["Region", "Country", "Site", "Code", "Active"])
        regions = (GeographyRegion.query.filter_by(organization_id=ctx["org_id"])
                   .order_by(GeographyRegion.display_order, GeographyRegion.name).all())
        for region in regions:
            _value_cell(ws.cell(row, 1), region.name, bold=True)
            _value_cell(ws.cell(row, 2), "")
            _value_cell(ws.cell(row, 3), "")
            _value_cell(ws.cell(row, 4), region.code or "", align="center")
            _value_cell(ws.cell(row, 5), "")
            row += 1
            for country in region.countries:
                _value_cell(ws.cell(row, 1), "")
                _value_cell(ws.cell(row, 2), country.name, bold=True)
                _value_cell(ws.cell(row, 3), "")
                _value_cell(ws.cell(row, 4), country.iso_code or country.code or "", align="center")
                _value_cell(ws.cell(row, 5), "")
                row += 1
                for site in country.sites:
                    _value_cell(ws.cell(row, 1), "")
                    _value_cell(ws.cell(row, 2), "")
                    _value_cell(ws.cell(row, 3), site.name, italic=True)
                    _value_cell(ws.cell(row, 4), site.code or "", align="center")
                    _value_cell(ws.cell(row, 5), "yes" if site.is_active else "no", align="center")
                    row += 1
