"""Generate People & Transition Readiness Excel workbook with all templates."""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Colors ──
TEAL = "14B8A6"
PINK = "EC4899"
BLUE = "3B82F6"
GREEN = "10B981"
RED = "EF4444"
ORANGE = "F97316"
PURPLE = "8B5CF6"
GREY = "64748B"
LIGHT_GREY = "F1F5F9"
WHITE = "FFFFFF"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
subheader_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, color=GREY, size=10)
body_font = Font(name="Calibri", size=10)
body_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

# Fit score validation
fit_dv = DataValidation(type="list", formula1='"1 - Significant gap,2 - Partial fit,3 - Strong fit"', allow_blank=True)
fit_dv.prompt = "1=gap, 2=partial, 3=strong"
fit_dv.promptTitle = "Fit Score"

# Readiness validation
ready_dv = DataValidation(type="list", formula1='"1 - 6+ months,2 - 3-6 months,3 - Ready now"', allow_blank=True)
ready_dv.prompt = "1=long, 2=medium, 3=ready"
ready_dv.promptTitle = "Readiness"

# Risk validation
risk_dv = DataValidation(type="list", formula1='"1 - Low,2 - Medium,3 - High"', allow_blank=True)

# Sourcing path validation
sourcing_dv = DataValidation(type="list", formula1='"Develop internally,External hire,Nearshore Lithuania,Outsource,Partner/contractor"', allow_blank=True)

# Yes/No
yesno_dv = DataValidation(type="list", formula1='"Yes,No,Partial,TBD"', allow_blank=True)


def setup_sheet(ws, title, headers, col_widths=None):
    """Set up a sheet with title row and headers."""
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    title_cell.fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Header row
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    return 3  # first data row


def add_example_row(ws, row, values):
    """Add an example row with grey italic font."""
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = Font(name="Calibri", size=10, italic=True, color="94A3B8")
        cell.alignment = body_align
        cell.border = thin_border


# ══════════════════════════════════════════════════════════════════════════
# SHEET 1: Role-to-Person Fit Matrix
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Role-Person Fit"
headers = [
    "Role Family", "Target Role", "Current Person", "Department",
    "Fit Score (1-3)", "Readiness (1-3)", "Transition Risk (1-3)",
    "Sourcing Path", "Development Plan", "Notes",
]
start = setup_sheet(ws1, "Role-to-Person Fit Assessment", headers,
                    [20, 30, 22, 18, 16, 16, 16, 22, 30, 30])
ws1.add_data_validation(fit_dv)
ws1.add_data_validation(ready_dv)
ws1.add_data_validation(risk_dv)
ws1.add_data_validation(sourcing_dv)

# Add fit/readiness/risk validations to columns
fit_dv.add(f"E3:E200")
ready_dv.add(f"F3:F200")
risk_dv.add(f"G3:G200")
sourcing_dv.add(f"H3:H200")

# Role family groups with example data
families = [
    ("Business-facing", [
        ("IT Business Partner - Divisions", "", "", "1 - Significant gap", "1 - 6+ months", "2 - Medium", "External hire", "", "Critical role for business engagement"),
        ("Service Owner - ERP", "", "", "", "", "", "", "", ""),
        ("Demand Manager", "", "", "", "", "", "", "", ""),
    ]),
    ("Application & Architecture", [
        ("Enterprise Architect", "", "", "2 - Partial fit", "2 - 3-6 months", "1 - Low", "Develop internally", "Architecture certification + mentoring", ""),
        ("SAP Solution Architect", "", "", "", "", "", "", "", ""),
        ("Domain Owner - CRM", "", "", "", "", "", "", "", ""),
        ("Data Steward", "", "", "", "", "", "", "", ""),
        ("Integration Lead", "", "", "", "", "", "", "", ""),
    ]),
    ("Infrastructure & Security", [
        ("Infrastructure Platform Lead", "", "", "", "", "", "", "", ""),
        ("Cybersecurity Officer", "", "", "", "", "", "", "", ""),
        ("IAM Lead", "", "", "", "", "", "", "", ""),
        ("Site IT Coordinator", "", "", "", "", "", "", "", ""),
        ("Service Desk Manager", "", "", "", "", "", "", "", ""),
    ]),
    ("Governance & Delivery", [
        ("PMO Lead", "", "", "", "", "", "", "", ""),
        ("Change Manager", "", "", "", "", "", "", "", ""),
        ("Vendor Manager", "", "", "", "", "", "", "", ""),
        ("IT Finance Controller", "", "", "", "", "", "", "", ""),
        ("Portfolio Manager", "", "", "", "", "", "", "", ""),
    ]),
]

row = start
for family, roles in families:
    # Family header
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws1.cell(row=row, column=1, value=family)
    cell.font = Font(name="Calibri", bold=True, color=PURPLE, size=11)
    cell.fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    row += 1
    for role_data in roles:
        ws1.cell(row=row, column=1, value=family)
        for i, v in enumerate(role_data, 2):
            cell = ws1.cell(row=row, column=i, value=v)
            cell.font = body_font if v else Font(name="Calibri", size=10, color="CBD5E1")
            cell.alignment = body_align
            cell.border = thin_border
        row += 1
    row += 1  # gap between families

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2: Capability Gap Register
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Capability Gaps")
headers = [
    "Capability Area", "Gap Description", "Severity (1-3)",
    "Single-Person Dependency?", "Mitigation Plan",
    "Action (Develop/Hire/Outsource)", "Owner", "Target Date",
    "Cost Estimate", "Status",
]
start = setup_sheet(ws2, "Capability Gap Register", headers,
                    [22, 35, 14, 18, 30, 22, 18, 14, 14, 14])

risk_dv2 = DataValidation(type="list", formula1='"1 - Low,2 - Medium,3 - Critical"', allow_blank=True)
ws2.add_data_validation(risk_dv2)
risk_dv2.add("C3:C100")
status_dv = DataValidation(type="list", formula1='"Not started,In progress,Completed,Deferred"', allow_blank=True)
ws2.add_data_validation(status_dv)
status_dv.add("J3:J100")

add_example_row(ws2, start, [
    "Cloud architecture", "No cloud-native skills in current team",
    "3 - Critical", "Yes", "External hire + training program for 2 juniors",
    "External hire", "CIO", "2026-06-30", "CHF 150K/yr", "Not started",
])
add_example_row(ws2, start + 1, [
    "Data governance", "No dedicated data steward role",
    "2 - Medium", "No", "Upskill existing BI analyst + formal training",
    "Develop internally", "IT Leadership", "2026-09-30", "CHF 20K training", "Not started",
])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3: Transition Roadmap
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Transition Roadmap")
headers = [
    "Wave", "Role / Change", "Current State", "Target State",
    "Dependencies", "Start Date", "End Date",
    "Risk Level", "Rollback Plan", "Status",
]
start = setup_sheet(ws3, "Transition Wave Planner", headers,
                    [12, 28, 22, 22, 25, 14, 14, 12, 25, 14])

wave_dv = DataValidation(type="list", formula1='"Wave 1 (Quick wins),Wave 2 (Development),Wave 3 (External hire),Wave 4 (Restructure)"', allow_blank=True)
ws3.add_data_validation(wave_dv)
wave_dv.add("A3:A100")
ws3.add_data_validation(status_dv)

add_example_row(ws3, start, [
    "Wave 1 (Quick wins)", "PMO Lead assignment",
    "No formal PMO", "Dedicated PMO Lead",
    "Budget approval", "2026-05-01", "2026-06-15",
    "1 - Low", "Revert to shared PM model", "Not started",
])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4: Sourcing Decision Register
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Sourcing Decisions")
headers = [
    "Role Family", "Target Role", "Sourcing Path",
    "Rationale", "Cost Estimate (Annual)",
    "JD Ready?", "Budget Approved?",
    "Timeline", "Risk", "Notes",
]
start = setup_sheet(ws4, "Sourcing Decision Register", headers,
                    [20, 28, 22, 30, 18, 12, 14, 14, 12, 25])

sourcing_dv2 = DataValidation(type="list", formula1='"Develop internally,External hire,Nearshore Lithuania,Outsource,Partner/contractor,TBD"', allow_blank=True)
ws4.add_data_validation(sourcing_dv2)
sourcing_dv2.add("C3:C200")
ws4.add_data_validation(yesno_dv)
yesno_dv.add("F3:G200")

add_example_row(ws4, start, [
    "Application & Architecture", "Enterprise Architect",
    "Develop internally", "Current person has 70% fit, needs cloud certification",
    "CHF 15K (training)", "No", "N/A",
    "Q3 2026", "1 - Low", "Strong internal candidate",
])
add_example_row(ws4, start + 1, [
    "Business-facing", "IT Business Partner - Divisions",
    "External hire", "No internal candidate with business relationship skills",
    "CHF 160K", "Partial", "TBD",
    "Q2 2026", "2 - Medium", "Need JD review with HR",
])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5: Lithuania Nearshoring
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Lithuania Nearshoring")
headers = [
    "Role / Function", "Nearshore Viable?",
    "Talent Availability", "Language Fit",
    "Timezone Fit", "Cost Saving (%)",
    "Mgmt Overhead", "Recommendation", "Notes",
]
start = setup_sheet(ws5, "Lithuania (Kaunas) Nearshoring Assessment", headers,
                    [28, 16, 18, 14, 14, 16, 16, 22, 30])

viable_dv = DataValidation(type="list", formula1='"Yes - recommended,Possible - needs validation,No - not viable,TBD"', allow_blank=True)
ws5.add_data_validation(viable_dv)
viable_dv.add("B3:B100")

add_example_row(ws5, start, [
    "Application support (L2/L3)", "Yes - recommended",
    "Good — strong Java/SAP talent pool", "English: good",
    "Same timezone (CET)", "35-40%",
    "Low — existing site", "Recommended for Wave 2",
    "Kaunas already has 3 IT people — can grow team",
])
add_example_row(ws5, start + 1, [
    "Enterprise Architecture", "No - not viable",
    "Limited senior talent", "English: good",
    "Same timezone", "25%",
    "High — needs close collaboration", "Keep local",
    "Too strategic for remote — needs daily face-to-face",
])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 6: Outsourcing Assessment
# ══════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Outsourcing Assessment")
headers = [
    "Service Tower", "Current Volume",
    "Complexity", "Vendor Market",
    "Min Viable Scale", "Cost-Benefit",
    "Risk Level", "Recommendation", "Notes",
]
start = setup_sheet(ws6, "Outsourcing Readiness per Service Tower", headers,
                    [25, 18, 14, 18, 18, 18, 14, 22, 30])

recommend_dv = DataValidation(type="list", formula1='"Outsource now,Defer - insufficient scale,Hybrid model,Keep in-house,TBD"', allow_blank=True)
ws6.add_data_validation(recommend_dv)
recommend_dv.add("H3:H100")

add_example_row(ws6, start, [
    "Service Desk (L1)", "~200 tickets/month",
    "Low", "Many vendors available",
    "500+ tickets/month", "Marginal at current volume",
    "1 - Low", "Defer - insufficient scale",
    "Volume too low for dedicated outsourced team. Revisit after ERP rollout.",
])
add_example_row(ws6, start + 1, [
    "Infrastructure Ops", "15 servers, 8 sites",
    "Medium", "Several regional MSPs",
    "50+ servers", "Possible with managed service",
    "2 - Medium", "Hybrid model",
    "Outsource monitoring + L1, keep L2/L3 in-house.",
])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 7: Communication Plan
# ══════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Communication Plan")
headers = [
    "Person / Group", "Current Role", "Impact",
    "1:1 Date", "1:1 Done?", "Key Message",
    "Development Offered", "HR Co-owner",
    "Follow-up Date", "Status", "Notes",
]
start = setup_sheet(ws7, "Stakeholder Communication & Change Plan", headers,
                    [22, 22, 16, 14, 12, 30, 25, 16, 14, 14, 25])

impact_dv = DataValidation(type="list", formula1='"No change,Role evolution,Role change,New hire (no impact),Potential displacement,TBD"', allow_blank=True)
ws7.add_data_validation(impact_dv)
impact_dv.add("C3:C200")

add_example_row(ws7, start, [
    "Example Person", "IT Manager",
    "Role evolution", "2026-05-15", "No",
    "Your role is evolving to include business partnership responsibilities",
    "Leadership coaching + business analysis training",
    "HR Business Partner", "2026-06-01", "Not started",
    "Handle with care — long tenure, high influence",
])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 8: CIO Summary Dashboard
# ══════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("CIO Summary")
ws8.merge_cells("A1:F1")
cell = ws8.cell(row=1, column=1, value="CIO Gap Analysis Summary — IT Function Target")
cell.font = Font(name="Calibri", bold=True, color=WHITE, size=16)
cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
cell.alignment = Alignment(horizontal="center", vertical="center")
ws8.row_dimensions[1].height = 40

summary_data = [
    ("", ""),
    ("OVERVIEW", ""),
    ("Total target roles", "=COUNTA('Role-Person Fit'!B3:B200)"),
    ("Roles assessed (fit score assigned)", "=COUNTIF('Role-Person Fit'!E3:E200,\"*\")"),
    ("Assessment coverage %", "=IF(B4>0,B5/B4*100,0)"),
    ("", ""),
    ("FIT DISTRIBUTION", ""),
    ("Strong fit (score 3)", "=COUNTIF('Role-Person Fit'!E3:E200,\"3*\")"),
    ("Partial fit (score 2)", "=COUNTIF('Role-Person Fit'!E3:E200,\"2*\")"),
    ("Significant gap (score 1)", "=COUNTIF('Role-Person Fit'!E3:E200,\"1*\")"),
    ("", ""),
    ("SOURCING DECISIONS", ""),
    ("Develop internally", "=COUNTIF('Sourcing Decisions'!C3:C200,\"Develop*\")"),
    ("External hire", "=COUNTIF('Sourcing Decisions'!C3:C200,\"External*\")"),
    ("Nearshore Lithuania", "=COUNTIF('Sourcing Decisions'!C3:C200,\"Nearshore*\")"),
    ("Outsource", "=COUNTIF('Sourcing Decisions'!C3:C200,\"Outsource*\")"),
    ("Partner/contractor", "=COUNTIF('Sourcing Decisions'!C3:C200,\"Partner*\")"),
    ("TBD", "=COUNTIF('Sourcing Decisions'!C3:C200,\"TBD\")"),
    ("", ""),
    ("CAPABILITY GAPS", ""),
    ("Total gaps identified", "=COUNTA('Capability Gaps'!B3:B100)"),
    ("Critical gaps (severity 3)", "=COUNTIF('Capability Gaps'!C3:C100,\"3*\")"),
    ("Single-person dependencies", "=COUNTIF('Capability Gaps'!D3:D100,\"Yes\")"),
    ("Gaps with mitigation plan", "=COUNTIF('Capability Gaps'!E3:E100,\"*\")"),
    ("", ""),
    ("TRANSITION", ""),
    ("Total roles in transition plan", "=COUNTA('Transition Roadmap'!B3:B100)"),
    ("Wave 1 (quick wins)", "=COUNTIF('Transition Roadmap'!A3:A100,\"Wave 1*\")"),
    ("1:1 conversations completed", "=COUNTIF('Communication Plan'!E3:E200,\"Yes\")"),
    ("", ""),
    ("KEY RISKS", ""),
    ("High transition risks", "=COUNTIF('Role-Person Fit'!G3:G200,\"3*\")"),
    ("Outsourcing scale insufficient", "See Outsourcing Assessment sheet"),
]

ws8.column_dimensions["A"].width = 35
ws8.column_dimensions["B"].width = 30

for i, (label, value) in enumerate(summary_data, 2):
    a = ws8.cell(row=i, column=1, value=label)
    b = ws8.cell(row=i, column=2, value=value)
    if label and label == label.upper() and label.strip():
        a.font = Font(name="Calibri", bold=True, color=TEAL, size=12)
    else:
        a.font = Font(name="Calibri", size=11)
    b.font = Font(name="Calibri", size=11)
    b.alignment = Alignment(horizontal="right")

# Save
output_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           "People_Transition_Readiness_Templates.xlsx")
wb.save(output_path)
print(f"Saved: {output_path}")
print("Sheets: Role-Person Fit, Capability Gaps, Transition Roadmap,")
print("        Sourcing Decisions, Lithuania Nearshoring, Outsourcing Assessment,")
print("        Communication Plan, CIO Summary")
